from __future__ import annotations

import importlib.util
import json
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
SCRIPT_PATH = PROJECT_ROOT / "scripts" / "reset_steven_demo_data.py"


def load_module():
    spec = importlib.util.spec_from_file_location("reset_steven_demo_data", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def sample_plan():
    expected_revision = load_module().EXPECTED_REVISION
    return {
        "database": "puiying_steven_demo",
        "revision": expected_revision,
        "binding": {
            "database": "puiying_steven_demo",
            "user": "puiying_steven_demo_app",
            "host": "127.0.0.1",
            "port": 5432,
            "postgres_major": 18,
            "revision": expected_revision,
            "file_root": "",
        },
        "namespace": {"s2_subject": "DEMO-S2-STANDARD-20260717"},
        "object_ids": {"s2": None, "s1": None, "s3": None},
        "row_counts": {"s1_jobs": 0, "s2_jobs": 0, "s3_counts": 0},
        "files": [],
        "actors": {
            "admin": "admin-id",
            "approver": "approver-id",
            "operator": "steven-id",
        },
        "preserved": ["accounts_roles_permissions_sessions"],
    }


def test_plan_hash_is_stable_across_dictionary_order():
    module = load_module()
    plan = sample_plan()
    reordered = dict(reversed(list(plan.items())))
    assert module.calculate_plan_hash(plan) == module.calculate_plan_hash(reordered)


def test_dry_run_is_default_and_explicit_flag_is_supported():
    module = load_module()
    default_arguments = module.parse_arguments([])
    explicit_arguments = module.parse_arguments(["--dry-run"])
    assert default_arguments.apply is False
    assert default_arguments.dry_run is False
    assert explicit_arguments.apply is False
    assert explicit_arguments.dry_run is True


def test_dry_run_summary_exposes_counts_not_object_ids():
    module = load_module()
    plan = sample_plan()
    plan["files"] = [
        {"module": "s1", "storage_key": "generated/tenders/job/v1.docx"},
        {"module": "s2", "storage_key": "generated/quotes/job/v1.xlsx"},
        {"module": "s3", "storage_key": "generated/inventory/job/v1.xlsx"},
    ]
    summary = module.summarize_plan(plan)
    assert summary["file_count"] == 3
    assert summary["file_counts_by_module"] == {"s1": 1, "s2": 1, "s3": 1}
    assert "actors" not in summary
    assert "object_ids" not in summary
    assert "files" not in summary


def test_safe_storage_path_accepts_only_generated_namespaces(tmp_path):
    module = load_module()
    resolved = module.safe_storage_path(tmp_path, "generated/tenders/job/v1.docx")
    assert resolved == (tmp_path / "generated" / "tenders" / "job" / "v1.docx").resolve()
    with pytest.raises(module.ResetBlocked):
        module.safe_storage_path(tmp_path, "../outside.txt")
    with pytest.raises(module.ResetBlocked):
        module.safe_storage_path(tmp_path, "uploads/source.pdf")


def test_validate_plan_uses_the_configured_file_root(tmp_path):
    module = load_module()
    plan = sample_plan()
    plan["binding"]["file_root"] = str(tmp_path.resolve())
    plan["files"] = [{"storage_key": "generated/inventory/count/v1.xlsx"}]
    module.validate_plan(plan, tmp_path)


def test_standard_namespace_constants_are_frozen():
    module = load_module()
    assert module.STANDARD_S2_SUBJECT == "DEMO-S2-STANDARD-20260717"
    assert module.STANDARD_S1_DOCUMENT == "DEMO-S1-STANDARD-20260717"
    assert module.STANDARD_S3_COUNT == "DEMO-S3-STANDARD-20260717"


def test_ranking_totals_use_supplier_id_mapping():
    module = load_module()
    comparison = SimpleNamespace(
        ranking=[
            SimpleNamespace(supplier_id="supplier-c", total=2583),
            SimpleNamespace(supplier_id="supplier-a", total=2605),
        ]
    )
    suppliers = [
        SimpleNamespace(id="supplier-a", supplier_code="SUP-A"),
        SimpleNamespace(id="supplier-c", supplier_code="SUP-C"),
    ]
    assert module.ranking_totals_by_supplier_code(comparison, suppliers) == {
        "SUP-C": "2583",
        "SUP-A": "2605",
    }


def test_standard_quote_import_contains_three_by_five_ground_truth():
    module = load_module()
    candidates = []
    for name in (
        "01_supplier_a_zh_scanned.ground-truth.json",
        "02_supplier_b_mixed_numbers.ground-truth.json",
        "03_supplier_c_bilingual_table.ground-truth.json",
    ):
        path = PROJECT_ROOT / "demo-data" / "steven-d0" / name
        candidates.append(json.loads(path.read_text(encoding="utf-8"))["candidate"])
    workbook = load_workbook(
        filename=__import__("io").BytesIO(module.build_standard_quote_import(candidates)),
        data_only=True,
    )
    assert workbook.sheetnames == ["Items", "Suppliers", "Offers"]
    assert workbook["Items"].max_row == 6
    assert workbook["Suppliers"].max_row == 4
    assert workbook["Offers"].max_row == 16


def test_plan_hash_mismatch_is_rejected():
    module = load_module()
    with pytest.raises(module.ResetBlocked, match="does not match"):
        module.validate_expected_plan_hash("a" * 64, "b" * 64)
    module.validate_expected_plan_hash("A" * 64, "a" * 64)


def test_request_id_context_uses_public_audit_context_api():
    module = load_module()
    from app.core.audit_context import current_request_id

    assert current_request_id() is None
    with module.request_id("demo-reset-test"):
        assert current_request_id() == "demo-reset-test"
    assert current_request_id() is None


def test_apply_requires_confirmation_and_plan_hash(monkeypatch, tmp_path):
    module = load_module()
    monkeypatch.setattr(module, "RUNTIME_STATE", tmp_path / "state.json")
    monkeypatch.setattr(module, "runtime_ports_in_use", lambda: [])
    with pytest.raises(module.ResetBlocked):
        module.validate_apply_arguments(
            SimpleNamespace(apply=True, confirm_local_redacted_demo=False, expected_plan_hash="a" * 64)
        )
    with pytest.raises(module.ResetBlocked):
        module.validate_apply_arguments(
            SimpleNamespace(apply=True, confirm_local_redacted_demo=True, expected_plan_hash="")
        )
    module.validate_apply_arguments(
        SimpleNamespace(apply=True, confirm_local_redacted_demo=True, expected_plan_hash="a" * 64)
    )


def test_apply_refuses_existing_runtime_state(monkeypatch, tmp_path):
    module = load_module()
    state = tmp_path / "state.json"
    state.write_text("{}", encoding="utf-8")
    monkeypatch.setattr(module, "RUNTIME_STATE", state)
    with pytest.raises(module.ResetBlocked, match="stop script"):
        module.validate_apply_arguments(
            SimpleNamespace(apply=True, confirm_local_redacted_demo=True, expected_plan_hash="b" * 64)
        )


def test_apply_refuses_active_runtime_ports(monkeypatch, tmp_path):
    module = load_module()
    monkeypatch.setattr(module, "RUNTIME_STATE", tmp_path / "state.json")
    monkeypatch.setattr(module, "runtime_ports_in_use", lambda: [9000, 15443])
    with pytest.raises(module.ResetBlocked, match="9000,15443"):
        module.validate_apply_arguments(
            SimpleNamespace(apply=True, confirm_local_redacted_demo=True, expected_plan_hash="b" * 64)
        )


def test_runtime_environment_rejects_non_demo_database(monkeypatch):
    module = load_module()
    monkeypatch.setenv("AUTH_MODE", "session")
    monkeypatch.setenv("DEMO_SEED_ENABLED", "false")
    monkeypatch.setenv("OCR_ENABLED", "false")
    monkeypatch.setenv("AI_STRUCTURING_ENABLED", "false")
    with pytest.raises(module.ResetBlocked, match="other than"):
        module.validate_runtime_environment(
            "postgresql+psycopg://puiying_steven_demo_app@127.0.0.1:5432/not_the_demo"
        )


def test_runtime_environment_rejects_live_ai_or_ocr(monkeypatch):
    module = load_module()
    monkeypatch.setenv("AUTH_MODE", "session")
    monkeypatch.setenv("DEMO_SEED_ENABLED", "false")
    monkeypatch.setenv("OCR_ENABLED", "true")
    monkeypatch.setenv("AI_STRUCTURING_ENABLED", "false")
    with pytest.raises(module.ResetBlocked, match="OCR"):
        module.validate_runtime_environment(
            "postgresql+psycopg://puiying_steven_demo_app@127.0.0.1:5432/puiying_steven_demo"
        )


@pytest.mark.parametrize(
    ("database_url", "message"),
    [
        (
            "postgresql+psycopg://other_role@127.0.0.1:5432/puiying_steven_demo",
            "database role",
        ),
        (
            "postgresql+psycopg://puiying_steven_demo_app@localhost:5432/puiying_steven_demo",
            "host",
        ),
        (
            "postgresql+psycopg://puiying_steven_demo_app@127.0.0.1:55432/puiying_steven_demo",
            "port",
        ),
    ],
)
def test_runtime_environment_requires_frozen_database_binding(monkeypatch, database_url, message):
    module = load_module()
    monkeypatch.setenv("APP_ENV", "development")
    monkeypatch.setenv("AUTH_MODE", "session")
    monkeypatch.setenv("DEMO_SEED_ENABLED", "false")
    monkeypatch.setenv("OCR_ENABLED", "false")
    monkeypatch.setenv("AI_STRUCTURING_ENABLED", "false")
    with pytest.raises(module.ResetBlocked, match=message):
        module.validate_runtime_environment(database_url)


def test_runtime_environment_rejects_non_development(monkeypatch):
    module = load_module()
    monkeypatch.setenv("APP_ENV", "staging")
    with pytest.raises(module.ResetBlocked, match="APP_ENV=development"):
        module.validate_runtime_environment(
            "postgresql+psycopg://puiying_steven_demo_app@127.0.0.1:5432/puiying_steven_demo"
        )


def test_file_root_is_frozen_to_project_demo_storage(monkeypatch, tmp_path):
    module = load_module()
    monkeypatch.setenv("FILE_STORAGE_ROOT", str(tmp_path))
    with pytest.raises(module.ResetBlocked, match="fixed project file root"):
        module.validate_file_root()
    monkeypatch.setenv("FILE_STORAGE_ROOT", str(module.DEFAULT_FILE_ROOT))
    assert module.validate_file_root() == module.DEFAULT_FILE_ROOT.resolve()


def test_database_url_is_built_from_pg_components_without_password(monkeypatch):
    module = load_module()
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.setenv("PGHOST", "127.0.0.1")
    monkeypatch.setenv("PGPORT", "5432")
    monkeypatch.setenv("PGUSER", "puiying_steven_demo_app")
    monkeypatch.setenv("PGDATABASE", "puiying_steven_demo")
    assert (
        module.database_url_from_environment()
        == "postgresql+psycopg://puiying_steven_demo_app@127.0.0.1:5432/puiying_steven_demo"
    )


def test_dry_run_arguments_reject_apply_only_flags():
    module = load_module()
    with pytest.raises(module.ResetBlocked):
        module.validate_apply_arguments(
            SimpleNamespace(apply=False, confirm_local_redacted_demo=True, expected_plan_hash="")
        )


def test_file_sha256_streams_file_content(tmp_path):
    module = load_module()
    target = tmp_path / "evidence.bin"
    target.write_bytes(b"steven-demo-evidence")
    assert module.file_sha256(target) == __import__("hashlib").sha256(target.read_bytes()).hexdigest()


def test_reset_journal_is_written_atomically(monkeypatch, tmp_path):
    module = load_module()
    monkeypatch.setattr(module, "JOURNAL_ROOT", tmp_path)
    target = module.write_reset_journal("run-001", {"phase": "staging_verified"})
    assert json.loads(target.read_text(encoding="utf-8")) == {"phase": "staging_verified"}
    assert list(tmp_path.glob(".*.tmp")) == []


def test_transaction_connection_view_does_not_close_or_commit_connection():
    module = load_module()
    connection = object()
    with module.TransactionConnectionView(connection).connect() as yielded:
        assert yielded is connection


def test_quarantined_files_can_be_restored(monkeypatch, tmp_path):
    module = load_module()
    file_root = tmp_path / "files"
    quarantine_root = tmp_path / "quarantine"
    monkeypatch.setattr(module, "QUARANTINE_ROOT", quarantine_root)
    source = file_root / "generated" / "quotes" / "job-1" / "v1.xlsx"
    source.parent.mkdir(parents=True)
    source.write_bytes(b"controlled-demo-file")
    plan = {
        "files": [
            {
                "exists": True,
                "storage_key": "generated/quotes/job-1/v1.xlsx",
            }
        ]
    }
    moved = module.quarantine_files(plan, "run-001", file_root)
    assert not source.exists()
    assert moved[0][1].is_file()
    module.restore_quarantined_files(moved)
    assert source.read_bytes() == b"controlled-demo-file"


def test_switch_staging_namespace_updates_only_demo_rows():
    module = load_module()

    class Result:
        rowcount = 1

    class Connection:
        def __init__(self):
            self.calls = []

        def execute(self, statement, parameters):
            self.calls.append((str(statement), parameters))
            return Result()

    connection = Connection()
    seeded = {
        "s2": {"id": "quote-id"},
        "s1": {"id": "tender-id"},
        "s3": {"id": "inventory-id"},
    }
    module.switch_staging_namespace(connection, seeded)
    assert len(connection.calls) == 3
    assert all("is_demo=true" in statement for statement, _ in connection.calls)
    assert [parameters["object_id"] for _, parameters in connection.calls] == [
        "quote-id",
        "tender-id",
        "inventory-id",
    ]


def test_seed_standard_demo_accepts_staging_names(monkeypatch):
    module = load_module()
    applications = (object(), object(), object())
    monkeypatch.setattr(module, "build_applications", lambda engine, file_root: applications)
    calls = []

    def record(module_name):
        def seed(application, steven_actor, approver_actor, run_id, namespace):
            calls.append((module_name, application, namespace))
            return {"id": f"{module_name}-id"}

        return seed

    monkeypatch.setattr(module, "seed_s2", record("s2"))
    monkeypatch.setattr(module, "seed_s1", record("s1"))
    monkeypatch.setattr(module, "seed_s3", record("s3"))
    result = module.seed_standard_demo(
        object(),
        {"operator": "steven-id", "approver": "approver-id"},
        "run-001",
        Path("."),
        s2_subject="DEMO-S2-STAGING-run-001",
        s1_document_number="DEMO-S1-STAGING-run-001",
        s3_count_number="DEMO-S3-STAGING-run-001",
    )
    assert result == {
        "s2": {"id": "s2-id"},
        "s1": {"id": "s1-id"},
        "s3": {"id": "s3-id"},
    }
    assert [namespace for _, _, namespace in calls] == [
        "DEMO-S2-STAGING-run-001",
        "DEMO-S1-STAGING-run-001",
        "DEMO-S3-STAGING-run-001",
    ]

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import runpy

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from app.core.api_response import ApiError
from app.core.audit import AuditRepository
from app.core.audit_context import reset_request_id, set_request_id
from app.core.config import Settings
from app.main import (
    LazyPostgresInventoryApplication,
    LazyPostgresTenderApplication,
    create_app,
)
from app.modules.accounts.repository import InMemoryAuthRepository
from app.modules.platform.append_only_storage import LocalAppendOnlyFileStorage
from app.modules.steven.inventory_application import StevenInventoryApplicationService
from app.modules.steven.inventory_excel import InventoryExcelRenderer
from app.modules.steven.inventory_import import InventoryImportParser
from app.modules.steven.inventory_repository import InMemoryInventoryRepository
from app.modules.steven.inventory_schemas import (
    InventoryCountCreateRequest,
    InventoryCountLineInput,
    InventoryCountUpdateRequest,
    InventoryItemCreateRequest,
)
from app.modules.steven.inventory_uow import InMemoryInventoryUnitOfWork


def item_payload(**overrides):
    values = {
        "sku": "DEMO-PAPER-A4",
        "item_name": "A4 影印纸（脱敏）",
        "category": "办公耗材",
        "location": "DEMO-STORE-A",
        "book_quantity": 12,
        "safety_stock": 10,
        "target_stock": 30,
        "is_demo": True,
    }
    values.update(overrides)
    return values


def inventory_xlsx(rows, headers=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or [
        "sku",
        "item_name",
        "category",
        "location",
        "book_quantity",
        "safety_stock",
        "target_stock",
    ])
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def make_application(tmp_path):
    repository = InMemoryInventoryRepository()
    audit = AuditRepository()
    renderer = InventoryExcelRenderer()
    application = StevenInventoryApplicationService(
        InMemoryInventoryUnitOfWork(repository, audit),
        renderer,
        LocalAppendOnlyFileStorage(tmp_path, "inventory", "xlsx", renderer.verify),
    )
    return application, repository, audit


def create_item(application, **overrides):
    return application.create_item(
        InventoryItemCreateRequest(**item_payload(**overrides)),
        "steven.test",
    )


def create_count(application, items, *, number="DEMO-S3-COUNT-001", lines=None):
    if lines is None:
        lines = [
            InventoryCountLineInput(
                inventory_item_id=items[0].id,
                counted_quantity=8,
                confirmed_order_quantity=22,
            ),
            InventoryCountLineInput(
                inventory_item_id=items[1].id,
                counted_quantity=25,
                confirmed_order_quantity=0,
            ),
        ]
    return application.create_count(
        InventoryCountCreateRequest(
            count_number=number,
            count_date=date(2026, 7, 17),
            lines=lines,
            is_demo=True,
        ),
        "steven.test",
    )


def make_items(application):
    return [
        create_item(application),
        create_item(
            application,
            sku="DEMO-PEN-BLUE",
            item_name="蓝色原子笔（脱敏）",
            location="DEMO-STORE-B",
            book_quantity=20,
            safety_stock=5,
            target_stock=20,
        ),
    ]


@pytest.mark.parametrize(
    "sku",
    ["demo-paper-a4", " ＤＥＭＯ－ＰＡＰＥＲ－Ａ４ ", "  DEMO-PAPER-A4  "],
)
def test_sku_nfkc_casefold_whitespace_uniqueness(tmp_path, sku):
    application, _, _ = make_application(tmp_path)
    create_item(application)
    with pytest.raises(ApiError) as captured:
        create_item(application, sku=sku)
    assert captured.value.code == "duplicate_sku"


@pytest.mark.parametrize(
    "field,value",
    [
        ("book_quantity", -1),
        ("safety_stock", -1),
        ("target_stock", -1),
        ("book_quantity", 1.5),
        ("book_quantity", "12"),
    ],
)
def test_inventory_quantities_require_non_negative_strict_integers(field, value):
    with pytest.raises(ValidationError):
        InventoryItemCreateRequest(**item_payload(**{field: value}))


def test_target_stock_cannot_be_below_safety_stock(tmp_path):
    application, _, _ = make_application(tmp_path)
    with pytest.raises(ApiError) as captured:
        create_item(application, safety_stock=10, target_stock=9)
    assert captured.value.code == "target_below_safety_stock"


def test_count_calculations_and_manual_reason_rule(tmp_path):
    application, _, _ = make_application(tmp_path)
    items = make_items(application)
    with pytest.raises(ApiError) as captured:
        create_count(
            application,
            items,
            lines=[
                InventoryCountLineInput(
                    inventory_item_id=items[0].id,
                    counted_quantity=8,
                    confirmed_order_quantity=21,
                )
            ],
        )
    assert captured.value.code == "manual_reason_required"

    count = create_count(application, items)
    first, second = count.lines
    assert first.difference_quantity == -4
    assert first.is_low_stock is True
    assert first.suggested_order_quantity == 22
    assert first.confirmed_order_quantity == 22
    assert second.difference_quantity == 5
    assert second.is_low_stock is False
    assert second.suggested_order_quantity == 0


def test_duplicate_count_item_and_number_are_blocked_atomically(tmp_path):
    application, repository, _ = make_application(tmp_path)
    items = make_items(application)
    duplicate_lines = [
        InventoryCountLineInput(inventory_item_id=items[0].id, counted_quantity=8),
        InventoryCountLineInput(inventory_item_id=items[0].id, counted_quantity=9),
    ]
    with pytest.raises(ApiError) as captured:
        create_count(application, items, lines=duplicate_lines)
    assert captured.value.code == "duplicate_count_item"
    assert repository.counts == {}
    assert repository.lines == {}

    create_count(application, items)
    with pytest.raises(ApiError) as duplicate_number:
        create_count(application, items)
    assert duplicate_number.value.code == "duplicate_count_number"


def test_submit_self_approval_return_and_editing_state_machine(tmp_path):
    application, _, _ = make_application(tmp_path)
    items = make_items(application)
    count = create_count(application, items)
    submitted = application.submit(count.id, "dual-role.test")
    assert submitted.status == "submitted"
    with pytest.raises(ApiError) as self_approval:
        application.approve(count.id, "dual-role.test", "不应允许自审")
    assert self_approval.value.status_code == 403
    assert self_approval.value.code == "self_approval_forbidden"
    with pytest.raises(ApiError) as edit_submitted:
        application.update_count(
            count.id,
            InventoryCountUpdateRequest(count_date=date(2026, 7, 18)),
            "steven.test",
        )
    assert edit_submitted.value.code == "inventory_count_not_editable"
    with pytest.raises(ApiError) as empty_return:
        application.return_for_revision(count.id, "approver.test", " ")
    assert empty_return.value.code == "return_opinion_required"
    returned = application.return_for_revision(count.id, "approver.test", "请复核差异")
    assert returned.status == "returned"
    updated = application.update_count(
        count.id,
        InventoryCountUpdateRequest(count_date=date(2026, 7, 18)),
        "steven.test",
    )
    assert updated.status == "draft"
    assert updated.submitted_by is None
    assert updated.decided_by is None


@pytest.mark.parametrize("status_name", ["draft", "submitted", "returned"])
def test_non_approved_counts_cannot_export(tmp_path, status_name):
    application, _, _ = make_application(tmp_path)
    items = make_items(application)
    count = create_count(application, items, number=f"DEMO-S3-{status_name}")
    if status_name in {"submitted", "returned"}:
        application.submit(count.id, "steven.test")
    if status_name == "returned":
        application.return_for_revision(count.id, "approver.test", "退回测试")
    with pytest.raises(ApiError) as captured:
        application.export(count.id, "steven.test")
    assert captured.value.code == "formal_export_forbidden"


def test_approved_excel_versions_are_append_only_reopenable_safe_and_audited(tmp_path):
    application, repository, audit = make_application(tmp_path)
    formula_item = create_item(
        application,
        sku="=DEMO-FORMULA",
        item_name="@库存文本",
        location="+DEMO-STORE",
        book_quantity=1,
        safety_stock=1,
        target_stock=2,
    )
    count = create_count(
        application,
        [formula_item],
        lines=[
            InventoryCountLineInput(
                inventory_item_id=formula_item.id,
                counted_quantity=1,
                confirmed_order_quantity=1,
                manual_reason="-人工说明",
                remark="=HYPERLINK(\"https://example.invalid\")",
            )
        ],
    )
    application.submit(count.id, "steven.test")
    application.approve(count.id, "approver.test", "独立审批人已人工复核")

    exports = [application.export(count.id, "steven.test") for _ in range(3)]
    assert [item.version.version_number for item in exports] == [1, 2, 3]
    versions = application.list_versions(count.id)
    assert len({item.storage_key for item in versions}) == 3
    assert len({item.sha256 for item in versions}) == 3

    for version in versions:
        path = application.version_file(count.id, version.version_number)
        InventoryExcelRenderer.verify(path)
        workbook = load_workbook(path, data_only=False)
        try:
            sheet = workbook["Inventory Count"]
            assert str(sheet["A10"].value).startswith("'=")
            assert str(sheet["B10"].value).startswith("'@")
            assert str(sheet["C10"].value).startswith("'+")
            assert str(sheet["L10"].value).startswith("'-")
            assert str(sheet["M10"].value).startswith("'=")
        finally:
            workbook.close()
        assert version.file_id in repository.files
        assert version.size_bytes == path.stat().st_size
        assert version.sha256 is not None and len(version.sha256) == 64

    actions = {event["action"] for event in audit.list_for_object(count.id)}
    assert {
        "inventory.count.create",
        "inventory.count.submit",
        "inventory.count.approve",
        "inventory.export_reserved",
        "inventory.export",
    } <= actions


def test_api_permissions_request_id_and_legacy_security():
    steven = TestClient(
        create_app(
            Settings(
                app_env="test",
                auth_mode="mock",
                demo_seed_enabled=False,
                mock_identity="steven",
            ),
            InMemoryAuthRepository(),
        )
    )
    created_item = steven.post(
        "/api/v1/steven/inventory-items",
        json=item_payload(),
        headers={"X-Request-Id": "s3-item-request-001"},
    )
    assert created_item.status_code == 201
    item_id = created_item.json()["data"]["id"]
    created_count = steven.post(
        "/api/v1/steven/inventory-counts",
        json={
            "count_number": "DEMO-S3-API-001",
            "count_date": "2026-07-17",
            "is_demo": True,
            "lines": [{"inventory_item_id": item_id, "counted_quantity": 8}],
        },
        headers={"X-Request-Id": "s3-count-request-001"},
    )
    assert created_count.status_code == 201
    count_id = created_count.json()["data"]["id"]
    audits = steven.get(f"/api/v1/steven/inventory-counts/{count_id}/audit-events")
    assert audits.status_code == 200
    assert any(
        event["action"] == "inventory.count.create"
        and event["request_id"] == "s3-count-request-001"
        for event in audits.json()["data"]
    )
    assert steven.post(
        f"/api/v1/steven/inventory-counts/{count_id}/approve",
        json={"opinion": "越权"},
    ).status_code == 403
    assert steven.get(
        "/api/v1/steven/inventory-items",
        headers={"X-Role": "admin"},
    ).status_code == 400

    approver = TestClient(
        create_app(
            Settings(
                app_env="test",
                auth_mode="mock",
                demo_seed_enabled=False,
                mock_identity="approver",
            ),
            InMemoryAuthRepository(),
        )
    )
    assert approver.get("/api/v1/steven/inventory-items").status_code == 200
    assert approver.post("/api/v1/steven/inventory-items", json=item_payload()).status_code == 403
    assert approver.get("/api/v1/steven/inventory-items/export").status_code == 403

    admin = TestClient(
        create_app(
            Settings(
                app_env="test",
                auth_mode="mock",
                demo_seed_enabled=False,
                mock_identity="admin",
            ),
            InMemoryAuthRepository(),
        )
    )
    assert admin.get("/api/v1/steven/inventory-items").status_code == 403


def test_export_all_inventory_items_returns_readable_snapshot():
    client = TestClient(
        create_app(
            Settings(
                app_env="test",
                auth_mode="mock",
                demo_seed_enabled=False,
                mock_identity="steven",
            ),
            InMemoryAuthRepository(),
        )
    )
    created = client.post(
        "/api/v1/steven/inventory-items",
        json=item_payload(
            sku="=DEMO-S3-EXPORT",
            item_name="@导出测试品项",
            book_quantity=2,
            safety_stock=5,
            target_stock=12,
        ),
    )
    assert created.status_code == 201

    response = client.get("/api/v1/steven/inventory-items/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "steven-inventory-all-" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    try:
        sheet = workbook["Inventory Items"]
        assert sheet["A6"].value == "SKU"
        assert sheet["A7"].value == "'=DEMO-S3-EXPORT"
        assert sheet["B7"].value == "'@导出测试品项"
        assert sheet["H7"].value == "是"
        assert sheet["I7"].value == 10
        assert isinstance(sheet["L7"].value, datetime)
        assert sheet["L7"].value.microsecond == 0
        assert sheet["L7"].number_format == "yyyy-mm-dd hh:mm:ss"
        assert isinstance(sheet["M7"].value, datetime)
        assert sheet["M7"].value.microsecond == 0
        assert sheet["M7"].number_format == "yyyy-mm-dd hh:mm:ss"
    finally:
        workbook.close()


def test_dual_role_api_self_approval_returns_403():
    client = TestClient(
        create_app(
            Settings(
                app_env="test",
                auth_mode="mock",
                demo_seed_enabled=False,
                mock_identity="dual",
            ),
            InMemoryAuthRepository(),
        )
    )
    item = client.post("/api/v1/steven/inventory-items", json=item_payload()).json()["data"]
    count = client.post(
        "/api/v1/steven/inventory-counts",
        json={
            "count_number": "DEMO-S3-DUAL-001",
            "count_date": "2026-07-17",
            "is_demo": True,
            "lines": [{"inventory_item_id": item["id"], "counted_quantity": 8}],
        },
    ).json()["data"]
    assert client.post(f"/api/v1/steven/inventory-counts/{count['id']}/submit").status_code == 200
    response = client.post(
        f"/api/v1/steven/inventory-counts/{count['id']}/approve",
        json={"opinion": "不应允许自审"},
    )
    assert response.status_code == 403
    assert response.json()["error"]["code"] == "self_approval_forbidden"


def test_inventory_import_parser_reports_row_level_blockers():
    parser = InventoryImportParser()
    parsed = parser.parse(
        "inventory.xlsx",
        inventory_xlsx([
            ["DEMO-S3-IMPORT-A", "导入品项 A（脱敏）", "演示类别", "DEMO-STORE-C", 1, 2, 5],
            [" demo-s3-import-a ", "导入品项 B（脱敏）", "演示类别", "DEMO-STORE-C", 2, 1, 3],
            ["DEMO-S3-IMPORT-C", "=风险文本", "演示类别", "STORE-C", -1, 5, 4],
            ["DEMO-S3-EXISTS", "已存在（脱敏）", "演示类别", "DEMO-STORE-C", 1, 1, 2],
        ]),
        {"demo-s3-exists"},
    )
    assert parsed.valid_count == 1
    assert parsed.invalid_count == 3
    codes = {
        issue["code"]
        for row in parsed.rows
        for issue in row["errors"]
    }
    assert {
        "duplicate_sku_in_file",
        "formula_injection_risk",
        "invalid_demo_location",
        "invalid_non_negative_integer",
        "target_below_safety_stock",
        "duplicate_sku",
    } <= codes


@pytest.mark.parametrize(
    ("filename", "content", "expected_code"),
    [
        ("inventory.csv", b"sku", "unsupported_file"),
        ("inventory.xlsx", b"", "empty_file"),
        ("inventory.xlsx", b"not-an-xlsx", "invalid_xlsx"),
        (
            "inventory.xlsx",
            inventory_xlsx([], headers=["sku", "wrong"]),
            "invalid_headers",
        ),
        ("inventory.xlsx", inventory_xlsx([]), "no_data_rows"),
    ],
)
def test_inventory_import_file_contract(filename, content, expected_code):
    parsed = InventoryImportParser().parse(filename, content, set())
    assert parsed.issues[0]["code"] == expected_code


def test_inventory_import_preflight_confirm_and_request_id(tmp_path):
    application, repository, audit = make_application(tmp_path)
    content = inventory_xlsx([
        ["DEMO-S3-IMPORT-001", "批量导入纸品（脱敏）", "办公耗材", "DEMO-STORE-C", 8, 5, 20],
        ["DEMO-S3-IMPORT-002", "批量导入笔品（脱敏）", "书写用品", "DEMO-STORE-D", 4, 3, 10],
    ])
    token = set_request_id("s3-import-request-001")
    try:
        batch = application.preflight_import("inventory.xlsx", content, "steven.test")
        confirmed = application.confirm_import(batch.id, "steven.test")
    finally:
        reset_request_id(token)
    assert batch.valid_count == 2 and batch.invalid_count == 0
    assert confirmed.status == "confirmed"
    assert len(confirmed.rows) == 2
    assert all(row.status == "confirmed" and row.imported_item_id for row in confirmed.rows)
    assert len(repository.items) == 2
    assert repository.import_batches[batch.id].request_id == "s3-import-request-001"
    actions = {event["action"] for event in audit.list_for_object(batch.id)}
    assert {"inventory.import.preflight", "inventory.import.confirm"} <= actions
    assert all(event["request_id"] == "s3-import-request-001" for event in audit.list_for_object(batch.id))


def test_inventory_import_blocked_batch_cannot_confirm(tmp_path):
    application, repository, _ = make_application(tmp_path)
    batch = application.preflight_import(
        "inventory.xlsx",
        inventory_xlsx([
            ["DEMO-S3-DUP", "品项一（脱敏）", "类别", "DEMO-STORE-C", 1, 1, 2],
            ["demo-s3-dup", "品项二（脱敏）", "类别", "DEMO-STORE-C", 1, 1, 2],
        ]),
        "steven.test",
    )
    with pytest.raises(ApiError) as captured:
        application.confirm_import(batch.id, "steven.test")
    assert captured.value.code == "import_preflight_blocked"
    assert repository.items == {}
    assert repository.import_batches[batch.id].status == "preflight_ready"


def test_inventory_import_confirmation_rolls_back_all_items_on_conflict(tmp_path):
    application, repository, audit = make_application(tmp_path)
    batch = application.preflight_import(
        "inventory.xlsx",
        inventory_xlsx([
            ["DEMO-S3-FIRST", "第一项（脱敏）", "类别", "DEMO-STORE-C", 1, 1, 2],
            ["DEMO-S3-CONFLICT", "冲突项（脱敏）", "类别", "DEMO-STORE-C", 1, 1, 2],
        ]),
        "steven.test",
    )
    create_item(application, sku="DEMO-S3-CONFLICT", item_name="并发创建冲突项（脱敏）")
    before_actions = len(audit.list_for_object(batch.id))
    with pytest.raises(ApiError) as captured:
        application.confirm_import(batch.id, "steven.test")
    assert captured.value.code == "duplicate_sku"
    assert not any(item.normalized_sku == "demo-s3-first" for item in repository.items.values())
    assert repository.import_batches[batch.id].status == "preflight_ready"
    assert all(row.status == "valid" for row in repository.import_rows_for(batch.id))
    assert len(audit.list_for_object(batch.id)) == before_actions


def test_smart_inventory_import_updates_existing_demo_and_creates_new_item_atomically(tmp_path):
    application, repository, audit = make_application(tmp_path)
    existing = create_item(
        application,
        sku="DEMO-S3-SMART-UPDATE",
        item_name="更新前品项（脱敏）",
        book_quantity=1,
        safety_stock=1,
        target_stock=2,
    )
    batch = application.preflight_import(
        "smart-inventory.xlsx",
        inventory_xlsx([
            ["DEMO-S3-SMART-UPDATE", "更新后品项（脱敏）", "更新类别", "DEMO-STORE-B", 12, 4, 30],
            ["DEMO-S3-SMART-NEW", "新增品项（脱敏）", "新增类别", "DEMO-STORE-C", 5, 2, 9],
        ]),
        "steven.test",
        allow_existing_demo_updates=True,
    )
    assert batch.invalid_count == 0
    assert [row.values["_import_action"] for row in batch.rows] == ["update", "create"]

    confirmed = application.confirm_import(batch.id, "steven.test")
    updated = repository.get_item_by_normalized_sku("demo-s3-smart-update")
    created = repository.get_item_by_normalized_sku("demo-s3-smart-new")
    assert confirmed.status == "confirmed"
    assert updated is not None and updated.id == existing.id
    assert (updated.item_name, updated.book_quantity, updated.target_stock) == ("更新后品项（脱敏）", 12, 30)
    assert created is not None and created.item_name == "新增品项（脱敏）"
    event = next(event for event in audit.list_for_object(batch.id) if event["action"] == "inventory.import.confirm")
    assert event["before_after"]["after"]["created_count"] == 1
    assert event["before_after"]["after"]["updated_count"] == 1


def test_smart_inventory_import_does_not_update_non_demo_item(tmp_path):
    application, repository, _ = make_application(tmp_path)
    create_item(
        application,
        sku="FORMAL-S3-LOCKED",
        item_name="非演示品项",
        is_demo=False,
    )
    batch = application.preflight_import(
        "smart-inventory.xlsx",
        inventory_xlsx([
            ["FORMAL-S3-LOCKED", "尝试更新", "类别", "DEMO-STORE-A", 2, 1, 3],
        ]),
        "steven.test",
        allow_existing_demo_updates=True,
    )
    assert batch.invalid_count == 1
    assert batch.rows[0].errors[0].code == "duplicate_sku"
    assert repository.get_item_by_normalized_sku("formal-s3-locked").item_name == "非演示品项"


def test_inventory_import_api_multipart_preflight_and_confirm():
    client = TestClient(
        create_app(
            Settings(
                app_env="test",
                auth_mode="mock",
                demo_seed_enabled=False,
                mock_identity="steven",
            ),
            InMemoryAuthRepository(),
        )
    )
    preflight = client.post(
        "/api/v1/steven/inventory-items/import/preflight",
        files={
            "file": (
                "inventory.xlsx",
                inventory_xlsx([
                    ["DEMO-S3-API-IMPORT", "API 导入品项（脱敏）", "演示类别", "DEMO-STORE-C", 3, 2, 8],
                ]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-Request-Id": "s3-import-api-001"},
    )
    assert preflight.status_code == 200
    batch = preflight.json()["data"]
    assert batch["valid_count"] == 1 and batch["invalid_count"] == 0
    confirmed = client.post(f"/api/v1/steven/inventory-items/import/{batch['id']}/confirm")
    assert confirmed.status_code == 200
    assert confirmed.json()["data"]["status"] == "confirmed"
    items = client.get("/api/v1/steven/inventory-items").json()["data"]
    assert any(item["sku"] == "DEMO-S3-API-IMPORT" for item in items)


def test_0009_migration_appends_inventory_tables_constraints_permissions_and_chain():
    migration = (
        Path(__file__).resolve().parents[1]
        / "alembic"
        / "versions"
        / "20260717_0009_steven_s3_inventory.py"
    ).read_text(encoding="utf-8")
    assert 'revision = "20260717_0009"' in migration
    assert 'down_revision = "20260717_0008"' in migration
    for table in (
        "steven_inventory_items",
        "steven_inventory_counts",
        "steven_inventory_count_lines",
        "steven_inventory_versions",
        "steven_inventory_candidate_links",
    ):
        assert table in migration
    for constraint in (
        "uq_steven_inventory_items_normalized_sku",
        "uq_steven_inventory_count_item",
        "uq_steven_inventory_versions_count_version",
        "ck_steven_inventory_lines_manual_reason",
    ):
        assert constraint in migration
    assert "steven:inventory:approve" in migration
    assert "inventory_sheet_extraction" in migration
    assert "inventory_exception_explanation" in migration


def test_0011_migration_appends_inventory_import_tables_and_chain():
    migration = (
        Path(__file__).resolve().parents[1]
        / "alembic"
        / "versions"
        / "20260717_0011_steven_s3_inventory_import.py"
    ).read_text(encoding="utf-8")
    assert 'revision = "20260717_0011"' in migration
    assert 'down_revision = "20260717_0010"' in migration
    for table in (
        "steven_inventory_import_batches",
        "steven_inventory_import_rows",
    ):
        assert table in migration
    for constraint in (
        "ck_steven_inventory_import_batches_status",
        "ck_steven_inventory_import_batches_count_total",
        "ck_steven_inventory_import_rows_status",
        "uq_steven_inventory_import_rows_batch_row",
    ):
        assert constraint in migration


def test_s3_seed_contains_at_least_twenty_redacted_items_across_three_locations():
    seed_path = Path(__file__).resolve().parents[3] / "scripts" / "seed_s3_demo.py"
    namespace = runpy.run_path(str(seed_path), run_name="s3_seed_contract")
    demo_items = namespace["DEMO_ITEMS"]
    locations = {item["location"] for item in demo_items}

    assert len(demo_items) >= 20
    assert len(locations) >= 3
    assert all(location.startswith("DEMO-STORE-") for location in locations)
    assert all(item["is_demo"] is True for item in demo_items)


def test_lazy_postgres_application_wiring_resolves_s1_and_s3(tmp_path):
    settings = Settings(file_storage_root=str(tmp_path))

    tender = LazyPostgresTenderApplication(settings, object())._resolve()
    inventory = LazyPostgresInventoryApplication(settings, object())._resolve()

    assert tender.__class__.__name__ == "StevenTenderApplicationService"
    assert inventory.__class__.__name__ == "StevenInventoryApplicationService"
    assert isinstance(inventory._parser, InventoryImportParser)

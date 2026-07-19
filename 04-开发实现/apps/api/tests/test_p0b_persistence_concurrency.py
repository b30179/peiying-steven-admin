from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.core.api_response import ApiError
from app.core.audit import AuditRepository
from app.core.config import Settings
from app.main import LazyPostgresQuoteApplication, create_app
from app.modules.accounts.repository import InMemoryAuthRepository
from app.modules.steven.quote_application import StevenQuoteApplicationService
from app.modules.steven.quote_excel import LocalAppendOnlyQuoteStorage, QuoteExcelExporter, QuoteImportParser
from app.modules.steven.quote_repository import StevenQuoteRepository
from app.modules.steven.quote_schemas import QuoteCreateRequest, QuoteRecommendationRequest
from app.modules.steven.quote_uow import InMemoryQuoteUnitOfWork


class PersistentAuditStub:
    def append(self, **values):
        return values

    def list(self):
        return []

    def list_for_object(self, object_id: str):
        del object_id
        return []


def standard_import_file() -> bytes:
    workbook = Workbook()
    items = workbook.active
    items.title = "Items"
    items.append(["item_code", "item", "specification", "qty", "unit"])
    for index in range(1, 6):
        items.append([f"ITEM-{index:03}", f"脱敏品项 {index}", "标准规格", index, "件"])
    suppliers = workbook.create_sheet("Suppliers")
    suppliers.append(["supplier_code", "supplier_name", "currency", "valid_until", "freight", "tax"])
    for index in range(1, 4):
        suppliers.append([f"SUP-{index}", f"供应商 {index}（脱敏）", "HKD", date(2027, 12, 31), 10, 5])
    offers = workbook.create_sheet("Offers")
    offers.append(["supplier_code", "item_code", "unit_price", "remark"])
    for supplier_index in range(1, 4):
        for item_index in range(1, 6):
            offers.append([f"SUP-{supplier_index}", f"ITEM-{item_index:03}", 10 + supplier_index + item_index, "脱敏报价"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_application(tmp_path, *, audit=None, storage=None):
    repository = StevenQuoteRepository(seed_demo=True)
    audit_repository = audit or AuditRepository()
    exporter = QuoteExcelExporter(tmp_path)
    quote_storage = storage or LocalAppendOnlyQuoteStorage(exporter.data_root)
    application = StevenQuoteApplicationService(
        InMemoryQuoteUnitOfWork(repository, audit_repository),
        QuoteImportParser(),
        exporter,
        quote_storage,
    )
    return application, repository, audit_repository


def make_approved_application(tmp_path, *, storage=None):
    application, repository, audit = make_application(tmp_path, storage=storage)
    quote_id = "demo-quote-hkd-2026"
    application.save_recommendation(
        quote_id,
        QuoteRecommendationRequest(
            recommended_supplier_id="demo-supplier-3",
            non_lowest_reason="",
            approval_opinion="",
        ),
        "steven.test",
    )
    application.submit_approval(quote_id, "steven.test")
    application.approve(quote_id, "approver.test", "已人工复核")
    return application, repository, audit, quote_id


class FailOnImportConfirmAudit(AuditRepository):
    def append(self, **values):
        if values["action"] == "quote.import_confirm":
            raise RuntimeError("simulated_audit_failure")
        return super().append(**values)


class FailOnceStorage(LocalAppendOnlyQuoteStorage):
    def __init__(self, data_root):
        super().__init__(data_root)
        self._failed = False

    def publish(self, **values):
        if not self._failed:
            self._failed = True
            raise OSError("simulated_render_failure")
        return super().publish(**values)


def test_import_and_audit_rollback_together(tmp_path):
    application, repository, audit = make_application(tmp_path, audit=FailOnImportConfirmAudit())
    quote = application.create_quote(QuoteCreateRequest(subject="脱敏事务回滚", currency="HKD", is_demo=True), "steven.test")
    preview = application.precheck_import(
        quote_id=quote.id,
        filename="sanitized.xlsx",
        content=standard_import_file(),
        actor="steven.test",
    )

    with pytest.raises(RuntimeError, match="simulated_audit_failure"):
        application.confirm_import(quote.id, preview.batch_id, "steven.test")

    assert repository.items_for(quote.id) == []
    assert repository.suppliers_for(quote.id) == []
    assert repository.offers_for(quote.id) == []
    assert repository.get_import_batch(preview.batch_id).confirmed is False
    assert not any(event["action"] == "quote.import_confirm" for event in audit.list())


def test_import_payload_hash_mismatch_writes_zero_rows(tmp_path):
    application, repository, _ = make_application(tmp_path)
    quote = application.create_quote(QuoteCreateRequest(subject="脱敏 hash 校验", currency="HKD", is_demo=True), "steven.test")
    preview = application.precheck_import(
        quote_id=quote.id,
        filename="sanitized.xlsx",
        content=standard_import_file(),
        actor="steven.test",
    )
    repository.get_import_batch(preview.batch_id).offers[0]["unit_price"] += 1

    with pytest.raises(ApiError) as captured:
        application.confirm_import(quote.id, preview.batch_id, "steven.test")

    assert captured.value.code == "import_payload_hash_mismatch"
    assert repository.items_for(quote.id) == []
    assert repository.suppliers_for(quote.id) == []
    assert repository.offers_for(quote.id) == []


def test_concurrent_import_confirmation_allows_one_success(tmp_path):
    application, repository, _ = make_application(tmp_path)
    quote = application.create_quote(QuoteCreateRequest(subject="脱敏并发导入", currency="HKD", is_demo=True), "steven.test")
    preview = application.precheck_import(
        quote_id=quote.id,
        filename="sanitized.xlsx",
        content=standard_import_file(),
        actor="steven.test",
    )

    def confirm():
        try:
            application.confirm_import(quote.id, preview.batch_id, "steven.test")
            return "success"
        except ApiError as error:
            return error.status_code, error.code

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(lambda _: confirm(), range(2)))

    assert outcomes.count("success") == 1
    assert any(outcome == (409, "import_already_confirmed") for outcome in outcomes)
    assert len(repository.items_for(quote.id)) == 5
    assert len(repository.suppliers_for(quote.id)) == 3
    assert len(repository.offers_for(quote.id)) == 15


def test_concurrent_approval_decision_allows_one_success(tmp_path):
    application, _, _, quote_id = make_approved_application_for_pending(tmp_path)

    def decide(actor):
        try:
            application.approve(quote_id, actor, "并发人工审批")
            return "success"
        except ApiError as error:
            return error.status_code, error.code

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(decide, ["approver.one", "approver.two"]))

    assert outcomes.count("success") == 1
    assert any(outcome == (409, "approval_closed") for outcome in outcomes)
    assert application.get_quote(quote_id).status == "approved"


def make_approved_application_for_pending(tmp_path):
    application, repository, audit = make_application(tmp_path)
    quote_id = "demo-quote-hkd-2026"
    application.save_recommendation(
        quote_id,
        QuoteRecommendationRequest(recommended_supplier_id="demo-supplier-3", non_lowest_reason="", approval_opinion=""),
        "steven.test",
    )
    application.submit_approval(quote_id, "steven.test")
    return application, repository, audit, quote_id


def test_ten_concurrent_exports_have_unique_versions_and_hashes(tmp_path):
    application, _, _, quote_id = make_approved_application(tmp_path)

    with ThreadPoolExecutor(max_workers=10) as executor:
        exports = list(executor.map(lambda _: application.export(quote_id, "steven.test"), range(10)))

    versions = application.list_versions(quote_id)
    assert sorted(result.version.version_number for result in exports) == list(range(1, 11))
    assert [version.version_number for version in versions] == list(range(1, 11))
    assert len({version.storage_key for version in versions}) == 10
    for version in versions:
        assert version.status == "ready"
        path = application.version_file(quote_id, version.version_number)
        load_workbook(path, read_only=True).close()
        assert version.size_bytes == path.stat().st_size
        assert version.sha256 is not None and len(version.sha256) == 64


def test_failed_export_version_is_not_reused(tmp_path):
    storage = FailOnceStorage(tmp_path)
    application, _, _, quote_id = make_approved_application(tmp_path, storage=storage)

    with pytest.raises(ApiError) as captured:
        application.export(quote_id, "steven.test")

    assert captured.value.code == "export_failed"
    successful = application.export(quote_id, "steven.test")
    versions = application.list_versions(quote_id)
    assert [(version.version_number, version.status) for version in versions] == [(1, "failed"), (2, "ready")]
    assert successful.version.version_number == 2
    assert any(event["action"] == "quote.export_failed" for event in application.list_audit_events(quote_id))


def test_published_file_can_be_reconciled_after_registration_failure(tmp_path, monkeypatch):
    application, _, _, quote_id = make_approved_application(tmp_path)
    original_complete = application._complete_export

    def fail_registration(*_):
        raise RuntimeError("simulated_database_registration_failure")

    monkeypatch.setattr(application, "_complete_export", fail_registration)
    with pytest.raises(RuntimeError, match="simulated_database_registration_failure"):
        application.export(quote_id, "steven.test")
    monkeypatch.setattr(application, "_complete_export", original_complete)

    reserved = application.list_versions(quote_id)[0]
    assert reserved.status == "reserved"
    published_path = application._storage.resolve(reserved.storage_key)
    assert published_path.is_file()

    reconciled = application.reconcile_export(quote_id, reserved.id, "admin.test")
    assert reconciled.status == "ready"
    assert reconciled.size_bytes == published_path.stat().st_size
    assert reconciled.sha256 is not None and len(reconciled.sha256) == 64


def test_invalid_published_file_is_failed_without_reusing_version(tmp_path, monkeypatch):
    application, _, _, quote_id = make_approved_application(tmp_path)
    original_complete = application._complete_export

    monkeypatch.setattr(application, "_complete_export", lambda *_: (_ for _ in ()).throw(RuntimeError("registration_failed")))
    with pytest.raises(RuntimeError, match="registration_failed"):
        application.export(quote_id, "steven.test")
    monkeypatch.setattr(application, "_complete_export", original_complete)

    reserved = application.list_versions(quote_id)[0]
    application._storage.resolve(reserved.storage_key).write_bytes(b"not-an-xlsx")
    with pytest.raises(ApiError) as captured:
        application.reconcile_export(quote_id, reserved.id, "admin.test")

    assert captured.value.code == "published_file_invalid"
    assert application.list_versions(quote_id)[0].status == "failed"
    assert any(event["action"] == "quote.export_reconciliation_failed" for event in application.list_audit_events(quote_id))


@pytest.mark.parametrize("app_env", ["staging", "production"])
def test_controlled_environments_do_not_fall_back_to_memory_s2(tmp_path, app_env):
    settings = Settings(
        app_env=app_env,
        auth_mode="session",
        demo_seed_enabled=False,
        database_url="postgresql+psycopg://placeholder.invalid/puiying_steven_controlled",
        file_storage_root=str(tmp_path / app_env),
        allowed_origins=("https://steven.example.invalid",),
        rate_limit_mode="gateway",
        trusted_proxy_cidrs=("10.0.0.0/8",),
    )
    application = create_app(
        settings,
        auth_repository=InMemoryAuthRepository(),
        audit_repository=PersistentAuditStub(),
    )
    assert isinstance(application.state.quote_application, LazyPostgresQuoteApplication)

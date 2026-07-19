from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.core.api_response import ApiError
from app.core.audit import AuditRepository
from app.modules.platform.append_only_storage import LocalAppendOnlyFileStorage
from app.modules.steven.ai_assist_service import (
    QuoteAiRankingEntry,
    QuoteAiRecommendation,
    StevenAiAssistService,
)
from app.modules.steven.inventory_application import StevenInventoryApplicationService
from app.modules.steven.inventory_excel import InventoryExcelRenderer
from app.modules.steven.inventory_repository import InMemoryInventoryRepository
from app.modules.steven.inventory_schemas import InventoryItemCreateRequest
from app.modules.steven.inventory_smart_import import build_standard_inventory_workbook, inspect_inventory_workbook
from app.modules.steven.inventory_uow import InMemoryInventoryUnitOfWork
from app.modules.steven.quote_excel import QuoteExcelExporter, QuoteImportParser
from app.modules.steven.quote_repository import StevenQuoteRepository
from app.modules.steven.quote_service import StevenQuoteService
from app.modules.steven.tender_application import StevenTenderApplicationService
from app.modules.steven.tender_repository import InMemoryTenderRepository
from app.modules.steven.tender_schemas import (
    TenderCreateRequest,
    TenderTemplateCreateRequest,
    TenderTemplateUpdateRequest,
)
from app.modules.steven.tender_uow import InMemoryTenderUnitOfWork
from app.modules.steven.tender_word import TenderWordRenderer


def mock_ai() -> StevenAiAssistService:
    return StevenAiAssistService(
        enabled=True,
        provider="mock",
        endpoint="https://example.invalid/v1",
        model="mock-model",
        timeout_seconds=1,
        transport=lambda _: {},
    )


def arbitrary_inventory_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["物资编码", "品名", "库存数", "库位", "安全量", "备注"])
    sheet.append(["DEMO-S3-SMART-001", "演示清洁布", 12, "DEMO-STORE-A", 3, "脱敏样本"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_inventory_application(tmp_path):
    repository = InMemoryInventoryRepository()
    audit = AuditRepository()
    renderer = InventoryExcelRenderer()
    application = StevenInventoryApplicationService(
        InMemoryInventoryUnitOfWork(repository, audit),
        renderer,
        LocalAppendOnlyFileStorage(tmp_path, "inventory", "xlsx", renderer.verify),
    )
    return application, repository, audit


def make_tender_application(tmp_path):
    repository = InMemoryTenderRepository(seed_demo=True)
    audit = AuditRepository()
    renderer = TenderWordRenderer()
    application = StevenTenderApplicationService(
        InMemoryTenderUnitOfWork(repository, audit),
        renderer,
        LocalAppendOnlyFileStorage(tmp_path, "tenders", "docx", renderer.verify),
    )
    return application, repository, audit


def test_mock_ai_inventory_mapping_and_quick_entry_are_reviewable():
    service = mock_ai()
    mapping = service.smart_inventory_mapping(
        ["物资编码", "品名", "库存数", "库位", "安全量", "未使用列"],
        [{"物资编码": "DEMO-S3-001", "品名": "演示品项", "库存数": 5, "库位": "DEMO-STORE-A"}],
    )
    assert mapping.mapping == {
        "物资编码": "item_code",
        "品名": "item",
        "库存数": "qty",
        "库位": "location",
        "安全量": "safety_stock",
    }
    assert mapping.unmapped_columns == ["未使用列"]

    parsed = service.quick_inventory_entry("演示清洁布 12包，库位 DEMO-STORE-A")
    assert parsed.items[0].item == "演示清洁布"
    assert parsed.items[0].qty == 12
    assert parsed.items[0].location == "DEMO-STORE-A"

    parsed_without_location = service.quick_inventory_entry("A4纸 20包")
    assert parsed_without_location.items[0].item == "A4纸"
    assert parsed_without_location.items[0].qty == 20
    assert parsed_without_location.items[0].unit == "包"
    assert parsed_without_location.items[0].location is None


def test_smart_inventory_mapping_builds_existing_fixed_import_contract():
    content = arbitrary_inventory_xlsx()
    converted = build_standard_inventory_workbook(
        content,
        {
            "物资编码": "item_code",
            "品名": "item",
            "库存数": "qty",
            "库位": "location",
            "安全量": "safety_stock",
            "备注": "remark",
        },
    )
    workbook = load_workbook(BytesIO(converted), data_only=True)
    try:
        values = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert values[1] == (
        "DEMO-S3-SMART-001",
        "演示清洁布",
        "脱敏样本",
        "DEMO-STORE-A",
        12,
        3,
        12,
    )


def test_smart_inventory_mapping_detects_export_header_after_metadata_and_preserves_target_stock():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["数据标识", "当前库存信息（本机脱敏演示）"])
    sheet.append(["生成时间", "2026-07-18 18:08"])
    sheet.append(["库存品项数", 1])
    sheet.append(["说明", "仅用于脱敏演示"])
    sheet.append([])
    sheet.append(["SKU", "品项名称", "类别", "库位", "账面数量", "安全库存", "目标库存", "低库存", "创建时间"])
    sheet.append(["DEMO-S3-SMART-EXPORT", "演示纸品", "纸品", "DEMO-STORE-A", 8, 3, 20, "是", "2026-07-18"])
    buffer = BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()

    headers, samples = inspect_inventory_workbook("inventory-export.xlsx", content)
    assert headers[:7] == ["SKU", "品项名称", "类别", "库位", "账面数量", "安全库存", "目标库存"]
    assert samples[0]["SKU"] == "DEMO-S3-SMART-EXPORT"

    converted = build_standard_inventory_workbook(
        content,
        {
            "SKU": "item_code",
            "品项名称": "item",
            "类别": "specification",
            "库位": "location",
            "账面数量": "qty",
            "安全库存": "safety_stock",
            "目标库存": "target_stock",
        },
    )
    converted_workbook = load_workbook(BytesIO(converted), data_only=True)
    try:
        row = list(converted_workbook.active.iter_rows(values_only=True))[1]
    finally:
        converted_workbook.close()
    assert row == ("DEMO-S3-SMART-EXPORT", "演示纸品", "纸品", "DEMO-STORE-A", 8, 3, 20)


@pytest.mark.parametrize(
    ("mapping", "expected_code"),
    [
        ({"物资编码": "item_code", "品名": "item", "库存数": "qty", "安全量": "qty", "库位": "location"}, "duplicate_system_mapping"),
        ({"品名": "item", "库存数": "qty", "库位": "location"}, "mapping_required_fields"),
    ],
)
def test_smart_inventory_mapping_rejects_unsafe_confirmations(mapping, expected_code):
    with pytest.raises(ApiError) as captured:
        build_standard_inventory_workbook(arbitrary_inventory_xlsx(), mapping)
    assert captured.value.code == expected_code


def test_quick_entry_confirmation_is_atomic_on_duplicate_sku(tmp_path):
    application, repository, audit = make_inventory_application(tmp_path)
    before_items = len(repository.list_items())
    before_audit = len(audit.list())
    payloads = [
        InventoryItemCreateRequest(
            sku="DEMO-S3-QUICK-DUP",
            item_name="演示快速品项甲",
            category="快速录入（待维护）",
            location="DEMO-STORE-A",
            book_quantity=2,
            safety_stock=0,
            target_stock=2,
            is_demo=True,
        ),
        InventoryItemCreateRequest(
            sku=" demo-s3-quick-dup ",
            item_name="演示快速品项乙",
            category="快速录入（待维护）",
            location="DEMO-STORE-B",
            book_quantity=3,
            safety_stock=0,
            target_stock=3,
            is_demo=True,
        ),
    ]
    with pytest.raises(ApiError) as captured:
        application.create_items(payloads, "steven.test")
    assert captured.value.code == "duplicate_sku"
    assert len(repository.list_items()) == before_items
    assert len(audit.list()) == before_audit


def test_quote_ai_recommendation_is_ephemeral_and_maps_current_supplier(tmp_path):
    repository = StevenQuoteRepository()
    audit = AuditRepository()
    service = StevenQuoteService(repository, audit, QuoteImportParser(), QuoteExcelExporter(tmp_path))
    before_quote = service.get_quote("demo-quote-hkd-2026")
    before_audit = len(audit.list())

    result = service.recommend_supplier("demo-quote-hkd-2026", mock_ai())

    assert result["recommendation"] == "文具供应商丙（脱敏）"
    assert result["recommended_supplier_id"] in {supplier.id for supplier in before_quote.suppliers}
    assert service.get_quote("demo-quote-hkd-2026").recommended_supplier_id is None
    assert len(audit.list()) == before_audit


def test_quote_ai_recommendation_rejects_unknown_supplier(tmp_path):
    repository = StevenQuoteRepository()
    audit = AuditRepository()
    service = StevenQuoteService(repository, audit, QuoteImportParser(), QuoteExcelExporter(tmp_path))

    class UnknownSupplierAi:
        @staticmethod
        def recommend_quote(_):
            return QuoteAiRecommendation(
                recommendation="不存在的供应商",
                reason="仅用于负向测试",
                ranking=[QuoteAiRankingEntry(name="不存在的供应商", score=99, pros=[], cons=[])],
            )

    with pytest.raises(ApiError) as captured:
        service.recommend_supplier("demo-quote-hkd-2026", UnknownSupplierAi())
    assert captured.value.code == "ai_supplier_unknown"
    assert service.get_quote("demo-quote-hkd-2026").recommended_supplier_id is None


def test_inquiry_draft_is_traditional_text_only_and_has_no_side_effects():
    result = mock_ai().inquiry_draft("演示供应商甲", ["A4 影印纸", "蓝色原子笔"], "脱敏办公用品询价")
    assert "敬啟者" in result.text
    assert "詢價" in result.text
    assert "演示供应商甲" in result.text


def test_custom_template_crud_and_demo_template_guards(tmp_path):
    application, repository, audit = make_tender_application(tmp_path)
    created = application.create_template(
        TenderTemplateCreateRequest(
            name="自定义脱敏服务通知",
            document_type="custom_service_notice",
            template_body="标题：{{title}}\n事项：{{subject}}\n地点：{{location}}",
            variables=["title", "subject", "location"],
            keywords=["服务", "通知"],
        ),
        "steven.test",
    )
    assert created.is_demo is False
    assert created.version == 1

    updated = application.update_template(
        created.id,
        TenderTemplateUpdateRequest(
            name="自定义脱敏服务通知（修订）",
            document_type="custom_service_notice",
            template_body="标题：{{title}}\n事项：{{subject}}\n地点：{{location}}\n条款：{{controlled_clauses}}",
            variables=["title", "subject", "location", "controlled_clauses"],
            keywords=["服务", "通知", "条款"],
        ),
        "steven.test",
    )
    assert updated.version == 2
    assert "controlled_clauses" in updated.variables

    deleted = application.delete_template(created.id, "steven.test")
    assert deleted["deleted"] is True
    assert repository.get_template(created.id) is None
    assert [event["action"] for event in audit.list()][-3:] == [
        "tender.template.create",
        "tender.template.update",
        "tender.template.delete",
    ]

    demo_template_id = application.list_templates()[0].id
    demo_payload = TenderTemplateUpdateRequest(
        name="禁止修改",
        document_type="blocked",
        template_body="{{title}}",
        variables=["title"],
        keywords=[],
    )
    with pytest.raises(ApiError) as update_error:
        application.update_template(demo_template_id, demo_payload, "steven.test")
    with pytest.raises(ApiError) as delete_error:
        application.delete_template(demo_template_id, "steven.test")
    assert update_error.value.code == "demo_template_immutable"
    assert delete_error.value.code == "demo_template_immutable"


def test_custom_template_variable_contract_and_in_use_delete_guard(tmp_path):
    application, _, _ = make_tender_application(tmp_path)
    with pytest.raises(ApiError) as mismatch:
        application.create_template(
            TenderTemplateCreateRequest(
                name="变量不一致模板",
                document_type="invalid",
                template_body="{{title}} {{subject}}",
                variables=["title"],
                keywords=[],
            ),
            "steven.test",
        )
    assert mismatch.value.code == "template_variables_mismatch"

    created = application.create_template(
        TenderTemplateCreateRequest(
            name="在用自定义模板",
            document_type="custom_in_use",
            template_body="{{title}}\n{{document_number}}\n{{subject}}\n{{generated_date}}\n{{deadline_date}}\n{{budget_min}}\n{{budget_max}}\n{{currency}}\n{{location}}\n{{supplier_list}}\n{{controlled_clauses}}",
            variables=[
                "title", "document_number", "subject", "generated_date", "deadline_date",
                "budget_min", "budget_max", "currency", "location", "supplier_list", "controlled_clauses",
            ],
            keywords=["在用"],
        ),
        "steven.test",
    )
    application.create_tender(
        TenderCreateRequest(
            template_id=created.id,
            title="脱敏文书",
            document_number="DEMO-S1-CUSTOM-IN-USE",
            subject="脱敏事项",
            generated_date="2026-07-18",
            deadline_date="2026-07-21",
            budget_min=0,
            budget_max=100,
            currency="HKD",
            location="演示地点（脱敏）",
            controlled_clauses="仅供脱敏演示。",
            supplier_names=["演示供应商甲"],
            is_demo=True,
        ),
        "steven.test",
    )
    with pytest.raises(ApiError) as captured:
        application.delete_template(created.id, "steven.test")
    assert captured.value.code == "template_in_use"


def test_0016_migration_extends_chain_without_schema_duplication():
    from pathlib import Path

    migration = (
        Path(__file__).resolve().parents[1]
        / "alembic"
        / "versions"
        / "20260718_0016_steven_ai_assist_extensions.py"
    ).read_text(encoding="utf-8")
    assert 'revision = "20260718_0016"' in migration
    assert 'down_revision = "20260718_0015"' in migration
    assert "create_table" not in migration

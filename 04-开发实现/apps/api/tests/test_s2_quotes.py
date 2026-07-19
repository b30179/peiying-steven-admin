from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.api_response import ApiError
from app.core.audit import AuditRepository
from app.main import app
from app.modules.steven.quote_excel import QuoteExcelExporter, QuoteImportParser
from app.modules.steven.quote_repository import StevenQuoteRepository
from app.modules.steven.quote_schemas import QuoteCreateRequest, QuoteRecommendationRequest, QuoteSupplierReuseRequest
from app.modules.steven.quote_service import StevenQuoteService


def workbook_bytes(*, offer_count: int = 15, currencies: tuple[str, str, str] = ("HKD", "HKD", "HKD"), expired: bool = False, quantity: int = 1, unit_price: int = 10, freight: int = 10, tax: int = 5, duplicate_offer: bool = False) -> bytes:
    workbook = Workbook()
    items = workbook.active
    items.title = "Items"
    items.append(["item_code", "item", "specification", "qty", "unit"])
    for index in range(1, 6):
        items.append([f"ITEM-{index:03}", f"脱敏品项 {index}", "标准规格", quantity if index == 1 else index, "件"])
    suppliers = workbook.create_sheet("Suppliers")
    suppliers.append(["supplier_code", "supplier_name", "currency", "valid_until", "freight", "tax"])
    valid_until = date(2026, 7, 15) if expired else date(2026, 8, 31)
    for index, currency in enumerate(currencies, start=1):
        suppliers.append([f"SUP-{index}", f"供应商 {index}（脱敏）", currency, valid_until, freight, tax])
    offers = workbook.create_sheet("Offers")
    offers.append(["supplier_code", "item_code", "unit_price", "remark"])
    rows = [(f"SUP-{supplier}", f"ITEM-{item:03}", unit_price if unit_price < 0 else unit_price + supplier + item, "脱敏报价") for supplier in range(1, 4) for item in range(1, 6)]
    for row in rows[:offer_count]:
        offers.append(row)
    if duplicate_offer:
        offers.append(rows[0])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_service(tmp_path):
    repository = StevenQuoteRepository()
    audit = AuditRepository()
    service = StevenQuoteService(repository, audit, QuoteImportParser(), QuoteExcelExporter(tmp_path))
    return service, repository, audit


def create_empty_quote(service: StevenQuoteService) -> str:
    return service.create_quote(QuoteCreateRequest(subject="脱敏导入验收", currency="HKD", is_demo=True), "steven.test").id


def test_demo_3x5_totals_and_system_ranking(tmp_path):
    service, _, _ = make_service(tmp_path)
    quote = service.get_quote("demo-quote-hkd-2026")
    assert quote.comparison.comparison_allowed is True
    assert quote.comparison.expected_offer_count == 15
    assert quote.comparison.actual_offer_count == 15
    assert quote.recommended_supplier_id is None
    assert [(entry.supplier_name, entry.subtotal, entry.freight, entry.tax, entry.total) for entry in quote.comparison.ranking] == [
        ("文具供应商丙（脱敏）", 2393, 90, 100, 2583),
        ("文具供应商甲（脱敏）", 2405, 120, 80, 2605),
        ("文具供应商乙（脱敏）", 2367, 180, 75, 2622),
    ]


def test_14_offers_and_currency_mismatch_block_comparison(tmp_path):
    service, _, _ = make_service(tmp_path)
    quote_id = create_empty_quote(service)
    preview = service.precheck_import(quote_id=quote_id, filename="missing.xlsx", content=workbook_bytes(offer_count=14), actor="steven.test")
    assert preview.valid is True
    assert any(issue.code == "missing_offer" for issue in preview.issues)
    quote = service.confirm_import(quote_id, preview.batch_id, "steven.test")
    assert quote.comparison.comparison_allowed is False
    assert quote.comparison.actual_offer_count == 14
    assert quote.comparison.ranking == []

    quote_id = create_empty_quote(service)
    preview = service.precheck_import(quote_id=quote_id, filename="currency.xlsx", content=workbook_bytes(currencies=("HKD", "USD", "HKD")), actor="steven.test")
    quote = service.confirm_import(quote_id, preview.batch_id, "steven.test")
    assert quote.comparison.comparison_allowed is False
    assert any("币种" in reason for reason in quote.comparison.blocking_reasons)
    assert quote.comparison.lowest_supplier_id is None


def test_expired_quote_is_prominent_warning_not_auto_selection(tmp_path):
    service, _, _ = make_service(tmp_path)
    quote_id = create_empty_quote(service)
    preview = service.precheck_import(quote_id=quote_id, filename="expired.xlsx", content=workbook_bytes(expired=True), actor="steven.test")
    assert any(issue.code == "expired_quote" for issue in preview.issues)
    quote = service.confirm_import(quote_id, preview.batch_id, "steven.test")
    assert quote.comparison.comparison_allowed is True
    assert len(quote.comparison.warnings) == 3
    assert quote.recommended_supplier_id is None


@pytest.mark.parametrize(
    ("kwargs", "code"),
    [
        ({"quantity": 0}, "quantity_not_positive"),
        ({"unit_price": -1}, "negative_amount"),
        ({"freight": -1}, "negative_amount"),
        ({"tax": -1}, "negative_amount"),
        ({"duplicate_offer": True}, "duplicate_supplier_item"),
    ],
)
def test_import_precheck_blocks_invalid_values(tmp_path, kwargs, code):
    service, _, _ = make_service(tmp_path)
    quote_id = create_empty_quote(service)
    preview = service.precheck_import(quote_id=quote_id, filename="invalid.xlsx", content=workbook_bytes(**kwargs), actor="steven.test")
    assert preview.valid is False
    assert any(issue.code == code and issue.severity == "error" for issue in preview.issues)
    with pytest.raises(ApiError) as error:
        service.confirm_import(quote_id, preview.batch_id, "steven.test")
    assert error.value.code == "import_has_errors"


def test_non_lowest_requires_reason_and_approval_opinion(tmp_path):
    service, _, _ = make_service(tmp_path)
    quote_id = "demo-quote-hkd-2026"
    for payload in [
        QuoteRecommendationRequest(recommended_supplier_id="demo-supplier-1", non_lowest_reason="", approval_opinion="同意"),
        QuoteRecommendationRequest(recommended_supplier_id="demo-supplier-1", non_lowest_reason="交付较快", approval_opinion=""),
    ]:
        with pytest.raises(ApiError) as error:
            service.save_recommendation(quote_id, payload, "steven.test")
        assert error.value.code == "non_lowest_justification_required"
    quote = service.save_recommendation(quote_id, QuoteRecommendationRequest(recommended_supplier_id="demo-supplier-1", non_lowest_reason="交付期满足校务安排", approval_opinion="同意采用并人工复核"), "steven.test")
    assert quote.recommended_supplier_id == "demo-supplier-1"


def test_manual_approval_audit_and_editable_append_only_excel(tmp_path):
    service, _, audit = make_service(tmp_path)
    quote_id = "demo-quote-hkd-2026"
    service.save_recommendation(quote_id, QuoteRecommendationRequest(recommended_supplier_id="demo-supplier-3", non_lowest_reason="", approval_opinion=""), "steven.test")
    service.submit_approval(quote_id, "steven.test")
    service.approve(quote_id, "approver.test", "资料已人工核对，同意导出")
    first = service.export(quote_id, "steven.test")
    second = service.export(quote_id, "steven.test")
    assert first.version.version_number == 1
    assert second.version.version_number == 2
    first_path = service.version_file(quote_id, 1)
    second_path = service.version_file(quote_id, 2)
    assert first_path != second_path and first_path.exists() and second_path.exists()
    workbook = load_workbook(first_path)
    workbook["Summary"]["B2"] = "可编辑复核"
    edited = tmp_path / "edited-proof.xlsx"
    workbook.save(edited)
    assert load_workbook(edited)["Summary"]["B2"].value == "可编辑复核"
    actions = [event["action"] for event in audit.list_for_object(quote_id)]
    assert {"quote.recommend", "quote.submit", "quote.approve", "quote.export"}.issubset(actions)


def test_api_cross_role_and_general_audit_denied():
    from app.core.config import Settings
    from app.main import create_app
    from app.modules.accounts.repository import InMemoryAuthRepository

    admin_client = TestClient(create_app(Settings(app_env="test", auth_mode="mock", demo_seed_enabled=True, mock_identity="admin"), InMemoryAuthRepository()))
    denied = admin_client.get("/api/v1/steven/quotes")
    assert denied.status_code == 403
    steven_client = TestClient(create_app(Settings(app_env="test", auth_mode="mock", demo_seed_enabled=True, mock_identity="steven"), InMemoryAuthRepository()))
    audit_denied = steven_client.get("/api/v1/audit/events")
    assert audit_denied.status_code == 403
    scoped = steven_client.get("/api/v1/steven/quotes/demo-quote-hkd-2026/audit-events")
    assert scoped.status_code == 200


def approve_demo_quote(service: StevenQuoteService) -> None:
    quote_id = "demo-quote-hkd-2026"
    service.save_recommendation(
        quote_id,
        QuoteRecommendationRequest(
            recommended_supplier_id="demo-supplier-3",
            non_lowest_reason="",
            approval_opinion="",
        ),
        "steven.test",
    )
    service.submit_approval(quote_id, "steven.test")
    service.approve(quote_id, "approver.test", "历史供应商搜索验收")


def test_supplier_search_only_uses_approved_history_and_matches_supported_fields(tmp_path):
    service, _, _ = make_service(tmp_path)
    assert service.search_suppliers("影印纸", 20) == []

    approve_demo_quote(service)

    by_item = service.search_suppliers("影印纸", 20)
    assert {item.supplier_code for item in by_item} == {"SUP-A", "SUP-B", "SUP-C"}
    assert all(item.matched_items == ["A4 影印纸"] for item in by_item)

    by_specification = service.search_suppliers("80gsm", 20)
    assert {item.supplier_code for item in by_specification} == {"SUP-A", "SUP-B", "SUP-C"}
    assert all(item.matched_items == ["A4 影印纸"] for item in by_specification)

    by_supplier = service.search_suppliers("供应商甲", 20)
    assert [item.supplier_code for item in by_supplier] == ["SUP-A"]
    assert set(by_supplier[0].matched_items) == {"A4 影印纸", "A4 文件夹", "白板笔", "蓝色原子笔", "订书钉"}
    assert [item.item_code for item in by_supplier[0].items] == ["ITEM-001", "ITEM-002", "ITEM-003", "ITEM-004", "ITEM-005"]
    assert by_supplier[0].items[0].specification == "80gsm，500 张/包"
    assert by_supplier[0].items[0].qty == 20
    assert by_supplier[0].quote_count == 1


def test_reuse_supplier_can_select_history_items_without_copying_prices(tmp_path):
    service, repository, audit = make_service(tmp_path)
    approve_demo_quote(service)
    quote_id = create_empty_quote(service)

    quote = service.reuse_supplier(
        quote_id,
        QuoteSupplierReuseRequest(
            supplier_code="SUP-A",
            supplier_name="文具供应商甲（脱敏）",
            valid_until=date(2026, 8, 31),
            item_codes=["ITEM-001", "ITEM-003"],
        ),
        "steven.test",
    )

    assert [(supplier.supplier_code, supplier.supplier_name) for supplier in quote.suppliers] == [("SUP-A", "文具供应商甲（脱敏）")]
    assert [(item.item_code, item.item, item.qty, item.unit) for item in quote.items] == [
        ("ITEM-001", "A4 影印纸", 20, "包"),
        ("ITEM-003", "订书钉", 50, "盒"),
    ]
    assert repository.offers_for(quote_id) == []
    assert quote.recommended_supplier_id is None
    event = next(event for event in audit.list_for_object(quote_id) if event["action"] == "quote.supplier_reuse")
    assert event["before_after"]["after"]["reused_item_codes"] == ["ITEM-001", "ITEM-003"]
    assert event["before_after"]["after"]["prices_reused"] is False


def test_reuse_supplier_rejects_item_not_in_approved_history(tmp_path):
    service, _, _ = make_service(tmp_path)
    approve_demo_quote(service)
    quote_id = create_empty_quote(service)

    with pytest.raises(ApiError) as error:
        service.reuse_supplier(
            quote_id,
            QuoteSupplierReuseRequest(
                supplier_code="SUP-A",
                supplier_name="文具供应商甲（脱敏）",
                valid_until=date(2026, 8, 31),
                item_codes=["ITEM-NOT-HISTORICAL"],
            ),
            "steven.test",
        )

    assert error.value.status_code == 422
    assert error.value.code == "supplier_history_item_invalid"
    quote = service.get_quote(quote_id)
    assert quote.suppliers == []
    assert quote.items == []


def test_supplier_search_respects_limit(tmp_path):
    service, _, _ = make_service(tmp_path)
    approve_demo_quote(service)

    results = service.search_suppliers("脱敏", 2)

    assert len(results) == 2
    assert all(item.supplier_name.endswith("（脱敏）") for item in results)


def test_supplier_search_api_validates_query_and_permission():
    from app.core.config import Settings
    from app.main import create_app
    from app.modules.accounts.repository import InMemoryAuthRepository

    steven_client = TestClient(
        create_app(
            Settings(app_env="test", auth_mode="mock", demo_seed_enabled=True, mock_identity="steven"),
            InMemoryAuthRepository(),
        )
    )
    response = steven_client.get("/api/v1/steven/quotes/suppliers/search", params={"q": "影印纸", "limit": 20})
    assert response.status_code == 200
    assert response.json()["data"] == []
    assert steven_client.get("/api/v1/steven/quotes/suppliers/search", params={"q": ""}).status_code == 422
    assert steven_client.get("/api/v1/steven/quotes/suppliers/search", params={"q": "纸", "limit": 101}).status_code == 422

    admin_client = TestClient(
        create_app(
            Settings(app_env="test", auth_mode="mock", demo_seed_enabled=True, mock_identity="admin"),
            InMemoryAuthRepository(),
        )
    )
    assert admin_client.get("/api/v1/steven/quotes/suppliers/search", params={"q": "影印纸"}).status_code == 403


def test_unapproved_quote_can_be_deleted_and_audited(tmp_path):
    service, _, audit = make_service(tmp_path)
    quote_id = create_empty_quote(service)

    result = service.delete_quote(quote_id, "steven.test")

    assert result == {"id": quote_id, "deleted": True}
    with pytest.raises(ApiError) as error:
        service.get_quote(quote_id)
    assert error.value.status_code == 404
    assert any(event["action"] == "quote.delete" for event in audit.list_for_object(quote_id))


def test_approved_quote_cannot_be_deleted(tmp_path):
    service, _, _ = make_service(tmp_path)
    approve_demo_quote(service)

    with pytest.raises(ApiError) as error:
        service.delete_quote("demo-quote-hkd-2026", "steven.test")

    assert error.value.status_code == 409
    assert error.value.code == "approved_record_delete_forbidden"

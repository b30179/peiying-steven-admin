from __future__ import annotations

import csv
import hashlib
import io
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ParsedQuoteImport:
    items: list[dict[str, Any]]
    suppliers: list[dict[str, Any]]
    offers: list[dict[str, Any]]
    issues: list[dict[str, Any]]

    @property
    def valid(self) -> bool:
        return not any(issue["severity"] == "error" for issue in self.issues)


@dataclass(frozen=True)
class ExportedQuoteFile:
    path: Path
    filename: str
    sha256: str


@dataclass(frozen=True)
class PublishedQuoteFile:
    path: Path
    storage_key: str
    sha256: str
    size_bytes: int


class QuoteImportParser:
    def parse(self, filename: str, content: bytes, job_currency: str) -> ParsedQuoteImport:
        suffix = Path(filename).suffix.lower()
        if suffix == ".xlsx":
            parsed = self._parse_xlsx(content)
        elif suffix == ".csv":
            parsed = self._parse_csv(content)
        else:
            return ParsedQuoteImport([], [], [], [self._issue("error", "file", 0, "file", "unsupported_file", "仅支持 XLSX 或 CSV。")])
        issues = list(parsed.issues)
        issues.extend(self._cross_validate(parsed.items, parsed.suppliers, parsed.offers, job_currency))
        return ParsedQuoteImport(parsed.items, parsed.suppliers, parsed.offers, issues)

    def _parse_xlsx(self, content: bytes) -> ParsedQuoteImport:
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            return ParsedQuoteImport([], [], [], [self._issue("error", "file", 0, "file", "invalid_xlsx", "无法打开 XLSX 文件。")])
        required = {"Items", "Suppliers", "Offers"}
        missing = sorted(required.difference(workbook.sheetnames))
        if missing:
            return ParsedQuoteImport([], [], [], [self._issue("error", "file", 0, "sheet", "missing_sheet", f"缺少工作表：{', '.join(missing)}")])
        issues: list[dict[str, Any]] = []
        items = self._read_items(workbook["Items"], issues)
        suppliers = self._read_suppliers(workbook["Suppliers"], issues)
        offers = self._read_offers(workbook["Offers"], issues)
        return ParsedQuoteImport(items, suppliers, offers, issues)

    def _parse_csv(self, content: bytes) -> ParsedQuoteImport:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            return ParsedQuoteImport([], [], [], [self._issue("error", "CSV", 0, "file", "invalid_encoding", "CSV 必须使用 UTF-8 编码。")])
        reader = csv.DictReader(io.StringIO(text))
        items_by_code: dict[str, dict[str, Any]] = {}
        suppliers_by_code: dict[str, dict[str, Any]] = {}
        offers: list[dict[str, Any]] = []
        issues: list[dict[str, Any]] = []
        for row_number, row in enumerate(reader, start=2):
            item = self._normalize_item(row, "CSV", row_number, issues)
            supplier = self._normalize_supplier(row, "CSV", row_number, issues)
            offer = self._normalize_offer(row, "CSV", row_number, issues)
            if item:
                existing = items_by_code.get(item["item_code"])
                if existing and existing != item:
                    issues.append(self._issue("error", "CSV", row_number, "item_code", "inconsistent_item", "同一品项编码资料不一致。"))
                items_by_code[item["item_code"]] = item
            if supplier:
                existing = suppliers_by_code.get(supplier["supplier_code"])
                if existing and existing != supplier:
                    issues.append(self._issue("error", "CSV", row_number, "supplier_code", "inconsistent_supplier", "同一供应商编码资料不一致。"))
                suppliers_by_code[supplier["supplier_code"]] = supplier
            if offer:
                offers.append(offer)
        return ParsedQuoteImport(list(items_by_code.values()), list(suppliers_by_code.values()), offers, issues)

    def _read_items(self, sheet, issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows = self._sheet_rows(sheet)
        return [item for row_number, row in rows if (item := self._normalize_item(row, sheet.title, row_number, issues))]

    def _read_suppliers(self, sheet, issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows = self._sheet_rows(sheet)
        return [supplier for row_number, row in rows if (supplier := self._normalize_supplier(row, sheet.title, row_number, issues))]

    def _read_offers(self, sheet, issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows = self._sheet_rows(sheet)
        return [offer for row_number, row in rows if (offer := self._normalize_offer(row, sheet.title, row_number, issues))]

    @staticmethod
    def _sheet_rows(sheet) -> list[tuple[int, dict[str, Any]]]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        result: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(rows[1:], start=2):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            result.append((row_number, {header: value for header, value in zip(headers, values, strict=False)}))
        return result

    def _normalize_item(self, row: dict[str, Any], sheet: str, row_number: int, issues: list[dict[str, Any]]) -> dict[str, Any] | None:
        code = self._text(row.get("item_code"))
        name = self._text(row.get("item") or row.get("item_name"))
        unit = self._text(row.get("unit"))
        qty = self._decimal(row.get("qty") if "qty" in row else row.get("quantity"), sheet, row_number, "qty", issues)
        if not code or not name or not unit:
            issues.append(self._issue("error", sheet, row_number, "item", "missing_item_field", "品项编码、名称和单位为必填。"))
            return None
        if qty is None:
            return None
        if qty <= 0:
            issues.append(self._issue("error", sheet, row_number, "qty", "quantity_not_positive", "数量必须大于 0。"))
        return {"item_code": code, "item": name, "specification": self._text(row.get("specification")), "qty": qty, "unit": unit}

    def _normalize_supplier(self, row: dict[str, Any], sheet: str, row_number: int, issues: list[dict[str, Any]]) -> dict[str, Any] | None:
        code = self._text(row.get("supplier_code"))
        name = self._text(row.get("supplier_name"))
        currency = self._text(row.get("currency")).upper()
        valid_until = self._date(row.get("valid_until"), sheet, row_number, issues)
        freight = self._decimal(row.get("freight"), sheet, row_number, "freight", issues)
        tax = self._decimal(row.get("tax"), sheet, row_number, "tax", issues)
        if not code or not name or not currency:
            issues.append(self._issue("error", sheet, row_number, "supplier", "missing_supplier_field", "供应商编码、名称和币种为必填。"))
            return None
        if valid_until is None or freight is None or tax is None:
            return None
        if freight < 0:
            issues.append(self._issue("error", sheet, row_number, "freight", "negative_amount", "运费不得为负数。"))
        if tax < 0:
            issues.append(self._issue("error", sheet, row_number, "tax", "negative_amount", "税费不得为负数。"))
        return {"supplier_code": code, "supplier_name": name, "currency": currency, "valid_until": valid_until, "freight": freight, "tax": tax}

    def _normalize_offer(self, row: dict[str, Any], sheet: str, row_number: int, issues: list[dict[str, Any]]) -> dict[str, Any] | None:
        supplier_code = self._text(row.get("supplier_code"))
        item_code = self._text(row.get("item_code"))
        unit_price = self._decimal(row.get("unit_price"), sheet, row_number, "unit_price", issues)
        if not supplier_code or not item_code:
            issues.append(self._issue("error", sheet, row_number, "offer", "missing_offer_reference", "报价必须包含供应商编码和品项编码。"))
            return None
        if unit_price is None:
            return None
        if unit_price < 0:
            issues.append(self._issue("error", sheet, row_number, "unit_price", "negative_amount", "单价不得为负数。"))
        return {"supplier_code": supplier_code, "item_code": item_code, "unit_price": unit_price, "remark": self._text(row.get("remark")), "row": row_number, "sheet": sheet}

    def _cross_validate(self, items: list[dict[str, Any]], suppliers: list[dict[str, Any]], offers: list[dict[str, Any]], job_currency: str) -> list[dict[str, Any]]:
        issues: list[dict[str, Any]] = []
        item_codes = [item["item_code"] for item in items]
        supplier_codes = [supplier["supplier_code"] for supplier in suppliers]
        for duplicate in self._duplicates(item_codes):
            issues.append(self._issue("error", "Items", 0, "item_code", "duplicate_item_code", f"重复品项编码：{duplicate}"))
        for duplicate in self._duplicates(supplier_codes):
            issues.append(self._issue("error", "Suppliers", 0, "supplier_code", "duplicate_supplier_code", f"重复供应商编码：{duplicate}"))
        pairs: set[tuple[str, str]] = set()
        for offer in offers:
            pair = (offer["supplier_code"], offer["item_code"])
            if pair in pairs:
                issues.append(self._issue("error", offer.get("sheet", "Offers"), int(offer.get("row", 0)), "supplier_code+item_code", "duplicate_supplier_item", "同一供应商与品项只能有一条报价。"))
            pairs.add(pair)
            if offer["supplier_code"] not in supplier_codes:
                issues.append(self._issue("error", offer.get("sheet", "Offers"), int(offer.get("row", 0)), "supplier_code", "unknown_supplier", "报价引用了不存在的供应商。"))
            if offer["item_code"] not in item_codes:
                issues.append(self._issue("error", offer.get("sheet", "Offers"), int(offer.get("row", 0)), "item_code", "unknown_item", "报价引用了不存在的品项。"))
        expected = len(items) * len(suppliers)
        if len(pairs) < expected:
            issues.append(self._issue("warning", "Offers", 0, "offer_count", "missing_offer", f"报价不完整：应有 {expected} 条，实际 {len(pairs)} 条；将阻断完整比较。"))
        currencies = {supplier["currency"] for supplier in suppliers}
        if currencies and (len(currencies) > 1 or currencies != {job_currency}):
            issues.append(self._issue("warning", "Suppliers", 0, "currency", "currency_mismatch", "供应商币种与采购事项不一致；将阻断比较。"))
        for supplier in suppliers:
            if supplier["valid_until"] < date.today():
                issues.append(self._issue("warning", "Suppliers", 0, "valid_until", "expired_quote", f"{supplier['supplier_name']} 的报价已过期。"))
        return issues

    @staticmethod
    def _duplicates(values: list[str]) -> set[str]:
        seen: set[str] = set()
        duplicates: set[str] = set()
        for value in values:
            key = value.casefold()
            if key in seen:
                duplicates.add(value)
            seen.add(key)
        return duplicates

    @staticmethod
    def _text(value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _decimal(self, value: Any, sheet: str, row: int, field: str, issues: list[dict[str, Any]]) -> Decimal | None:
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            issues.append(self._issue("error", sheet, row, field, "invalid_number", "必须是有效数字。"))
            return None

    def _date(self, value: Any, sheet: str, row: int, issues: list[dict[str, Any]]) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value))
        except (TypeError, ValueError):
            issues.append(self._issue("error", sheet, row, "valid_until", "invalid_date", "报价有效期必须是 YYYY-MM-DD 日期。"))
            return None

    @staticmethod
    def _issue(severity: str, sheet: str, row: int, field: str, code: str, message: str) -> dict[str, Any]:
        return {"severity": severity, "sheet": sheet, "row": row, "field": field, "code": code, "message": message}


class QuoteExcelExporter:
    def __init__(self, data_root: Path | None = None) -> None:
        api_root = Path(__file__).resolve().parents[3]
        self.data_root = data_root or api_root / "data"
        self.output_root = self.data_root / "generated" / "quotes"

    def export(self, *, quote, items, suppliers, offers, comparison, version_number: int) -> ExportedQuoteFile:
        quote_dir = self.output_root / quote.id
        quote_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{date.today():%Y%m%d}_采购比价_{self._safe_segment(quote.subject)}_v{version_number}.xlsx"
        target = quote_dir / filename
        if target.exists():
            raise FileExistsError(f"版本文件已存在，拒绝覆盖：{target}")

        self.render_to(
            target=target,
            quote=quote,
            items=items,
            suppliers=suppliers,
            offers=offers,
            comparison=comparison,
        )
        digest = hashlib.sha256(target.read_bytes()).hexdigest()
        return ExportedQuoteFile(target, filename, digest)

    def render_to(self, *, target: Path, quote, items, suppliers, offers, comparison) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary_rows = [
            ("数据标识", quote.demo_label or "用户录入数据"),
            ("采购事项", quote.subject),
            ("统一币种", quote.currency),
            ("系统比较状态", "允许" if comparison["comparison_allowed"] else "阻断"),
            ("系统最低价供应商", comparison.get("lowest_supplier_name") or "未产生"),
            ("人工推荐供应商", comparison.get("recommended_supplier_name") or "空白"),
            ("非最低价理由", quote.non_lowest_reason or ""),
            ("审批意见", quote.approval_opinion or ""),
            ("说明", "系统只计算和排序，不自动推荐、审批、下单或发信。"),
        ]
        for row in summary_rows:
            summary.append(row)
        summary["A1"].font = Font(bold=True)

        items_sheet = workbook.create_sheet("Items")
        items_sheet.append(["item_code", "item", "specification", "qty", "unit"])
        for item in items:
            items_sheet.append([item.item_code, item.item, item.specification, float(item.qty), item.unit])

        suppliers_sheet = workbook.create_sheet("Suppliers")
        suppliers_sheet.append(["supplier_code", "supplier_name", "currency", "valid_until", "freight", "tax", "subtotal", "total", "expired"])
        for supplier in suppliers:
            suppliers_sheet.append([
                supplier.supplier_code,
                supplier.supplier_name,
                supplier.currency,
                supplier.valid_until.isoformat(),
                float(supplier.freight),
                float(supplier.tax),
                float(supplier.subtotal),
                float(supplier.total),
                "是" if supplier.valid_until < date.today() else "否",
            ])

        offers_sheet = workbook.create_sheet("Offers")
        offers_sheet.append(["supplier_code", "item_code", "unit_price", "line_total", "remark"])
        supplier_map = {supplier.id: supplier.supplier_code for supplier in suppliers}
        item_map = {item.id: item.item_code for item in items}
        for offer in offers:
            offers_sheet.append([
                supplier_map[offer.quote_supplier_id],
                item_map[offer.quote_item_id],
                float(offer.unit_price),
                float(offer.line_total),
                offer.remark,
            ])

        ranking_sheet = workbook.create_sheet("Ranking")
        ranking_sheet.append(["rank", "supplier_name", "subtotal", "freight", "tax", "total", "expired"])
        for entry in comparison["ranking"]:
            ranking_sheet.append([
                entry["rank"], entry["supplier_name"], float(entry["subtotal"]), float(entry["freight"]),
                float(entry["tax"]), float(entry["total"]), "是" if entry["expired"] else "否",
            ])

        for sheet in workbook.worksheets:
            self._style_sheet(sheet)
        workbook.save(target)

    @staticmethod
    def _style_sheet(sheet) -> None:
        header_fill = PatternFill("solid", fgColor="D9E7F5")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 12), 36)
        sheet.freeze_panes = "A2"

    @staticmethod
    def _safe_segment(value: str) -> str:
        cleaned = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "_", value.strip())
        cleaned = re.sub(r"\s+", "", cleaned)
        return cleaned[:24] or "未命名"


class LocalAppendOnlyQuoteStorage:
    def __init__(self, data_root: Path) -> None:
        self.data_root = data_root.resolve()
        self.temp_root = self.data_root / ".tmp" / "quotes"
        self.final_root = self.data_root / "generated" / "quotes"

    def filename_template(self, subject: str) -> str:
        segment = QuoteExcelExporter._safe_segment(subject)
        return f"{date.today():%Y%m%d}_采购比价_{segment}_v{{version}}.xlsx"

    @staticmethod
    def storage_template(quote_id: str, filename_template: str) -> str:
        return f"generated/quotes/{quote_id}/{filename_template}"

    def publish(self, *, quote_id: str, version_number: int, filename: str, render) -> PublishedQuoteFile:
        temporary = self.temp_root / quote_id / f"v{version_number}-{uuid4()}.xlsx"
        target = self.final_root / quote_id / filename
        temporary.parent.mkdir(parents=True, exist_ok=True)
        target.parent.mkdir(parents=True, exist_ok=True)
        try:
            render(temporary)
            load_workbook(temporary, read_only=True).close()
            digest = hashlib.sha256(temporary.read_bytes()).hexdigest()
            size_bytes = temporary.stat().st_size
            if target.exists():
                raise FileExistsError(f"版本文件已存在，拒绝覆盖：{target}")
            os.link(temporary, target)
            temporary.unlink()
        except Exception:
            temporary.unlink(missing_ok=True)
            raise
        storage_key = target.relative_to(self.data_root).as_posix()
        return PublishedQuoteFile(target, storage_key, digest, size_bytes)

    def resolve(self, storage_key: str) -> Path:
        target = (self.data_root / storage_key).resolve()
        if target != self.data_root and self.data_root not in target.parents:
            raise ValueError("storage_key_outside_root")
        return target

    def inspect_existing(self, storage_key: str) -> PublishedQuoteFile:
        target = self.resolve(storage_key)
        if not target.is_file():
            raise FileNotFoundError(target)
        load_workbook(target, read_only=True).close()
        return PublishedQuoteFile(
            path=target,
            storage_key=storage_key,
            sha256=hashlib.sha256(target.read_bytes()).hexdigest(),
            size_bytes=target.stat().st_size,
        )

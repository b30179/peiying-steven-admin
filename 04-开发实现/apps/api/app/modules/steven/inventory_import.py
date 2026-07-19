from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.modules.steven.inventory_service import normalize_sku


INVENTORY_IMPORT_HEADERS = (
    "sku",
    "item_name",
    "category",
    "location",
    "book_quantity",
    "safety_stock",
    "target_stock",
)
MAX_IMPORT_ROWS = 500
DANGEROUS_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True)
class ParsedInventoryImport:
    original_filename: str
    content_sha256: str
    rows: list[dict[str, Any]]
    issues: list[dict[str, Any]]

    @property
    def valid_count(self) -> int:
        return sum(row["status"] == "valid" for row in self.rows)

    @property
    def invalid_count(self) -> int:
        return sum(row["status"] == "invalid" for row in self.rows)


class InventoryImportParser:
    def parse(
        self,
        filename: str,
        content: bytes,
        existing_normalized_skus: set[str],
        *,
        updatable_normalized_skus: set[str] | None = None,
    ) -> ParsedInventoryImport:
        safe_filename = Path(filename).name[:255] or "inventory.xlsx"
        digest = hashlib.sha256(content).hexdigest()
        if Path(safe_filename).suffix.lower() != ".xlsx":
            return ParsedInventoryImport(
                safe_filename,
                digest,
                [],
                [self._issue(0, "file", "unsupported_file", "仅支持 XLSX 文件。")],
            )
        if not content:
            return ParsedInventoryImport(
                safe_filename,
                digest,
                [],
                [self._issue(0, "file", "empty_file", "导入文件不能为空。")],
            )
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
        except Exception:
            return ParsedInventoryImport(
                safe_filename,
                digest,
                [],
                [self._issue(0, "file", "invalid_xlsx", "无法打开 XLSX 文件。")],
            )
        try:
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not values:
            return ParsedInventoryImport(
                safe_filename,
                digest,
                [],
                [self._issue(0, "file", "empty_workbook", "XLSX 不包含可导入内容。")],
            )
        headers = tuple("" if value is None else str(value).strip() for value in values[0])
        if headers != INVENTORY_IMPORT_HEADERS:
            return ParsedInventoryImport(
                safe_filename,
                digest,
                [],
                [
                    self._issue(
                        1,
                        "header",
                        "invalid_headers",
                        "标题必须严格为：" + ", ".join(INVENTORY_IMPORT_HEADERS),
                    )
                ],
            )
        data_rows = [
            (row_number, row)
            for row_number, row in enumerate(values[1:], start=2)
            if any(value is not None and str(value).strip() for value in row)
        ]
        issues: list[dict[str, Any]] = []
        if len(data_rows) > MAX_IMPORT_ROWS:
            issues.append(
                self._issue(
                    0,
                    "file",
                    "too_many_rows",
                    f"单次导入不得超过 {MAX_IMPORT_ROWS} 行。",
                )
            )
            data_rows = data_rows[:MAX_IMPORT_ROWS]
        parsed_rows: list[dict[str, Any]] = []
        seen: set[str] = set()
        for row_number, cells in data_rows:
            raw = {
                header: self._json_value(cells[index] if index < len(cells) else None)
                for index, header in enumerate(INVENTORY_IMPORT_HEADERS)
            }
            row_errors: list[dict[str, Any]] = []
            text_values: dict[str, str] = {}
            for field in ("sku", "item_name", "category", "location"):
                value = "" if raw[field] is None else str(raw[field]).strip()
                if not value:
                    row_errors.append(self._issue(row_number, field, "required", f"{field} 必填。"))
                if value.startswith(DANGEROUS_PREFIXES):
                    row_errors.append(
                        self._issue(
                            row_number,
                            field,
                            "formula_injection_risk",
                            f"{field} 不得以 =、+、- 或 @ 开头。",
                        )
                    )
                text_values[field] = value
            sku, normalized_sku = normalize_sku(text_values["sku"])
            if normalized_sku:
                if normalized_sku in seen:
                    row_errors.append(
                        self._issue(row_number, "sku", "duplicate_sku_in_file", "文件内存在重复 SKU。")
                    )
                if normalized_sku in existing_normalized_skus and normalized_sku not in (updatable_normalized_skus or set()):
                    row_errors.append(
                        self._issue(row_number, "sku", "duplicate_sku", "SKU 已存在于库存。")
                    )
                seen.add(normalized_sku)
            if text_values["location"] and not text_values["location"].startswith("DEMO-STORE-"):
                row_errors.append(
                    self._issue(
                        row_number,
                        "location",
                        "invalid_demo_location",
                        "脱敏 Demo 库位必须使用 DEMO-STORE-* 前缀。",
                    )
                )
            quantities: dict[str, int | None] = {}
            for field in ("book_quantity", "safety_stock", "target_stock"):
                quantities[field] = self._strict_non_negative_int(raw[field])
                if quantities[field] is None:
                    row_errors.append(
                        self._issue(
                            row_number,
                            field,
                            "invalid_non_negative_integer",
                            f"{field} 必须是非负整数。",
                        )
                    )
            if (
                quantities["safety_stock"] is not None
                and quantities["target_stock"] is not None
                and quantities["target_stock"] < quantities["safety_stock"]
            ):
                row_errors.append(
                    self._issue(
                        row_number,
                        "target_stock",
                        "target_below_safety_stock",
                        "目标库存不得低于安全库存。",
                    )
                )
            normalized_values = {
                "sku": sku,
                "item_name": text_values["item_name"],
                "category": text_values["category"],
                "location": text_values["location"],
                **quantities,
                "is_demo": True,
                "_import_action": "update" if normalized_sku in (updatable_normalized_skus or set()) else "create",
            }
            parsed_rows.append(
                {
                    "row_number": row_number,
                    "raw_values": raw,
                    "values": normalized_values,
                    "normalized_sku": normalized_sku or None,
                    "status": "invalid" if row_errors else "valid",
                    "errors": row_errors,
                }
            )
        if not parsed_rows and not issues:
            issues.append(self._issue(0, "file", "no_data_rows", "XLSX 没有可导入的数据行。"))
        return ParsedInventoryImport(safe_filename, digest, parsed_rows, issues)

    @staticmethod
    def _strict_non_negative_int(value: Any) -> int | None:
        if isinstance(value, bool) or not isinstance(value, int) or value < 0:
            return None
        return value

    @staticmethod
    def _json_value(value: Any) -> Any:
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    @staticmethod
    def _issue(row: int, field: str, code: str, message: str) -> dict[str, Any]:
        return {"row": row, "field": field, "code": code, "message": message}

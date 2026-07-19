from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def safe_excel_text(value: str | None) -> str:
    text = value or ""
    if text.startswith(("=", "+", "-", "@", "\t", "\r", "\n")):
        return f"'{text}"
    return text


DEMO_TIMEZONE = timezone(timedelta(hours=8))


def excel_local_datetime(value: datetime) -> datetime:
    localized = value if value.tzinfo is None else value.astimezone(DEMO_TIMEZONE).replace(tzinfo=None)
    return localized.replace(microsecond=0)


class InventoryExcelRenderer:
    def render_items(self, *, items) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventory Items"
        sheet.append(["数据标识", "当前库存信息（本机脱敏演示）"])
        sheet.append(["生成时间", excel_local_datetime(datetime.now(timezone.utc))])
        sheet.append(["库存品项数", len(items)])
        sheet.append(["说明", "低库存与建议补货量按当前账面数量计算，仅供人工复核。"])
        sheet.append([])
        sheet.append(
            [
                "SKU",
                "品项名称",
                "类别",
                "库位",
                "账面数量",
                "安全库存",
                "目标库存",
                "低库存",
                "当前建议补货量",
                "状态",
                "脱敏演示",
                "创建时间",
                "更新时间",
            ]
        )
        for item in items:
            sheet.append(
                [
                    safe_excel_text(item.sku),
                    safe_excel_text(item.item_name),
                    safe_excel_text(item.category),
                    safe_excel_text(item.location),
                    item.book_quantity,
                    item.safety_stock,
                    item.target_stock,
                    "是" if item.book_quantity < item.safety_stock else "否",
                    max(0, item.target_stock - item.book_quantity),
                    safe_excel_text(item.status),
                    "是" if item.is_demo else "否",
                    excel_local_datetime(item.created_at),
                    excel_local_datetime(item.updated_at),
                ]
            )
        header_fill = PatternFill("solid", fgColor="D9E7F5")
        for row in (1, 6):
            for cell in sheet[row]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        sheet["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
        for row_number in range(7, sheet.max_row + 1):
            sheet.cell(row=row_number, column=12).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(row=row_number, column=13).number_format = "yyyy-mm-dd hh:mm:ss"
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 12), 36)
        sheet.freeze_panes = "A7"
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def verify_items(content: bytes) -> None:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            if "Inventory Items" not in workbook.sheetnames:
                raise ValueError("inventory_items_workbook_sheet_missing")
            sheet = workbook["Inventory Items"]
            if sheet["A1"].value != "数据标识" or sheet["B1"].value != "当前库存信息（本机脱敏演示）":
                raise ValueError("inventory_items_workbook_demo_label_missing")
            if sheet["A6"].value != "SKU":
                raise ValueError("inventory_items_workbook_header_missing")
        finally:
            workbook.close()

    def render_to(self, target: Path, *, count, lines, version_number: int) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventory Count"
        sheet.append(["数据标识", "当前使用脱敏演示资料"])
        sheet.append(["盘点单编号", safe_excel_text(count.count_number)])
        sheet.append(["盘点日期", count.count_date.isoformat()])
        sheet.append(["审批状态", count.status])
        sheet.append(["导出版本", version_number])
        sheet.append(["生成时间", datetime.now(timezone.utc).isoformat()])
        sheet.append(["说明", "系统仅计算差异与建议订货量；人工确认不等于自动下单。"])
        sheet.append([])
        sheet.append(
            [
                "SKU",
                "品项",
                "库位",
                "账面量",
                "盘点量",
                "差异",
                "安全库存",
                "目标库存",
                "低库存",
                "系统建议订货量",
                "人工确认补货量",
                "人工理由",
                "备注",
            ]
        )
        for line in lines:
            sheet.append(
                [
                    safe_excel_text(line.sku_snapshot),
                    safe_excel_text(line.item_name_snapshot),
                    safe_excel_text(line.location_snapshot),
                    line.book_quantity_snapshot,
                    line.counted_quantity,
                    line.difference_quantity,
                    line.safety_stock_snapshot,
                    line.target_stock_snapshot,
                    "是" if line.is_low_stock else "否",
                    line.suggested_order_quantity,
                    line.confirmed_order_quantity,
                    safe_excel_text(line.manual_reason),
                    safe_excel_text(line.remark),
                ]
            )
        header_fill = PatternFill("solid", fgColor="D9E7F5")
        for row in (1, 9):
            for cell in sheet[row]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 12), 36)
        sheet.freeze_panes = "A10"
        workbook.save(target)

    @staticmethod
    def verify(path: Path) -> None:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if "Inventory Count" not in workbook.sheetnames:
                raise ValueError("inventory_workbook_sheet_missing")
            sheet = workbook["Inventory Count"]
            if sheet["A1"].value != "数据标识" or sheet["B1"].value != "当前使用脱敏演示资料":
                raise ValueError("inventory_workbook_demo_label_missing")
            if sheet["A9"].value != "SKU":
                raise ValueError("inventory_workbook_header_missing")
        finally:
            workbook.close()

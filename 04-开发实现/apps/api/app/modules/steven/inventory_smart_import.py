from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.api_response import ApiError
from app.modules.steven.inventory_import import INVENTORY_IMPORT_HEADERS, MAX_IMPORT_ROWS


MAX_HEADER_SCAN_ROWS = 20
HEADER_HINTS = (
    "sku", "itemcode", "物资编码", "物資編碼", "品项名称", "品項名稱", "品名",
    "类别", "類別", "规格", "規格", "库位", "庫位", "账面数量", "帳面數量",
    "库存数量", "庫存數量", "安全库存", "安全庫存", "目标库存", "目標庫存",
)


@dataclass(frozen=True)
class WorkbookLayout:
    header_row_index: int
    columns: tuple[tuple[str, int], ...]


def inspect_inventory_workbook(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ApiError(422, "unsupported_file", "智能导入仅支持 XLSX 文件。")
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise ApiError(422, "invalid_xlsx", "无法打开 XLSX 文件。") from error
    try:
        rows = list(workbook.active.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS + 5, values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ApiError(422, "empty_workbook", "XLSX 不包含可导入内容。")
    layout = _locate_inventory_layout(rows)
    headers = [header for header, _ in layout.columns]
    if len(headers) > 80:
        raise ApiError(422, "too_many_columns", "智能导入最多支持 80 列。")
    samples = [
        {
            header: _json_value(row[source_index] if source_index < len(row) else None)
            for header, source_index in layout.columns
        }
        for row in rows[layout.header_row_index + 1:]
        if any(value is not None and str(value).strip() for value in row)
    ][:5]
    return headers, samples


def build_standard_inventory_workbook(content: bytes, mapping: dict[str, str]) -> bytes:
    try:
        source = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise ApiError(422, "invalid_xlsx", "无法打开 XLSX 文件。") from error
    try:
        values = list(source.active.iter_rows(values_only=True))
    finally:
        source.close()
    if not values:
        raise ApiError(422, "empty_workbook", "XLSX 不包含可导入内容。")
    layout = _locate_inventory_layout(values[:MAX_HEADER_SCAN_ROWS])
    indexes = dict(layout.columns)
    missing = set(mapping) - set(indexes)
    if missing:
        raise ApiError(422, "mapping_column_missing", "列映射引用了不存在的列。")
    reverse = {system_field: source_header for source_header, system_field in mapping.items()}
    if len(reverse) != len(mapping):
        raise ApiError(422, "duplicate_system_mapping", "同一系统字段不可重复映射。")
    if "item" not in reverse or "item_code" not in reverse or "qty" not in reverse or "location" not in reverse:
        raise ApiError(422, "mapping_required_fields", "至少必须映射品项、SKU、数量和库位。")

    target = Workbook()
    sheet = target.active
    sheet.title = "inventory_import"
    sheet.append(list(INVENTORY_IMPORT_HEADERS))
    data_rows = [
        row
        for row in values[layout.header_row_index + 1:]
        if any(value is not None and str(value).strip() for value in row)
    ]
    if len(data_rows) > MAX_IMPORT_ROWS:
        raise ApiError(422, "too_many_rows", f"单次导入不得超过 {MAX_IMPORT_ROWS} 行。")
    for row in data_rows:
        def value(field: str) -> Any:
            source_header = reverse.get(field)
            if source_header is None:
                return None
            index = indexes[source_header]
            return row[index] if index < len(row) else None

        quantity = _integer(value("qty"), "qty")
        safety = _integer(value("safety_stock") or 0, "safety_stock")
        target_stock = _integer(value("target_stock"), "target_stock") if "target_stock" in reverse else max(quantity, safety)
        specification = _text(value("specification"))
        unit = _text(value("unit"))
        remark = _text(value("remark"))
        category = specification or unit or remark or "智能导入（待维护）"
        sheet.append([
            _text(value("item_code")), _text(value("item")), category,
            _text(value("location")), quantity, safety, target_stock,
        ])
    output = io.BytesIO()
    target.save(output)
    target.close()
    return output.getvalue()


def _locate_inventory_layout(rows: list[tuple[Any, ...]]) -> WorkbookLayout:
    candidates: list[tuple[int, int, tuple[tuple[str, int], ...]]] = []
    for row_index, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        raw_columns = [
            (str(value).strip(), source_index)
            for source_index, value in enumerate(row)
            if value is not None and str(value).strip()
        ]
        if len(raw_columns) < 4:
            continue
        columns = _deduplicate_columns(raw_columns)
        score = len(columns) + sum(
            6
            for header, _ in columns
            if any(hint in _normalized_header(header) for hint in HEADER_HINTS)
        )
        candidates.append((score, row_index, columns))
    if not candidates:
        raise ApiError(422, "invalid_headers", "未找到有效的库存资料表头。")
    _, header_row_index, columns = max(candidates, key=lambda item: (item[0], -item[1]))
    return WorkbookLayout(header_row_index=header_row_index, columns=columns)


def _deduplicate_columns(columns: list[tuple[str, int]]) -> tuple[tuple[str, int], ...]:
    occurrences: dict[str, int] = {}
    result: list[tuple[str, int]] = []
    for header, source_index in columns:
        occurrences[header] = occurrences.get(header, 0) + 1
        suffix = occurrences[header]
        display_header = header if suffix == 1 else f"{header} ({suffix})"
        result.append((display_header, source_index))
    return tuple(result)


def _normalized_header(value: str) -> str:
    return re.sub(r"[\s_\-／/()（）]+", "", value).casefold()


def _integer(value: Any, field: str) -> int:
    if isinstance(value, bool):
        raise ApiError(422, "invalid_integer", f"{field} 必须是非负整数。")
    try:
        number = int(value)
    except (TypeError, ValueError) as error:
        raise ApiError(422, "invalid_integer", f"{field} 必须是非负整数。") from error
    if number < 0 or str(value).strip() not in {str(number), f"{number}.0"}:
        raise ApiError(422, "invalid_integer", f"{field} 必须是非负整数。")
    return number


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)

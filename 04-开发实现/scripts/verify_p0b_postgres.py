from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import date
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import re
import subprocess
import sys

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import make_url

PROJECT_ROOT = Path(__file__).resolve().parents[1]
API_ROOT = PROJECT_ROOT / "apps" / "api"
ORIGIN = "https://testserver"
EXPECTED_TABLES = {
    "steven_quote_jobs",
    "steven_quote_items",
    "steven_quote_suppliers",
    "steven_quote_offer_lines",
    "steven_quote_import_batches",
    "steven_quote_approvals",
    "steven_quote_versions",
    "steven_quote_audit_events",
    "platform_audit_events",
}
EXPECTED_CONSTRAINTS = {
    "uq_steven_quote_offer_lines_supplier_item",
    "uq_steven_quote_versions_job_version",
    "uq_steven_quote_versions_storage_key",
    "fk_steven_quote_jobs_approval",
    "fk_steven_quote_import_batches_confirmed_by",
    "fk_steven_quote_approvals_submitted_by",
    "fk_steven_quote_approvals_decided_by",
}
EXPECTED_INDEXES = {
    "uq_steven_quote_approvals_one_pending",
    "ix_steven_quote_import_batches_job_status",
    "ix_steven_quote_versions_status_time",
    "ix_steven_quote_audit_object_time",
    "uq_users_username_lower",
    "ix_platform_audit_request_time",
}


def result(step: str, status: str, **details: object) -> None:
    print(json.dumps({"step": step, "status": status, **details}, ensure_ascii=False))


def run_api_command(*arguments: str, expected_codes: set[int] | None = None) -> subprocess.CompletedProcess[str]:
    process = subprocess.run(
        [sys.executable, *arguments],
        cwd=API_ROOT,
        env=os.environ.copy(),
        capture_output=True,
        text=True,
        check=False,
    )
    if process.returncode not in (expected_codes or {0}):
        raise RuntimeError(f"Command failed with exit code {process.returncode}: {' '.join(arguments)}")
    return process


def validate_storage_root(raw_path: str) -> Path:
    root = Path(raw_path).resolve()
    allowed_root = (PROJECT_ROOT / "data" / "p0b-smoke").resolve()
    if root != allowed_root and allowed_root not in root.parents:
        raise RuntimeError(f"P0B_FILE_STORAGE_ROOT must stay under {allowed_root}")
    if root.exists() and any(root.iterdir()):
        raise RuntimeError("P0-B smoke file root must be absent or empty; existing files are never deleted")
    root.mkdir(parents=True, exist_ok=True)
    return root


def prepare_empty_database(admin_url: str, database_url: str) -> str:
    try:
        import psycopg
    except ModuleNotFoundError as error:
        raise RuntimeError("PostgreSQL driver is unavailable; use an approved project environment with requirements.txt installed") from error
    target = make_url(database_url)
    database_name = target.database or ""
    if not re.fullmatch(r"puiying_steven_p0b_smoke[_a-zA-Z0-9-]*", database_name):
        raise RuntimeError("Target database name must start with puiying_steven_p0b_smoke")
    admin = make_url(admin_url)
    if admin.database == database_name:
        raise RuntimeError("Administrator connection must target a different maintenance database")
    connection_kwargs = {
        "host": admin.host,
        "port": admin.port or 5432,
        "dbname": admin.database or "postgres",
        "user": admin.username,
        "password": admin.password,
    }
    with psycopg.connect(**connection_kwargs, autocommit=True) as connection:
        exists = connection.execute("SELECT 1 FROM pg_database WHERE datname=%s", (database_name,)).fetchone()
        if exists:
            if os.getenv("P0B_ALLOW_RECREATE", "").strip().lower() != "true":
                raise RuntimeError("Smoke database already exists; set P0B_ALLOW_RECREATE=true for this dedicated smoke database only")
            connection.execute("SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname=%s AND pid<>pg_backend_pid()", (database_name,))
            connection.execute(f'DROP DATABASE "{database_name}"')
        connection.execute(f'CREATE DATABASE "{database_name}"')
    result("create_empty_database", "passed", database=database_name)
    return database_name


def migration_and_schema_smoke(database_url: str) -> None:
    upgrade = run_api_command("-m", "alembic", "upgrade", "head")
    result("alembic_upgrade_head", "passed", output=upgrade.stdout.strip())
    current = run_api_command("-m", "alembic", "current")
    if "20260716_0006" not in current.stdout:
        raise RuntimeError("alembic current did not report 20260716_0006")

    engine = create_engine(database_url, future=True)
    schema = inspect(engine)
    missing_tables = sorted(EXPECTED_TABLES - set(schema.get_table_names()))
    if missing_tables:
        raise RuntimeError(f"Missing S2 tables: {missing_tables}")
    with engine.connect() as connection:
        constraints = set(connection.execute(text("SELECT conname FROM pg_constraint")).scalars())
        indexes = set(connection.execute(text("SELECT indexname FROM pg_indexes WHERE schemaname=current_schema()" )).scalars())
        pending_index = connection.execute(text("SELECT indexdef FROM pg_indexes WHERE indexname='uq_steven_quote_approvals_one_pending'" )).scalar_one_or_none()
        actor_constraints = [
            dict(row)
            for row in connection.execute(text("""
                SELECT conname,convalidated
                  FROM pg_constraint
                 WHERE conname IN (
                       'fk_steven_quote_approvals_submitted_by',
                       'fk_steven_quote_approvals_decided_by'
                 )
                 ORDER BY conname
            """)).mappings()
        ]
    missing_constraints = sorted(EXPECTED_CONSTRAINTS - constraints)
    missing_indexes = sorted(EXPECTED_INDEXES - indexes)
    if missing_constraints or missing_indexes or not pending_index or "WHERE" not in pending_index.upper():
        raise RuntimeError(f"Schema contract mismatch: constraints={missing_constraints}, indexes={missing_indexes}, pending_partial={bool(pending_index)}")
    result("schema_constraints", "passed", tables=len(EXPECTED_TABLES), revision="20260716_0006", actor_constraints=actor_constraints)


def bootstrap_smoke(password: str) -> None:
    first = run_api_command("-m", "app.cli", "bootstrap-admin")
    combined = first.stdout + first.stderr
    if password in combined or '"status": "created"' not in first.stdout:
        raise RuntimeError("Bootstrap output contract failed")
    duplicate = run_api_command("-m", "app.cli", "bootstrap-admin", expected_codes={3})
    duplicate_output = duplicate.stdout + duplicate.stderr
    if password in duplicate_output or "admin_already_exists" not in duplicate.stderr:
        raise RuntimeError("Duplicate bootstrap rejection contract failed")
    result("bootstrap_admin", "passed", duplicate_rejected=True, password_echoed=False)


def username_case_conflict_smoke(database_url: str) -> None:
    engine = create_engine(database_url, future=True)
    with engine.begin() as connection:
        role_id = connection.execute(text("SELECT id FROM roles WHERE code='operator'")).scalar_one()
        first_id = os.urandom(16).hex()
        connection.execute(text("INSERT INTO users (id,username,display_name,password_hash,status,failed_login_count,created_at,updated_at) VALUES (:id,'P0B1.CaseUser','case','not-a-login-hash','active',0,now(),now())"), {"id": first_id})
        connection.execute(text("INSERT INTO user_roles (user_id,role_id,assigned_at) VALUES (:user_id,:role_id,now())"), {"user_id": first_id, "role_id": role_id})
    try:
        with engine.begin() as connection:
            connection.execute(text("INSERT INTO users (id,username,display_name,password_hash,status,failed_login_count,created_at,updated_at) VALUES (:id,'p0b1.caseuser','case duplicate','not-a-login-hash','active',0,now(),now())"), {"id": os.urandom(16).hex()})
    except Exception:
        result("case_insensitive_username_unique", "passed", duplicate_rejected=True)
        return
    raise RuntimeError("Database accepted a case-insensitive duplicate username")


def standard_import_file() -> bytes:
    workbook = Workbook()
    items = workbook.active
    items.title = "Items"
    items.append(["item_code", "item", "specification", "qty", "unit"])
    for index in range(1, 6):
        items.append([f"ITEM-{index:03}", f"脱敏品项 {index}", "标准规格", index, "件"])
    suppliers = workbook.create_sheet("Suppliers")
    suppliers.append(["supplier_code", "supplier_name", "currency", "valid_until", "freight", "tax"])
    for index in range(1, 4):
        suppliers.append([f"SUP-{index}", f"供应商 {index}（脱敏）", "HKD", date(2027, 12, 31), 10, 5])
    offers = workbook.create_sheet("Offers")
    offers.append(["supplier_code", "item_code", "unit_price", "remark"])
    for supplier_index in range(1, 4):
        for item_index in range(1, 6):
            offers.append([f"SUP-{supplier_index}", f"ITEM-{item_index:03}", 10 + supplier_index + item_index, "脱敏报价"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def login(application, username: str, password: str) -> TestClient:
    client = TestClient(application, base_url=ORIGIN)
    response = client.post("/api/v1/auth/login", json={"username": username, "password": password}, headers={"Origin": ORIGIN})
    if response.status_code != 200:
        raise RuntimeError(f"Login failed for {username}: {response.status_code}")
    return client


def csrf_headers(client: TestClient, settings) -> dict[str, str]:
    token = client.cookies.get(settings.csrf_cookie_name)
    if not token:
        raise RuntimeError("CSRF cookie missing")
    return {"Origin": ORIGIN, settings.csrf_header_name: token}


def api_data(response, expected_status: int = 200):
    if response.status_code != expected_status:
        raise RuntimeError(f"Unexpected API response {response.status_code}: {response.text[:500]}")
    return response.json()["data"]


def create_account(admin: TestClient, settings, username: str, password: str, roles: list[str]) -> dict:
    return api_data(admin.post(
        "/api/v1/admin/accounts",
        json={"username": username, "display_name": f"{username}（脱敏验证）", "password": password, "roles": roles},
        headers=csrf_headers(admin, settings),
    ), 201)


def create_prechecked_quote(client: TestClient, settings, subject: str) -> tuple[str, str]:
    headers = csrf_headers(client, settings)
    quote = api_data(client.post(
        "/api/v1/steven/quotes",
        json={"subject": subject, "currency": "HKD", "is_demo": True},
        headers=headers,
    ), 201)
    preview = api_data(client.post(
        "/api/v1/steven/quotes/import",
        data={"quote_id": quote["id"]},
        files={"file": ("sanitized-p0b.xlsx", standard_import_file(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    ))
    if (preview["item_count"], preview["supplier_count"], preview["offer_count"], preview["valid"]) != (5, 3, 15, True):
        raise RuntimeError("3x5 import precheck contract failed")
    return quote["id"], preview["batch_id"]


def concurrent_write(clients: list[TestClient], path: str, payload: dict, settings) -> list[tuple[int, str]]:
    def invoke(client: TestClient) -> tuple[int, str]:
        response = client.post(path, json=payload, headers=csrf_headers(client, settings))
        body = response.json()
        return response.status_code, body.get("error", {}).get("code", "success")

    with ThreadPoolExecutor(max_workers=len(clients)) as executor:
        return list(executor.map(invoke, clients))


def application_smoke(database_url: str, file_root: Path) -> None:
    sys.path.insert(0, str(API_ROOT))
    from app.core.config import Settings
    from app.main import create_app

    settings = Settings(
        app_env="staging",
        auth_mode="session",
        demo_seed_enabled=False,
        database_url=database_url,
        file_storage_root=str(file_root),
        allowed_origins=(ORIGIN,),
        rate_limit_mode=os.environ["RATE_LIMIT_MODE"],
        trusted_proxy_cidrs=tuple(
            item.strip()
            for item in os.environ["TRUSTED_PROXY_CIDRS"].split(",")
            if item.strip()
        ),
    )
    application = create_app(settings)
    admin = login(application, os.environ["BOOTSTRAP_ADMIN_USERNAME"], os.environ["BOOTSTRAP_ADMIN_PASSWORD"])
    suffix = os.urandom(5).hex()

    case_username = f"P0B1.Case-{suffix}"

    def create_case_variant(username: str) -> tuple[int, str]:
        client = login(application, os.environ["BOOTSTRAP_ADMIN_USERNAME"], os.environ["BOOTSTRAP_ADMIN_PASSWORD"])
        response = client.post(
            "/api/v1/admin/accounts",
            json={
                "username": username,
                "display_name": "大小写唯一并发验证（脱敏）",
                "password": f"P0B-case-{os.urandom(12).hex()}",
                "roles": ["operator"],
            },
            headers=csrf_headers(client, settings),
        )
        body = response.json()
        return response.status_code, body.get("error", {}).get("code", "success")

    with ThreadPoolExecutor(max_workers=2) as executor:
        username_outcomes = list(executor.map(create_case_variant, [case_username, case_username.lower()]))
    if sorted(status for status, _ in username_outcomes) != [201, 409] or "duplicate_username" not in {code for _, code in username_outcomes}:
        raise RuntimeError(f"Concurrent case-insensitive username contract failed: {username_outcomes}")

    passwords = {name: f"P0B-{name}-{os.urandom(12).hex()}" for name in ("steven", "approver1", "approver2", "dual")}
    users = {
        "steven": create_account(admin, settings, f"p0b-steven-{suffix}", passwords["steven"], ["operator"]),
        "approver1": create_account(admin, settings, f"p0b-approver1-{suffix}", passwords["approver1"], ["approver"]),
        "approver2": create_account(admin, settings, f"p0b-approver2-{suffix}", passwords["approver2"], ["approver"]),
        "dual": create_account(admin, settings, f"p0b-dual-{suffix}", passwords["dual"], ["operator", "approver"]),
    }
    usernames = {key: value["username"] for key, value in users.items()}
    steven = login(application, usernames["steven"], passwords["steven"])
    quote_id, batch_id = create_prechecked_quote(steven, settings, "脱敏 P0-B 并发与重启验证")

    import_outcomes = concurrent_write(
        [login(application, usernames["steven"], passwords["steven"]) for _ in range(2)],
        f"/api/v1/steven/quotes/{quote_id}/confirm-import",
        {"batch_id": batch_id},
        settings,
    )
    if sorted(status for status, _ in import_outcomes) != [200, 409]:
        raise RuntimeError(f"Concurrent import contract failed: {import_outcomes}")

    restarted_application = create_app(settings)
    restarted_steven = login(restarted_application, usernames["steven"], passwords["steven"])
    persisted = api_data(restarted_steven.get(f"/api/v1/steven/quotes/{quote_id}"))
    if (len(persisted["items"]), len(persisted["suppliers"]), len(persisted["offer_lines"])) != (5, 3, 15):
        raise RuntimeError("Restart persistence contract failed")
    totals = [entry["total"] for entry in persisted["comparison"]["ranking"]]
    if [float(value) for value in totals] != [235.0, 250.0, 265.0]:
        raise RuntimeError(f"Restarted supplier totals are incorrect: {totals}")
    lowest_supplier_id = persisted["comparison"]["lowest_supplier_id"]
    api_data(restarted_steven.post(
        f"/api/v1/steven/quotes/{quote_id}/recommendation",
        json={"recommended_supplier_id": lowest_supplier_id, "non_lowest_reason": "", "approval_opinion": ""},
        headers=csrf_headers(restarted_steven, settings),
    ))
    api_data(restarted_steven.post(
        f"/api/v1/steven/quotes/{quote_id}/submit-approval",
        headers=csrf_headers(restarted_steven, settings),
    ))

    approval_outcomes = concurrent_write(
        [
            login(restarted_application, usernames["approver1"], passwords["approver1"]),
            login(restarted_application, usernames["approver2"], passwords["approver2"]),
        ],
        f"/api/v1/steven/quotes/{quote_id}/approve",
        {"opinion": "脱敏并发审批验证"},
        settings,
    )
    if sorted(status for status, _ in approval_outcomes) != [200, 409]:
        raise RuntimeError(f"Concurrent approval contract failed: {approval_outcomes}")

    dual = login(restarted_application, usernames["dual"], passwords["dual"])
    self_quote_id, self_batch_id = create_prechecked_quote(dual, settings, "脱敏自审阻断验证")
    api_data(dual.post(
        f"/api/v1/steven/quotes/{self_quote_id}/confirm-import",
        json={"batch_id": self_batch_id},
        headers=csrf_headers(dual, settings),
    ))
    self_quote = api_data(dual.get(f"/api/v1/steven/quotes/{self_quote_id}"))
    api_data(dual.post(
        f"/api/v1/steven/quotes/{self_quote_id}/recommendation",
        json={"recommended_supplier_id": self_quote["comparison"]["lowest_supplier_id"], "non_lowest_reason": "", "approval_opinion": ""},
        headers=csrf_headers(dual, settings),
    ))
    api_data(dual.post(f"/api/v1/steven/quotes/{self_quote_id}/submit-approval", headers=csrf_headers(dual, settings)))
    self_approval = dual.post(
        f"/api/v1/steven/quotes/{self_quote_id}/approve",
        json={"opinion": "不得成功"},
        headers=csrf_headers(dual, settings),
    )
    if self_approval.status_code != 403 or self_approval.json()["error"]["code"] != "self_approval_forbidden":
        raise RuntimeError("Self-approval was not rejected")

    export_outcomes = concurrent_write(
        [login(restarted_application, usernames["steven"], passwords["steven"]) for _ in range(10)],
        f"/api/v1/steven/quotes/{quote_id}/export",
        {},
        settings,
    )
    if any(status != 200 for status, _ in export_outcomes):
        raise RuntimeError(f"Concurrent export failed: {export_outcomes}")

    final_application = create_app(settings)
    final_steven = login(final_application, usernames["steven"], passwords["steven"])
    versions = api_data(final_steven.get(f"/api/v1/steven/quotes/{quote_id}/versions"))
    if [item["version_number"] for item in versions] != list(range(1, 11)):
        raise RuntimeError("Export version numbers are not unique and contiguous for this clean smoke run")
    for version in versions:
        if version["status"] != "ready":
            raise RuntimeError("A concurrent export version is not ready")
        path = (file_root / version["storage_key"]).resolve()
        if file_root not in path.parents or not path.is_file():
            raise RuntimeError("Export path escaped the smoke root or is missing")
        load_workbook(path, read_only=True).close()
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest != version["sha256"] or path.stat().st_size != version["size_bytes"]:
            raise RuntimeError("Export file metadata mismatch")
    audits = api_data(final_steven.get(f"/api/v1/steven/quotes/{quote_id}/audit-events"))
    actions = {event["action"] for event in audits}
    if not {"quote.import_confirm", "quote.recommend", "quote.submit", "quote.approve", "quote.export"}.issubset(actions):
        raise RuntimeError("Persistent quote audit events are incomplete")
    admin_after_restart = login(final_application, os.environ["BOOTSTRAP_ADMIN_USERNAME"], os.environ["BOOTSTRAP_ADMIN_PASSWORD"])
    request_id = f"p0b1-{os.urandom(8).hex()}"
    denied = final_steven.get("/api/v1/audit/events", headers={"X-Request-Id": request_id})
    if denied.status_code != 403:
        raise RuntimeError("Expected a traceable permission denial")
    platform_audits = api_data(admin_after_restart.get("/api/v1/audit/events"))
    if not any(event["request_id"] == request_id and event["action"] == "auth.authorization_rejected" for event in platform_audits):
        raise RuntimeError("Persistent platform audit request_id association is missing")
    result(
        "s2_restart_transaction_concurrency",
        "passed",
        quote_id=quote_id,
        item_count=5,
        supplier_count=3,
        offer_count=15,
        totals=totals,
        concurrent_import=import_outcomes,
        concurrent_approval=approval_outcomes,
        export_versions=[item["version_number"] for item in versions],
        concurrent_username=username_outcomes,
        self_approval_rejected=True,
    )


def main() -> int:
    required = [
        "P0B_ADMIN_DATABASE_URL",
        "DATABASE_URL",
        "P0B_FILE_STORAGE_ROOT",
        "BOOTSTRAP_ADMIN_USERNAME",
        "BOOTSTRAP_ADMIN_DISPLAY_NAME",
        "BOOTSTRAP_ADMIN_PASSWORD",
        "RATE_LIMIT_MODE",
        "TRUSTED_PROXY_CIDRS",
    ]
    missing = [name for name in required if not os.getenv(name)]
    if missing:
        result("configuration", "failed", missing=missing)
        return 2
    if os.getenv("APP_ENV") not in {"development", "test", "staging"} or os.getenv("AUTH_MODE") != "session" or os.getenv("DEMO_SEED_ENABLED", "").lower() != "false":
        result("configuration", "failed", reason="APP_ENV must be development/test/staging, AUTH_MODE=session, DEMO_SEED_ENABLED=false")
        return 2
    if os.getenv("FILE_STORAGE_ROOT") != os.getenv("P0B_FILE_STORAGE_ROOT"):
        result("configuration", "failed", reason="FILE_STORAGE_ROOT must exactly match P0B_FILE_STORAGE_ROOT")
        return 2
    if os.getenv("RATE_LIMIT_MODE") != "gateway":
        result("configuration", "failed", reason="P0-B smoke requires RATE_LIMIT_MODE=gateway; Redis is not contacted by this script")
        return 2
    try:
        file_root = validate_storage_root(os.environ["P0B_FILE_STORAGE_ROOT"])
        prepare_empty_database(os.environ["P0B_ADMIN_DATABASE_URL"], os.environ["DATABASE_URL"])
        migration_and_schema_smoke(os.environ["DATABASE_URL"])
        bootstrap_smoke(os.environ["BOOTSTRAP_ADMIN_PASSWORD"])
        username_case_conflict_smoke(os.environ["DATABASE_URL"])
        application_smoke(os.environ["DATABASE_URL"], file_root)
    except Exception as error:
        result("verification", "failed", error=type(error).__name__, message=str(error))
        return 1
    result("verification", "passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

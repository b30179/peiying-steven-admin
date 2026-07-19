from __future__ import annotations

import argparse
from contextlib import contextmanager
from datetime import date
from decimal import Decimal
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import socket
import shutil
import sys
from typing import Any, Iterator
from uuid import uuid4

from sqlalchemy import Connection, Engine, create_engine, text
from sqlalchemy.engine import make_url


PROJECT_ROOT = Path(__file__).resolve().parents[1]
API_ROOT = PROJECT_ROOT / "apps" / "api"
EXPECTED_DATABASE = "puiying_steven_demo"
EXPECTED_DATABASE_USER = "puiying_steven_demo_app"
EXPECTED_DATABASE_HOST = "127.0.0.1"
EXPECTED_DATABASE_PORT = 5432
EXPECTED_POSTGRES_MAJOR = 18
EXPECTED_REVISION = "20260717_0014"
STANDARD_S2_SUBJECT = "DEMO-S2-STANDARD-20260717"
STANDARD_S1_DOCUMENT = "DEMO-S1-STANDARD-20260717"
STANDARD_S3_COUNT = "DEMO-S3-STANDARD-20260717"
STANDARD_TEMPLATE_ID = "template-s1-demo-001"
STANDARD_TEMPLATE_CODE = "DEMO-SERVICE-INVITATION"
RUNTIME_STATE = PROJECT_ROOT / "data" / "runtime" / "steven-demo-processes.json"
DEFAULT_FILE_ROOT = PROJECT_ROOT / "data" / "steven-demo-d1"
QUARANTINE_ROOT = PROJECT_ROOT / "data" / "demo-reset-quarantine"
JOURNAL_ROOT = PROJECT_ROOT / "data" / "demo-reset-journals"
ALLOWED_STORAGE_PREFIXES = (
    "generated/quotes/",
    "generated/tenders/",
    "generated/inventory/",
)
ADVISORY_LOCK_KEY = 0x53544556454E
RUNTIME_PORTS = (9000, 4300, 15443, 20219)

sys.path.insert(0, str(API_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))


class ResetBlocked(RuntimeError):
    pass


def database_url_from_environment() -> str:
    configured = os.environ.get("DATABASE_URL", "").strip()
    if configured:
        return configured
    host = os.environ.get("PGHOST", EXPECTED_DATABASE_HOST).strip()
    port = os.environ.get("PGPORT", str(EXPECTED_DATABASE_PORT)).strip()
    user = os.environ.get("PGUSER", EXPECTED_DATABASE_USER).strip()
    database = os.environ.get("PGDATABASE", EXPECTED_DATABASE).strip()
    if not all((host, port, user, database)):
        raise ResetBlocked("PostgreSQL connection settings are incomplete.")
    return f"postgresql+psycopg://{user}@{host}:{port}/{database}"


def validate_runtime_environment(database_url: str) -> None:
    app_env = os.environ.get("APP_ENV", "development").strip().lower()
    if app_env != "development":
        raise ResetBlocked("Demo reset requires APP_ENV=development.")
    if os.environ.get("AUTH_MODE", "session").strip().lower() != "session":
        raise ResetBlocked("AUTH_MODE=session is required.")
    if os.environ.get("DEMO_SEED_ENABLED", "false").strip().lower() not in {"false", "0", "no", "off"}:
        raise ResetBlocked("DEMO_SEED_ENABLED must remain false.")
    if os.environ.get("OCR_ENABLED", "false").strip().lower() in {"true", "1", "yes", "on"}:
        raise ResetBlocked("OCR live mode must remain disabled.")
    if os.environ.get("AI_STRUCTURING_ENABLED", "false").strip().lower() in {"true", "1", "yes", "on"}:
        raise ResetBlocked("AI live mode must remain disabled.")
    parsed = make_url(database_url)
    if parsed.database != EXPECTED_DATABASE:
        raise ResetBlocked(f"Refusing to access a database other than {EXPECTED_DATABASE}.")
    if parsed.username != EXPECTED_DATABASE_USER:
        raise ResetBlocked(f"Demo reset requires database role {EXPECTED_DATABASE_USER}.")
    if parsed.host != EXPECTED_DATABASE_HOST:
        raise ResetBlocked(f"Demo reset requires PostgreSQL host {EXPECTED_DATABASE_HOST}.")
    if (parsed.port or 5432) != EXPECTED_DATABASE_PORT:
        raise ResetBlocked(f"Demo reset requires PostgreSQL port {EXPECTED_DATABASE_PORT}.")


def validate_file_root() -> Path:
    expected = DEFAULT_FILE_ROOT.resolve()
    configured = os.environ.get("FILE_STORAGE_ROOT", "").strip()
    if configured and Path(configured).resolve() != expected:
        raise ResetBlocked(f"Demo reset requires the fixed project file root: {expected}")
    return expected


def runtime_ports_in_use() -> list[int]:
    active: list[int] = []
    for port in RUNTIME_PORTS:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as probe:
            probe.settimeout(0.2)
            if probe.connect_ex((EXPECTED_DATABASE_HOST, port)) == 0:
                active.append(port)
    return active


def database_identity(connection: Connection) -> dict[str, Any]:
    row = connection.execute(
        text(
            """
            SELECT current_database() AS database_name,
                   current_user AS database_user,
                   host(inet_server_addr()) AS server_address,
                   inet_server_port() AS server_port,
                   current_setting('server_version_num')::integer AS server_version_num,
                   (SELECT version_num FROM alembic_version) AS revision
            """
        )
    ).mappings().one()
    identity = {
        "database": row["database_name"],
        "user": row["database_user"],
        "host": row["server_address"],
        "port": int(row["server_port"]),
        "postgres_major": int(row["server_version_num"]) // 10000,
        "revision": row["revision"],
        "file_root": str(DEFAULT_FILE_ROOT.resolve()),
    }
    expected = {
        "database": EXPECTED_DATABASE,
        "user": EXPECTED_DATABASE_USER,
        "host": EXPECTED_DATABASE_HOST,
        "port": EXPECTED_DATABASE_PORT,
        "postgres_major": EXPECTED_POSTGRES_MAJOR,
        "revision": EXPECTED_REVISION,
        "file_root": str(DEFAULT_FILE_ROOT.resolve()),
    }
    if identity != expected:
        raise ResetBlocked("Connected PostgreSQL instance does not match the frozen Steven Demo binding.")
    return identity


def canonical_plan_payload(plan: dict[str, Any]) -> bytes:
    return json.dumps(plan, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")


def calculate_plan_hash(plan: dict[str, Any]) -> str:
    return hashlib.sha256(canonical_plan_payload(plan)).hexdigest()


def summarize_plan(plan: dict[str, Any]) -> dict[str, Any]:
    return {
        "database": plan["database"],
        "revision": plan["revision"],
        "binding": plan["binding"],
        "row_counts": plan["row_counts"],
        "file_count": len(plan["files"]),
        "file_counts_by_module": {
            module: sum(1 for entry in plan["files"] if entry["module"] == module)
            for module in ("s1", "s2", "s3")
        },
        "preserved": plan["preserved"],
    }


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def safe_storage_path(file_root: Path, storage_key: str) -> Path:
    normalized = storage_key.replace("\\", "/").lstrip("/")
    if not any(normalized.startswith(prefix) for prefix in ALLOWED_STORAGE_PREFIXES):
        raise ResetBlocked(f"Storage key is outside the controlled Demo namespace: {storage_key}")
    root = file_root.resolve()
    candidate = (root / Path(normalized)).resolve()
    if candidate == root or root not in candidate.parents:
        raise ResetBlocked(f"Storage key escapes the controlled file root: {storage_key}")
    return candidate


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Controlled Steven redacted Demo reset.")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--dry-run",
        action="store_true",
        help="Print the controlled reset plan without changing database rows or files (default).",
    )
    mode.add_argument("--apply", action="store_true", help="Apply the reviewed plan.")
    parser.add_argument(
        "--confirm-local-redacted-demo",
        action="store_true",
        help="Confirm this is the approved local redacted Demo database.",
    )
    parser.add_argument("--expected-plan-hash", default="", help="SHA-256 printed by the immediately preceding dry-run.")
    return parser.parse_args(argv)


def validate_apply_arguments(arguments: argparse.Namespace) -> None:
    if not arguments.apply:
        if arguments.confirm_local_redacted_demo or arguments.expected_plan_hash:
            raise ResetBlocked("Confirmation flags are only valid together with --apply.")
        return
    if not arguments.confirm_local_redacted_demo:
        raise ResetBlocked("--apply requires --confirm-local-redacted-demo.")
    if len(arguments.expected_plan_hash) != 64:
        raise ResetBlocked("--apply requires a 64-character --expected-plan-hash from dry-run.")
    try:
        int(arguments.expected_plan_hash, 16)
    except ValueError as error:
        raise ResetBlocked("--expected-plan-hash must be hexadecimal.") from error
    if RUNTIME_STATE.exists():
        raise ResetBlocked("Steven Demo is running or has a state file. Run the stop script before --apply.")
    active_ports = runtime_ports_in_use()
    if active_ports:
        ports = ",".join(str(port) for port in active_ports)
        raise ResetBlocked(f"Steven Demo runtime ports are active ({ports}). Run the stop script before --apply.")


def scalar(connection: Connection, statement: str, **parameters: Any) -> int:
    return int(connection.execute(text(statement), parameters).scalar_one())


def actor_for_role(connection: Connection, role_code: str) -> str:
    actor = connection.execute(
        text(
            """
            SELECT u.id
              FROM users u
              JOIN user_roles ur ON ur.user_id=u.id
              JOIN roles r ON r.id=ur.role_id
             WHERE u.status='active'
               AND r.status='active'
               AND r.code=:role_code
             ORDER BY u.created_at,u.id
             LIMIT 1
            """
        ),
        {"role_code": role_code},
    ).scalar_one_or_none()
    if actor is None:
        raise ResetBlocked(f"No active {role_code} actor is available.")
    return str(actor)


def inspect_namespace_job(
    connection: Connection,
    table: str,
    key_column: str,
    key_value: str,
) -> dict[str, Any] | None:
    rows = connection.execute(
        text(f"SELECT id,is_demo FROM {table} WHERE {key_column}=:key_value"),
        {"key_value": key_value},
    ).mappings().all()
    if len(rows) > 1:
        raise ResetBlocked(f"Multiple rows use the controlled namespace in {table}.")
    if not rows:
        return None
    row = rows[0]
    if not row["is_demo"]:
        raise ResetBlocked(f"Protected non-Demo row uses the standard namespace in {table}.")
    return {"id": str(row["id"]), "is_demo": True}


def file_shared_reference_count(
    connection: Connection,
    file_id: str,
    tender_version_ids: list[str],
    inventory_version_ids: list[str],
) -> int:
    return scalar(
        connection,
        """
        SELECT
            (SELECT count(*) FROM ocr_jobs WHERE file_id=:file_id)
          + (SELECT count(*) FROM ai_jobs WHERE file_id=:file_id)
          + (SELECT count(*) FROM review_candidates WHERE source_file_id=:file_id)
          + (SELECT count(*) FROM steven_quote_jobs WHERE source_file_id=:file_id)
          + (SELECT count(*) FROM steven_quote_suppliers WHERE source_file_id=:file_id)
          + (SELECT count(*) FROM steven_tender_versions
              WHERE file_id=:file_id AND NOT (id = ANY(CAST(:tender_ids AS text[]))))
          + (SELECT count(*) FROM steven_inventory_versions
              WHERE file_id=:file_id AND NOT (id = ANY(CAST(:inventory_ids AS text[]))))
        """,
        file_id=file_id,
        tender_ids=tender_version_ids,
        inventory_ids=inventory_version_ids,
    )


def version_rows(connection: Connection, table: str, foreign_key: str, object_id: str | None) -> list[dict[str, Any]]:
    if object_id is None:
        return []
    rows = connection.execute(
        text(
            f"""
            SELECT id,storage_key,status,sha256,size_bytes
                   {",file_id" if table != "steven_quote_versions" else ""}
              FROM {table}
             WHERE {foreign_key}=:object_id
             ORDER BY version_number,id
            """
        ),
        {"object_id": object_id},
    ).mappings().all()
    return [
        {
            "id": str(row["id"]),
            "storage_key": row["storage_key"],
            "status": row["status"],
            "sha256": row["sha256"],
            "size_bytes": row["size_bytes"],
            **({"file_id": str(row["file_id"]) if row["file_id"] else None} if "file_id" in row else {}),
        }
        for row in rows
    ]


def row_counts(connection: Connection, s2_id: str | None, s1_id: str | None, s3_id: str | None) -> dict[str, int]:
    counts = {
        "s2_jobs": 1 if s2_id else 0,
        "s1_jobs": 1 if s1_id else 0,
        "s3_counts": 1 if s3_id else 0,
    }
    if s2_id:
        counts.update(
            {
                "s2_items": scalar(connection, "SELECT count(*) FROM steven_quote_items WHERE quote_job_id=:id", id=s2_id),
                "s2_suppliers": scalar(connection, "SELECT count(*) FROM steven_quote_suppliers WHERE quote_job_id=:id", id=s2_id),
                "s2_offers": scalar(
                    connection,
                    """
                    SELECT count(*) FROM steven_quote_offer_lines
                     WHERE quote_supplier_id IN (SELECT id FROM steven_quote_suppliers WHERE quote_job_id=:id)
                        OR quote_item_id IN (SELECT id FROM steven_quote_items WHERE quote_job_id=:id)
                    """,
                    id=s2_id,
                ),
                "s2_batches": scalar(connection, "SELECT count(*) FROM steven_quote_import_batches WHERE quote_job_id=:id", id=s2_id),
                "s2_approvals": scalar(connection, "SELECT count(*) FROM steven_quote_approvals WHERE quote_job_id=:id", id=s2_id),
                "s2_versions": scalar(connection, "SELECT count(*) FROM steven_quote_versions WHERE quote_job_id=:id", id=s2_id),
                "s2_candidate_links": scalar(connection, "SELECT count(*) FROM steven_quote_import_candidates WHERE quote_job_id=:id", id=s2_id),
            }
        )
    if s1_id:
        counts.update(
            {
                "s1_suppliers": scalar(connection, "SELECT count(*) FROM steven_tender_suppliers WHERE tender_job_id=:id", id=s1_id),
                "s1_versions": scalar(connection, "SELECT count(*) FROM steven_tender_versions WHERE tender_job_id=:id", id=s1_id),
                "s1_candidate_links": scalar(connection, "SELECT count(*) FROM steven_tender_candidate_links WHERE tender_job_id=:id", id=s1_id),
            }
        )
    if s3_id:
        counts.update(
            {
                "s3_lines": scalar(connection, "SELECT count(*) FROM steven_inventory_count_lines WHERE inventory_count_id=:id", id=s3_id),
                "s3_versions": scalar(connection, "SELECT count(*) FROM steven_inventory_versions WHERE inventory_count_id=:id", id=s3_id),
                "s3_candidate_links": scalar(connection, "SELECT count(*) FROM steven_inventory_candidate_links WHERE inventory_count_id=:id", id=s3_id),
            }
        )
    return dict(sorted(counts.items()))


def build_plan(connection: Connection, file_root: Path) -> dict[str, Any]:
    binding = database_identity(connection)
    revision = binding["revision"]
    if revision != EXPECTED_REVISION:
        raise ResetBlocked(f"Alembic revision must be {EXPECTED_REVISION}; found {revision}.")

    actors = {role: actor_for_role(connection, role) for role in ("operator", "approver", "admin")}
    s2 = inspect_namespace_job(connection, "steven_quote_jobs", "subject", STANDARD_S2_SUBJECT)
    s1 = inspect_namespace_job(connection, "steven_tender_jobs", "document_number", STANDARD_S1_DOCUMENT)
    s3 = inspect_namespace_job(connection, "steven_inventory_counts", "count_number", STANDARD_S3_COUNT)
    s2_id = s2["id"] if s2 else None
    s1_id = s1["id"] if s1 else None
    s3_id = s3["id"] if s3 else None

    from seed_s3_demo import DEMO_ITEMS

    normalized_skus = [item["sku"].strip().upper() for item in DEMO_ITEMS]
    non_demo_item_references = scalar(
        connection,
        """
        SELECT count(*)
          FROM steven_inventory_count_lines line
          JOIN steven_inventory_counts count ON count.id=line.inventory_count_id
          JOIN steven_inventory_items item ON item.id=line.inventory_item_id
         WHERE item.normalized_sku = ANY(CAST(:skus AS text[]))
           AND count.is_demo=false
        """,
        skus=normalized_skus,
    )
    if non_demo_item_references:
        raise ResetBlocked("A non-Demo inventory count references a standard Demo SKU.")

    quote_versions = version_rows(connection, "steven_quote_versions", "quote_job_id", s2_id)
    tender_versions = version_rows(connection, "steven_tender_versions", "tender_job_id", s1_id)
    inventory_versions = version_rows(connection, "steven_inventory_versions", "inventory_count_id", s3_id)
    tender_version_ids = [row["id"] for row in tender_versions]
    inventory_version_ids = [row["id"] for row in inventory_versions]

    files: list[dict[str, Any]] = []
    for module, rows in (
        ("s2", quote_versions),
        ("s1", tender_versions),
        ("s3", inventory_versions),
    ):
        for row in rows:
            path = safe_storage_path(file_root, row["storage_key"])
            if row.get("file_id"):
                shared = file_shared_reference_count(
                    connection,
                    row["file_id"],
                    tender_version_ids,
                    inventory_version_ids,
                )
                if shared:
                    raise ResetBlocked(f"Generated file {row['file_id']} has shared references.")
            if row["status"] == "ready" and not path.is_file():
                raise ResetBlocked(f"Ready export file is missing: {row['storage_key']}")
            actual_size = path.stat().st_size if path.is_file() else None
            actual_sha256 = file_sha256(path) if path.is_file() else None
            if row["status"] == "ready" and (
                row["size_bytes"] != actual_size or row["sha256"] != actual_sha256
            ):
                raise ResetBlocked(f"Ready export metadata differs from the published file: {row['storage_key']}")
            files.append(
                {
                    "module": module,
                    "version_id": row["id"],
                    "file_id": row.get("file_id"),
                    "storage_key": row["storage_key"],
                    "exists": path.is_file(),
                    "db_size_bytes": row["size_bytes"],
                    "db_sha256": row["sha256"],
                    "actual_size_bytes": actual_size,
                    "actual_sha256": actual_sha256,
                }
            )

    return {
        "database": EXPECTED_DATABASE,
        "revision": EXPECTED_REVISION,
        "binding": binding,
        "namespace": {
            "s2_subject": STANDARD_S2_SUBJECT,
            "s1_document_number": STANDARD_S1_DOCUMENT,
            "s3_count_number": STANDARD_S3_COUNT,
            "s1_template_id": STANDARD_TEMPLATE_ID,
            "s1_template_code": STANDARD_TEMPLATE_CODE,
            "s3_skus": normalized_skus,
        },
        "object_ids": {"s2": s2_id, "s1": s1_id, "s3": s3_id},
        "row_counts": row_counts(connection, s2_id, s1_id, s3_id),
        "files": sorted(files, key=lambda item: (item["module"], item["storage_key"])),
        "actors": actors,
        "preserved": [
            "accounts_roles_permissions_sessions",
            "alembic_schema_and_revision",
            "platform_and_business_audit_history",
            "non_standard_demo_and_non_demo_data",
            "project_tls_and_runtime_configuration",
        ],
    }


def validate_plan(plan: dict[str, Any], file_root: Path) -> None:
    if plan["database"] != EXPECTED_DATABASE or plan["revision"] != EXPECTED_REVISION:
        raise ResetBlocked("Plan database or revision changed.")
    if plan["binding"]["file_root"] != str(file_root.resolve()):
        raise ResetBlocked("Plan file root differs from the frozen Steven Demo binding.")
    for file_entry in plan["files"]:
        safe_storage_path(file_root, file_entry["storage_key"])


def validate_expected_plan_hash(supplied: str, current: str) -> None:
    if supplied.lower() != current:
        raise ResetBlocked("The supplied plan hash does not match the current dry-run.")


def write_reset_journal(run_id: str, payload: dict[str, Any]) -> Path:
    JOURNAL_ROOT.mkdir(parents=True, exist_ok=True)
    target = JOURNAL_ROOT / f"{run_id}.json"
    temporary = JOURNAL_ROOT / f".{run_id}.{uuid4().hex}.tmp"
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2, default=str),
        encoding="utf-8",
    )
    os.replace(temporary, target)
    return target


def quarantine_files(plan: dict[str, Any], run_id: str, file_root: Path) -> list[tuple[Path, Path]]:
    moved: list[tuple[Path, Path]] = []
    for entry in plan["files"]:
        if not entry["exists"]:
            continue
        source = safe_storage_path(file_root, entry["storage_key"])
        target = safe_storage_path(QUARANTINE_ROOT / run_id, entry["storage_key"])
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.exists():
            raise ResetBlocked(f"Quarantine target already exists: {entry['storage_key']}")
        shutil.move(str(source), str(target))
        moved.append((source, target))
    return moved


def restore_quarantined_files(moved: list[tuple[Path, Path]]) -> None:
    for source, target in reversed(moved):
        if not target.exists():
            continue
        source.parent.mkdir(parents=True, exist_ok=True)
        if source.exists():
            raise ResetBlocked(f"Cannot restore quarantined file because the original path exists: {source.name}")
        shutil.move(str(target), str(source))


def append_reset_audit(
    connection: Connection,
    *,
    actor: str,
    action: str,
    run_id: str,
    plan_hash: str,
    outcome: str,
    details: dict[str, Any],
) -> None:
    connection.execute(
        text(
            """
            INSERT INTO platform_audit_events
                (id,actor_user_id,actor_label,action,outcome,object_type,object_id,
                 request_id,occurred_at,before_after)
            VALUES
                (:id,:actor,NULL,:action,:outcome,'steven_demo_reset',:run_id,
                 :request_id,now(),CAST(:before_after AS jsonb))
            """
        ),
        {
            "id": str(uuid4()),
            "actor": actor,
            "action": action,
            "outcome": outcome,
            "run_id": run_id,
            "request_id": f"demo-reset-{run_id}",
            "before_after": json.dumps(
                {
                    "plan_hash": plan_hash,
                    "row_counts": details.get("row_counts", {}),
                    "file_count": len(details.get("files", [])),
                    "stage": details.get("stage"),
                    "error_type": details.get("error_type"),
                },
                ensure_ascii=False,
            ),
        },
    )


def delete_standard_namespace(connection: Connection, plan: dict[str, Any]) -> None:
    s2_id = plan["object_ids"]["s2"]
    s1_id = plan["object_ids"]["s1"]
    s3_id = plan["object_ids"]["s3"]

    if s2_id:
        connection.execute(
            text("UPDATE steven_quote_jobs SET recommended_supplier_id=NULL,approval_id=NULL WHERE id=:id"),
            {"id": s2_id},
        )
        connection.execute(text("DELETE FROM steven_quote_import_candidates WHERE quote_job_id=:id"), {"id": s2_id})
        connection.execute(
            text(
                """
                DELETE FROM steven_quote_offer_lines
                 WHERE quote_supplier_id IN (SELECT id FROM steven_quote_suppliers WHERE quote_job_id=:id)
                    OR quote_item_id IN (SELECT id FROM steven_quote_items WHERE quote_job_id=:id)
                """
            ),
            {"id": s2_id},
        )
        for table in (
            "steven_quote_versions",
            "steven_quote_import_batches",
            "steven_quote_approvals",
            "steven_quote_suppliers",
            "steven_quote_items",
        ):
            connection.execute(text(f"DELETE FROM {table} WHERE quote_job_id=:id"), {"id": s2_id})
        connection.execute(text("DELETE FROM steven_quote_jobs WHERE id=:id AND is_demo=true"), {"id": s2_id})

    if s1_id:
        connection.execute(text("DELETE FROM steven_tender_candidate_links WHERE tender_job_id=:id"), {"id": s1_id})
        connection.execute(text("DELETE FROM steven_tender_versions WHERE tender_job_id=:id"), {"id": s1_id})
        connection.execute(text("DELETE FROM steven_tender_suppliers WHERE tender_job_id=:id"), {"id": s1_id})
        connection.execute(text("DELETE FROM steven_tender_jobs WHERE id=:id AND is_demo=true"), {"id": s1_id})

    if s3_id:
        connection.execute(text("DELETE FROM steven_inventory_candidate_links WHERE inventory_count_id=:id"), {"id": s3_id})
        connection.execute(text("DELETE FROM steven_inventory_versions WHERE inventory_count_id=:id"), {"id": s3_id})
        connection.execute(text("DELETE FROM steven_inventory_count_lines WHERE inventory_count_id=:id"), {"id": s3_id})
        connection.execute(text("DELETE FROM steven_inventory_counts WHERE id=:id AND is_demo=true"), {"id": s3_id})

    file_ids = [entry["file_id"] for entry in plan["files"] if entry.get("file_id")]
    if file_ids:
        connection.execute(
            text(
                """
                DELETE FROM files
                 WHERE id = ANY(CAST(:file_ids AS text[]))
                   AND is_demo=true
                   AND NOT EXISTS (SELECT 1 FROM ocr_jobs WHERE ocr_jobs.file_id=files.id)
                   AND NOT EXISTS (SELECT 1 FROM ai_jobs WHERE ai_jobs.file_id=files.id)
                   AND NOT EXISTS (SELECT 1 FROM review_candidates WHERE review_candidates.source_file_id=files.id)
                   AND NOT EXISTS (SELECT 1 FROM steven_tender_versions WHERE steven_tender_versions.file_id=files.id)
                   AND NOT EXISTS (SELECT 1 FROM steven_inventory_versions WHERE steven_inventory_versions.file_id=files.id)
                """
            ),
            {"file_ids": file_ids},
        )


def switch_staging_namespace(
    connection: Connection,
    seeded: dict[str, Any],
) -> None:
    updates = (
        (
            "steven_quote_jobs",
            "subject",
            seeded["s2"]["id"],
            STANDARD_S2_SUBJECT,
        ),
        (
            "steven_tender_jobs",
            "document_number",
            seeded["s1"]["id"],
            STANDARD_S1_DOCUMENT,
        ),
        (
            "steven_inventory_counts",
            "count_number",
            seeded["s3"]["id"],
            STANDARD_S3_COUNT,
        ),
    )
    for table, column, object_id, value in updates:
        result = connection.execute(
            text(
                f"""
                UPDATE {table}
                   SET {column}=:value
                 WHERE id=:object_id
                   AND is_demo=true
                """
            ),
            {"value": value, "object_id": object_id},
        )
        if result.rowcount != 1:
            raise ResetBlocked(f"Staging namespace switch failed for {table}.")


@contextmanager
def request_id(value: str) -> Iterator[None]:
    from app.core.audit_context import reset_request_id, set_request_id

    token = set_request_id(value)
    try:
        yield
    finally:
        reset_request_id(token)


def build_applications(engine: Engine, file_root: Path):
    from app.modules.platform.append_only_storage import LocalAppendOnlyFileStorage
    from app.modules.steven.inventory_application import StevenInventoryApplicationService
    from app.modules.steven.inventory_excel import InventoryExcelRenderer
    from app.modules.steven.inventory_uow import PostgresInventoryUnitOfWork
    from app.modules.steven.quote_application import StevenQuoteApplicationService
    from app.modules.steven.quote_excel import LocalAppendOnlyQuoteStorage, QuoteExcelExporter, QuoteImportParser
    from app.modules.steven.quote_uow import PostgresQuoteUnitOfWork
    from app.modules.steven.tender_application import StevenTenderApplicationService
    from app.modules.steven.tender_uow import PostgresTenderUnitOfWork
    from app.modules.steven.tender_word import TenderWordRenderer

    quote_exporter = QuoteExcelExporter(file_root)
    tender_renderer = TenderWordRenderer()
    inventory_renderer = InventoryExcelRenderer()
    return (
        StevenQuoteApplicationService(
            PostgresQuoteUnitOfWork(engine),
            QuoteImportParser(),
            quote_exporter,
            LocalAppendOnlyQuoteStorage(quote_exporter.data_root),
        ),
        StevenTenderApplicationService(
            PostgresTenderUnitOfWork(engine),
            tender_renderer,
            LocalAppendOnlyFileStorage(file_root, "tenders", "docx", tender_renderer.verify),
        ),
        StevenInventoryApplicationService(
            PostgresInventoryUnitOfWork(engine),
            inventory_renderer,
            LocalAppendOnlyFileStorage(file_root, "inventory", "xlsx", inventory_renderer.verify),
        ),
    )


def ranking_totals_by_supplier_code(comparison: Any, suppliers: list[Any]) -> dict[str, str]:
    supplier_codes = {supplier.id: supplier.supplier_code for supplier in suppliers}
    try:
        return {
            supplier_codes[entry.supplier_id]: str(entry.total)
            for entry in comparison.ranking
        }
    except KeyError as error:
        raise ResetBlocked("Quote comparison references an unknown supplier.") from error


def build_standard_quote_import(candidates: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    item_sheet = workbook.active
    item_sheet.title = "Items"
    item_sheet.append(["item_code", "item", "specification", "qty", "unit"])
    for item in candidates[0]["items"]:
        item_sheet.append(
            [
                item["item_code"],
                item["item"],
                item["specification"],
                Decimal(item["qty"]),
                item["unit"],
            ]
        )

    supplier_sheet = workbook.create_sheet("Suppliers")
    supplier_sheet.append(
        ["supplier_code", "supplier_name", "currency", "valid_until", "freight", "tax"]
    )
    offer_sheet = workbook.create_sheet("Offers")
    offer_sheet.append(["supplier_code", "item_code", "unit_price", "remark"])
    for candidate in candidates:
        supplier_sheet.append(
            [
                candidate["supplier_code"],
                candidate["supplier_name"],
                candidate["currency"],
                date.fromisoformat(candidate["valid_until"]),
                Decimal(candidate["freight"]),
                Decimal(candidate["tax"]),
            ]
        )
        for item in candidate["items"]:
            offer_sheet.append(
                [
                    candidate["supplier_code"],
                    item["item_code"],
                    Decimal(item["unit_price"]),
                    "脱敏标准演示报价",
                ]
            )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def seed_s2(
    application,
    steven_actor: str,
    approver_actor: str,
    run_id: str,
    subject: str = STANDARD_S2_SUBJECT,
) -> dict[str, Any]:
    from app.modules.steven.quote_schemas import (
        QuoteCreateRequest,
        QuoteRecommendationRequest,
    )

    ground_truth_paths = [
        PROJECT_ROOT / "demo-data" / "steven-d0" / name
        for name in (
            "01_supplier_a_zh_scanned.ground-truth.json",
            "02_supplier_b_mixed_numbers.ground-truth.json",
            "03_supplier_c_bilingual_table.ground-truth.json",
        )
    ]
    candidates = [
        json.loads(path.read_text(encoding="utf-8"))["candidate"]
        for path in ground_truth_paths
    ]
    quote = application.create_quote(
        QuoteCreateRequest(subject=subject, currency="HKD", is_demo=True),
        steven_actor,
    )
    preview = application.precheck_import(
        quote_id=quote.id,
        filename="steven-standard-redacted-quotes.xlsx",
        content=build_standard_quote_import(candidates),
        actor=steven_actor,
    )
    imported = application.confirm_import(quote.id, preview.batch_id, steven_actor)
    suppliers = {supplier.supplier_code: supplier for supplier in imported.suppliers}
    application.save_recommendation(
        quote.id,
        QuoteRecommendationRequest(
            recommended_supplier_id=suppliers["SUP-A"].id,
            non_lowest_reason="脱敏演示：综合交付安排后人工选择非最低价供应商。",
            approval_opinion="脱敏演示：审批前已人工复核金额、币种和完整性。",
        ),
        steven_actor,
    )
    application.submit_approval(quote.id, steven_actor)
    application.approve(quote.id, approver_actor, "独立审批人已核对脱敏标准报价。")
    versions = [application.export(quote.id, steven_actor).version for _ in range(3)]
    quote_view = application.get_quote(quote.id)
    totals = ranking_totals_by_supplier_code(quote_view.comparison, quote_view.suppliers)
    return {
        "id": quote.id,
        "versions": [version.version_number for version in versions],
        "totals": totals,
        "request_id": f"demo-reset-{run_id}-s2",
    }


def seed_s1(
    application,
    steven_actor: str,
    approver_actor: str,
    run_id: str,
    document_number: str = STANDARD_S1_DOCUMENT,
) -> dict[str, Any]:
    from app.modules.steven.tender_schemas import TenderCreateRequest

    template = application.ensure_demo_template(steven_actor)
    if template.id != STANDARD_TEMPLATE_ID or template.code != STANDARD_TEMPLATE_CODE:
        raise ResetBlocked("The controlled S1 Demo template identity does not match the frozen contract.")
    tender = application.create_tender(
        TenderCreateRequest(
            template_id=template.id,
            title="脱敏采购服务邀请文书",
            document_number=document_number,
            subject="脱敏设施保养服务",
            generated_date=date(2026, 7, 17),
            deadline_date=date(2026, 7, 20),
            budget_min=Decimal("1000"),
            budget_max=Decimal("3000"),
            currency="HKD",
            location="演示地点 A（脱敏）",
            controlled_clauses="仅接受脱敏演示资料；最终内容必须由人工复核。",
            supplier_names=["演示供应商甲", "演示供应商乙", "演示供应商丙"],
            is_demo=True,
        ),
        steven_actor,
    )
    application.preview(tender.id, steven_actor)
    application.submit(tender.id, steven_actor)
    application.approve(tender.id, approver_actor, "独立审批人已核对脱敏标准文书。")
    versions = [application.export(tender.id, steven_actor).version for _ in range(3)]
    return {
        "id": tender.id,
        "versions": [version.version_number for version in versions],
        "request_id": f"demo-reset-{run_id}-s1",
    }


def seed_s3(
    application,
    steven_actor: str,
    approver_actor: str,
    run_id: str,
    count_number: str = STANDARD_S3_COUNT,
) -> dict[str, Any]:
    from app.modules.steven.inventory_schemas import (
        InventoryCountCreateRequest,
        InventoryCountLineInput,
        InventoryItemCreateRequest,
    )
    from app.modules.steven.inventory_service import normalize_sku
    from seed_s3_demo import DEMO_ITEMS

    existing = {normalize_sku(item.sku)[1]: item for item in application.list_items()}
    items = []
    for values in DEMO_ITEMS:
        normalized = normalize_sku(values["sku"])[1]
        item = existing.get(normalized)
        if item is None:
            item = application.create_item(InventoryItemCreateRequest(**values), steven_actor)
        else:
            expected = {key: values[key] for key in (
                "sku",
                "item_name",
                "category",
                "location",
                "book_quantity",
                "safety_stock",
                "target_stock",
                "is_demo",
            )}
            actual = {key: getattr(item, key) for key in expected}
            if actual != expected:
                raise ResetBlocked(f"Standard Demo inventory item differs from the frozen seed: {values['sku']}")
        items.append(item)

    lines = [
        InventoryCountLineInput(inventory_item_id=items[0].id, counted_quantity=8, confirmed_order_quantity=22, remark="低库存与负差异脱敏案例"),
        InventoryCountLineInput(inventory_item_id=items[1].id, counted_quantity=25, confirmed_order_quantity=0, remark="高于账面且建议订货量为零"),
        InventoryCountLineInput(
            inventory_item_id=items[2].id,
            counted_quantity=7,
            confirmed_order_quantity=20,
            manual_reason="人工确认演示：保留两件缓冲量",
        ),
        InventoryCountLineInput(inventory_item_id=items[3].id, counted_quantity=18, confirmed_order_quantity=0),
        InventoryCountLineInput(inventory_item_id=items[4].id, counted_quantity=35, confirmed_order_quantity=5),
    ]
    count = application.create_count(
        InventoryCountCreateRequest(
            count_number=count_number,
            count_date=date(2026, 7, 17),
            lines=lines,
            is_demo=True,
        ),
        steven_actor,
    )
    application.submit(count.id, steven_actor)
    application.approve(count.id, approver_actor, "独立审批人已核对脱敏标准盘点。")
    versions = [application.export(count.id, steven_actor).version for _ in range(3)]
    return {
        "id": count.id,
        "versions": [version.version_number for version in versions],
        "request_id": f"demo-reset-{run_id}-s3",
    }


def seed_standard_demo(
    engine: Engine,
    actors: dict[str, str],
    run_id: str,
    file_root: Path,
    *,
    s2_subject: str = STANDARD_S2_SUBJECT,
    s1_document_number: str = STANDARD_S1_DOCUMENT,
    s3_count_number: str = STANDARD_S3_COUNT,
) -> dict[str, Any]:
    quote_application, tender_application, inventory_application = build_applications(engine, file_root)
    with request_id(f"demo-reset-{run_id}-s2"):
        s2 = seed_s2(
            quote_application,
            actors["operator"],
            actors["approver"],
            run_id,
            s2_subject,
        )
    with request_id(f"demo-reset-{run_id}-s1"):
        s1 = seed_s1(
            tender_application,
            actors["operator"],
            actors["approver"],
            run_id,
            s1_document_number,
        )
    with request_id(f"demo-reset-{run_id}-s3"):
        s3 = seed_s3(
            inventory_application,
            actors["operator"],
            actors["approver"],
            run_id,
            s3_count_number,
        )
    return {"s2": s2, "s1": s1, "s3": s3}


def verify_demo_objects(
    engine: Engine,
    seeded: dict[str, Any],
    file_root: Path,
    *,
    expected_names: dict[str, str],
) -> dict[str, Any]:
    from app.modules.steven.inventory_excel import InventoryExcelRenderer
    from app.modules.steven.tender_word import TenderWordRenderer
    from openpyxl import load_workbook

    with engine.connect() as connection:
        s2 = connection.execute(
            text(
                """
                SELECT id,status,subject FROM steven_quote_jobs
                 WHERE id=:id AND is_demo=true
                """
            ),
            {"id": seeded["s2"]["id"]},
        ).mappings().one()
        s1 = connection.execute(
            text(
                """
                SELECT id,status,document_number FROM steven_tender_jobs
                 WHERE id=:id AND is_demo=true
                """
            ),
            {"id": seeded["s1"]["id"]},
        ).mappings().one()
        s3 = connection.execute(
            text(
                """
                SELECT id,status,count_number FROM steven_inventory_counts
                 WHERE id=:id AND is_demo=true
                """
            ),
            {"id": seeded["s3"]["id"]},
        ).mappings().one()
        actual_names = {
            "s2": s2["subject"],
            "s1": s1["document_number"],
            "s3": s3["count_number"],
        }
        if actual_names != expected_names:
            raise ResetBlocked(f"Demo namespace verification failed: {actual_names}")
        if s1["status"] != "approved" or s3["status"] != "approved" or s2["status"] != "exported":
            raise ResetBlocked("Demo workflow status verification failed.")

        totals = {
            row["supplier_code"]: str(row["total"])
            for row in connection.execute(
                text(
                    """
                    SELECT supplier.supplier_code,
                           coalesce(sum(offer.line_total),0)
                           + supplier.freight
                           + supplier.tax AS total
                      FROM steven_quote_suppliers supplier
                      LEFT JOIN steven_quote_offer_lines offer
                        ON offer.quote_supplier_id=supplier.id
                     WHERE supplier.quote_job_id=:id
                     GROUP BY supplier.id,supplier.supplier_code,supplier.freight,supplier.tax
                    """
                ),
                {"id": s2["id"]},
            ).mappings()
        }
        expected_totals = {"SUP-C": "2583.00", "SUP-A": "2605.00", "SUP-B": "2622.00"}
        if totals != expected_totals:
            raise ResetBlocked(f"Standard S2 totals verification failed: {totals}")

        version_specs = {
            "s2": ("steven_quote_versions", "quote_job_id", s2["id"]),
            "s1": ("steven_tender_versions", "tender_job_id", s1["id"]),
            "s3": ("steven_inventory_versions", "inventory_count_id", s3["id"]),
        }
        verified_files: dict[str, list[dict[str, Any]]] = {}
        for module, (table, foreign_key, object_id) in version_specs.items():
            rows = version_rows(connection, table, foreign_key, str(object_id))
            ready = [row for row in rows if row["status"] == "ready"]
            if len(rows) != 3 or len(ready) != 3:
                raise ResetBlocked(f"{module.upper()} Demo must contain exactly three ready export versions.")
            version_numbers = connection.execute(
                text(
                    f"""
                    SELECT version_number
                      FROM {table}
                     WHERE {foreign_key}=:object_id
                     ORDER BY version_number
                    """
                ),
                {"object_id": object_id},
            ).scalars().all()
            if list(version_numbers) != [1, 2, 3]:
                raise ResetBlocked(f"{module.upper()} Demo version sequence is not v1/v2/v3.")

            file_evidence: list[dict[str, Any]] = []
            for row in ready:
                path = safe_storage_path(file_root, row["storage_key"])
                if not path.is_file():
                    raise ResetBlocked(f"{module.upper()} Demo export is missing: {row['storage_key']}")
                if module == "s2":
                    load_workbook(path, read_only=True, data_only=True).close()
                elif module == "s1":
                    TenderWordRenderer.verify(path)
                else:
                    InventoryExcelRenderer.verify(path)
                digest = file_sha256(path)
                size_bytes = path.stat().st_size
                if row["sha256"] != digest or row["size_bytes"] != size_bytes:
                    raise ResetBlocked(f"{module.upper()} Demo export metadata mismatch.")
                if row.get("file_id"):
                    file_row = connection.execute(
                        text(
                            """
                            SELECT storage_key,sha256,size_bytes,status,is_demo
                              FROM files
                             WHERE id=:file_id
                            """
                        ),
                        {"file_id": row["file_id"]},
                    ).mappings().one()
                    if (
                        file_row["storage_key"] != row["storage_key"]
                        or file_row["sha256"] != digest
                        or file_row["size_bytes"] != size_bytes
                        or file_row["status"] != "stored"
                        or not file_row["is_demo"]
                    ):
                        raise ResetBlocked(f"{module.upper()} files metadata mismatch.")
                file_evidence.append(
                    {
                        "version_id": row["id"],
                        "storage_key": row["storage_key"],
                        "sha256": digest,
                        "size_bytes": size_bytes,
                    }
                )
            verified_files[module] = file_evidence

        return {
            "object_ids": {"s2": str(s2["id"]), "s1": str(s1["id"]), "s3": str(s3["id"])},
            "statuses": {"s2": s2["status"], "s1": s1["status"], "s3": s3["status"]},
            "ready_versions": {"s2": 3, "s1": 3, "s3": 3},
            "totals": totals,
            "files": verified_files,
        }


def verify_standard_demo(engine: Engine, file_root: Path) -> dict[str, Any]:
    with engine.connect() as connection:
        seeded = {
            "s2": inspect_namespace_job(connection, "steven_quote_jobs", "subject", STANDARD_S2_SUBJECT),
            "s1": inspect_namespace_job(
                connection,
                "steven_tender_jobs",
                "document_number",
                STANDARD_S1_DOCUMENT,
            ),
            "s3": inspect_namespace_job(
                connection,
                "steven_inventory_counts",
                "count_number",
                STANDARD_S3_COUNT,
            ),
        }
    if any(value is None for value in seeded.values()):
        raise ResetBlocked("One or more standard Demo namespaces are missing.")
    return verify_demo_objects(
        engine,
        seeded,
        file_root,
        expected_names={
            "s2": STANDARD_S2_SUBJECT,
            "s1": STANDARD_S1_DOCUMENT,
            "s3": STANDARD_S3_COUNT,
        },
    )


class TransactionConnectionView:
    def __init__(self, connection: Connection):
        self.connection = connection

    @contextmanager
    def connect(self) -> Iterator[Connection]:
        yield self.connection


def verify_switched_demo(
    connection: Connection,
    seeded: dict[str, Any],
    file_root: Path,
) -> dict[str, Any]:
    return verify_demo_objects(
        TransactionConnectionView(connection),
        seeded,
        file_root,
        expected_names={
            "s2": STANDARD_S2_SUBJECT,
            "s1": STANDARD_S1_DOCUMENT,
            "s3": STANDARD_S3_COUNT,
        },
    )


def apply_reset(engine: Engine, plan: dict[str, Any], plan_hash: str, file_root: Path) -> dict[str, Any]:
    run_id = uuid4().hex
    moved: list[tuple[Path, Path]] = []
    actors = dict(plan["actors"])
    seeded: dict[str, Any] | None = None
    journal = {
        "run_id": run_id,
        "plan_hash": plan_hash,
        "phase": "initializing",
        "old_plan": plan,
        "staging": None,
        "quarantined_files": [],
        "error_type": None,
    }
    write_reset_journal(run_id, journal)
    lock_connection = engine.connect()
    try:
        locked = lock_connection.execute(
            text("SELECT pg_try_advisory_lock(:key)"),
            {"key": ADVISORY_LOCK_KEY},
        ).scalar_one()
        if not locked:
            raise ResetBlocked("Another Steven Demo reset is already running.")
        database_identity(lock_connection)
        current_plan = build_plan(lock_connection, file_root)
        if calculate_plan_hash(current_plan) != plan_hash:
            raise ResetBlocked("The database changed after dry-run; generate and approve a new plan hash.")
        lock_connection.rollback()

        staging_names = {
            "s2": f"DEMO-S2-STAGING-{run_id}",
            "s1": f"DEMO-S1-STAGING-{run_id}",
            "s3": f"DEMO-S3-STAGING-{run_id}",
        }
        journal["phase"] = "seeding_staging"
        journal["staging"] = {"names": staging_names, "objects": None}
        write_reset_journal(run_id, journal)
        seeded = seed_standard_demo(
            engine,
            actors,
            run_id,
            file_root,
            s2_subject=staging_names["s2"],
            s1_document_number=staging_names["s1"],
            s3_count_number=staging_names["s3"],
        )
        staged_verified = verify_demo_objects(
            engine,
            seeded,
            file_root,
            expected_names=staging_names,
        )
        journal["phase"] = "staging_verified"
        journal["staging"]["objects"] = seeded
        journal["staging"]["verification"] = staged_verified
        write_reset_journal(run_id, journal)

        current_plan = build_plan(lock_connection, file_root)
        if calculate_plan_hash(current_plan) != plan_hash:
            raise ResetBlocked("The standard Demo namespace changed while staging data was generated.")
        lock_connection.rollback()

        transaction = lock_connection.begin()
        try:
            switch_plan = build_plan(lock_connection, file_root)
            if calculate_plan_hash(switch_plan) != plan_hash:
                raise ResetBlocked("The standard Demo namespace changed before the atomic switch.")
            moved = quarantine_files(switch_plan, run_id, file_root)
            journal["phase"] = "files_quarantined"
            journal["quarantined_files"] = [
                {"source": str(source), "quarantine": str(target)}
                for source, target in moved
            ]
            write_reset_journal(run_id, journal)
            append_reset_audit(
                lock_connection,
                actor=actors["admin"],
                action="steven.demo_reset.started",
                run_id=run_id,
                plan_hash=plan_hash,
                outcome="success",
                details={**switch_plan, "stage": "atomic_switch"},
            )
            delete_standard_namespace(lock_connection, switch_plan)
            switch_staging_namespace(lock_connection, seeded)
            verified = verify_switched_demo(lock_connection, seeded, file_root)
            append_reset_audit(
                lock_connection,
                actor=actors["admin"],
                action="steven.demo_reset.completed",
                run_id=run_id,
                plan_hash=plan_hash,
                outcome="success",
                details={
                    "row_counts": plan["row_counts"],
                    "files": plan["files"],
                    "stage": "verified_before_commit",
                },
            )
            journal["phase"] = "switch_verified_pending_commit"
            journal["standard_verification"] = verified
            write_reset_journal(run_id, journal)
            transaction.commit()
        except Exception:
            transaction.rollback()
            restore_quarantined_files(moved)
            moved = []
            journal["phase"] = "switch_failed_old_files_restored"
            write_reset_journal(run_id, journal)
            raise

        journal["phase"] = "completed"
        journal_warning = None
        try:
            write_reset_journal(run_id, journal)
        except OSError as error:
            journal_warning = type(error).__name__
        return {
            "run_id": run_id,
            "plan_hash": plan_hash,
            "journal": str(JOURNAL_ROOT / f"{run_id}.json"),
            "quarantined_file_count": len(moved),
            "seeded": seeded,
            "verified": verified,
            "journal_warning": journal_warning,
        }
    except Exception as error:
        if moved:
            restore_quarantined_files(moved)
        journal["phase"] = "failed_standard_namespace_preserved"
        journal["error_type"] = type(error).__name__
        write_reset_journal(run_id, journal)
        if actors.get("admin"):
            try:
                with engine.begin() as connection:
                    append_reset_audit(
                        connection,
                        actor=actors["admin"],
                        action="steven.demo_reset.failed",
                        run_id=run_id,
                        plan_hash=plan_hash,
                        outcome="failure",
                        details={
                            "row_counts": plan.get("row_counts", {}),
                            "files": plan.get("files", []),
                            "stage": journal["phase"],
                            "error_type": type(error).__name__,
                        },
                    )
            except Exception:
                pass
        raise
    finally:
        try:
            lock_connection.execute(
                text("SELECT pg_advisory_unlock(:key)"),
                {"key": ADVISORY_LOCK_KEY},
            )
        finally:
            lock_connection.close()


def emit(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str))


def main(argv: list[str] | None = None) -> int:
    try:
        arguments = parse_arguments(argv)
        validate_apply_arguments(arguments)
        database_url = database_url_from_environment()
        validate_runtime_environment(database_url)
        os.environ["DATABASE_URL"] = database_url
        file_root = validate_file_root()
        engine = create_engine(database_url, pool_pre_ping=True)
        with engine.connect() as connection:
            plan = build_plan(connection, file_root)
        validate_plan(plan, file_root)
        plan_hash = calculate_plan_hash(plan)
        if not arguments.apply:
            emit({"mode": "dry-run", "plan_hash": plan_hash, "summary": summarize_plan(plan)})
            return 0
        validate_expected_plan_hash(arguments.expected_plan_hash, plan_hash)
        result = apply_reset(engine, plan, plan_hash, file_root)
        emit({"mode": "applied", **result})
        return 0
    except ResetBlocked as error:
        print(f"RESET_BLOCKED: {error}", file=sys.stderr)
        return 2
    except Exception as error:
        print(f"RESET_FAILED: {type(error).__name__}: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

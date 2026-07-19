from __future__ import annotations

from contextlib import contextmanager
from datetime import date
from io import BytesIO
import json
import os
from pathlib import Path
import sys
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import make_url
from sqlalchemy.exc import IntegrityError


PROJECT_ROOT = Path(__file__).resolve().parents[1]
API_ROOT = PROJECT_ROOT / "apps" / "api"
EXPECTED_DATABASE = "puiying_steven_demo"
EXPECTED_REVISION = "20260717_0011"
EXPECTED_TABLES = {
    "steven_inventory_items",
    "steven_inventory_counts",
    "steven_inventory_count_lines",
    "steven_inventory_versions",
    "steven_inventory_candidate_links",
    "steven_inventory_import_batches",
    "steven_inventory_import_rows",
}
EXPECTED_CONSTRAINTS = {
    "ck_steven_inventory_items_book_quantity",
    "ck_steven_inventory_items_safety_stock",
    "ck_steven_inventory_items_target_stock",
    "ck_steven_inventory_items_status",
    "ck_steven_inventory_counts_status",
    "ck_steven_inventory_counts_next_export_version",
    "ck_steven_inventory_lines_counted_quantity",
    "ck_steven_inventory_lines_confirmed_order",
    "ck_steven_inventory_lines_manual_reason",
    "uq_steven_inventory_items_normalized_sku",
    "uq_steven_inventory_counts_number",
    "uq_steven_inventory_count_item",
    "uq_steven_inventory_versions_count_version",
    "uq_steven_inventory_versions_storage_key",
    "ck_steven_inventory_import_batches_status",
    "ck_steven_inventory_import_batches_counts",
    "ck_steven_inventory_import_batches_count_total",
    "ck_steven_inventory_import_rows_number",
    "ck_steven_inventory_import_rows_status",
    "uq_steven_inventory_import_rows_batch_row",
}
EXPECTED_INDEXES = {
    "ix_steven_inventory_items_status_location",
    "ix_steven_inventory_counts_status_updated",
    "ix_steven_inventory_count_lines_count",
    "ix_steven_inventory_count_lines_low_stock",
    "ix_steven_inventory_versions_count_status",
    "ix_steven_inventory_candidate_links_count",
    "ix_steven_inventory_import_batches_status_created",
    "ix_steven_inventory_import_batches_request_id",
    "ix_steven_inventory_import_rows_batch_status",
    "ix_steven_inventory_import_rows_normalized_sku",
}

sys.path.insert(0, str(API_ROOT))

from app.core.api_response import ApiError  # noqa: E402
from app.core.audit_context import reset_request_id, set_request_id  # noqa: E402
from app.core.config import Settings  # noqa: E402
from app.main import LazyPostgresInventoryApplication, create_app  # noqa: E402
from app.modules.steven.inventory_excel import InventoryExcelRenderer  # noqa: E402
from app.modules.steven.inventory_schemas import (  # noqa: E402
    InventoryCountCreateRequest,
    InventoryCountLineInput,
    InventoryItemCreateRequest,
)
from app.modules.steven.inventory_service import normalize_sku  # noqa: E402

from seed_s3_demo import DEMO_ITEMS  # noqa: E402


def emit(step: str, status: str, **details: object) -> None:
    print(json.dumps({"step": step, "status": status, **details}, ensure_ascii=False, default=str))


@contextmanager
def request_id(value: str):
    token = set_request_id(value)
    try:
        yield
    finally:
        reset_request_id(token)


def actor_for_role(engine, role_code: str) -> str:
    with engine.connect() as connection:
        actor = connection.execute(
            text(
                """
                SELECT u.id
                  FROM users u
                  JOIN user_roles ur ON ur.user_id=u.id
                  JOIN roles r ON r.id=ur.role_id
                 WHERE u.status='active'
                   AND r.status='active'
                   AND r.code=:role_code
                 ORDER BY u.created_at,u.id
                 LIMIT 1
                """
            ),
            {"role_code": role_code},
        ).scalar_one_or_none()
    if actor is None:
        raise RuntimeError(f"active_{role_code}_actor_missing")
    return actor


def schema_smoke(engine) -> None:
    schema = inspect(engine)
    missing_tables = sorted(EXPECTED_TABLES - set(schema.get_table_names()))
    with engine.connect() as connection:
        revision = connection.execute(text("SELECT version_num FROM alembic_version")).scalar_one()
        constraints = set(
            connection.execute(
                text(
                    """
                    SELECT conname
                      FROM pg_constraint
                     WHERE conrelid IN (
                        'steven_inventory_items'::regclass,
                        'steven_inventory_counts'::regclass,
                        'steven_inventory_count_lines'::regclass,
                        'steven_inventory_versions'::regclass,
                        'steven_inventory_import_batches'::regclass,
                        'steven_inventory_import_rows'::regclass
                     )
                    """
                )
            ).scalars()
        )
        indexes = set(
            connection.execute(
                text(
                    """
                    SELECT indexname
                      FROM pg_indexes
                     WHERE schemaname='public'
                       AND tablename LIKE 'steven_inventory_%'
                    """
                )
            ).scalars()
        )
        permissions = set(
            connection.execute(
                text("SELECT code FROM permissions WHERE code LIKE 'steven:inventory:%'")
            ).scalars()
        )
    missing_constraints = sorted(EXPECTED_CONSTRAINTS - constraints)
    missing_indexes = sorted(EXPECTED_INDEXES - indexes)
    if (
        revision != EXPECTED_REVISION
        or missing_tables
        or missing_constraints
        or missing_indexes
        or len(permissions) != 7
    ):
        raise RuntimeError(
            f"schema_contract_failed revision={revision} tables={missing_tables} "
            f"constraints={missing_constraints} indexes={missing_indexes} "
            f"permissions={sorted(permissions)}"
        )
    emit(
        "schema",
        "passed",
        revision=revision,
        tables=sorted(EXPECTED_TABLES),
        constraint_count=len(EXPECTED_CONSTRAINTS),
        index_count=len(EXPECTED_INDEXES),
        permission_count=len(permissions),
    )


def ensure_demo_items(inventory, actor: str):
    existing = {normalize_sku(item.sku)[1]: item for item in inventory.list_items()}
    items = []
    for values in DEMO_ITEMS:
        normalized = normalize_sku(values["sku"])[1]
        item = existing.get(normalized)
        if item is None:
            item = inventory.create_item(InventoryItemCreateRequest(**values), actor)
            existing[normalized] = item
        items.append(item)
    return items


def inventory_xlsx(rows) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "sku",
            "item_name",
            "category",
            "location",
            "book_quantity",
            "safety_stock",
            "target_stock",
        ]
    )
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def assert_import_workflow(inventory, engine, steven_actor: str, approver_actor: str, suffix: str):
    with engine.connect() as connection:
        before_count = connection.execute(
            text("SELECT count(*) FROM steven_inventory_items")
        ).scalar_one()
    invalid_batch = inventory.preflight_import(
        f"DEMO-S3-invalid-{suffix}.xlsx",
        inventory_xlsx(
            [
                ["DEMO-S3-PAPER-A4", "已存在品项（脱敏）", "演示类别", "DEMO-STORE-C", 1, 1, 2],
                [f"DEMO-S3-DUP-{suffix}", "重复一（脱敏）", "演示类别", "DEMO-STORE-C", 1, 1, 2],
                [f" demo-s3-dup-{suffix} ", "重复二（脱敏）", "演示类别", "DEMO-STORE-C", 1, 1, 2],
                [f"DEMO-S3-NEG-{suffix}", "负数（脱敏）", "演示类别", "DEMO-STORE-C", -1, 1, 2],
                [f"DEMO-S3-DEC-{suffix}", "小数（脱敏）", "演示类别", "DEMO-STORE-C", 1.5, 1, 2],
                [f"DEMO-S3-TARGET-{suffix}", "目标异常（脱敏）", "演示类别", "DEMO-STORE-C", 1, 5, 4],
            ]
        ),
        steven_actor,
    )
    if invalid_batch.invalid_count < 5:
        raise RuntimeError("inventory_import_invalid_rows_not_reported")
    try:
        inventory.confirm_import(invalid_batch.id, steven_actor)
    except ApiError as error:
        if error.code != "import_preflight_blocked":
            raise
    else:
        raise RuntimeError("inventory_import_invalid_batch_not_blocked")
    with engine.connect() as connection:
        after_invalid_count = connection.execute(
            text("SELECT count(*) FROM steven_inventory_items")
        ).scalar_one()
    if after_invalid_count != before_count:
        raise RuntimeError("inventory_import_invalid_batch_partially_written")

    request = f"s3-import-{suffix}"
    with request_id(request):
        valid_batch = inventory.preflight_import(
            f"DEMO-S3-valid-{suffix}.xlsx",
            inventory_xlsx(
                [
                    [
                        f"DEMO-S3-IMPORT-PAPER-{suffix}",
                        "批量导入纸品（脱敏）",
                        "批量导入演示",
                        "DEMO-STORE-C",
                        4,
                        5,
                        12,
                    ],
                    [
                        f"DEMO-S3-IMPORT-PEN-{suffix}",
                        "批量导入笔品（脱敏）",
                        "批量导入演示",
                        "DEMO-STORE-C",
                        15,
                        5,
                        15,
                    ],
                ]
            ),
            steven_actor,
        )
        if valid_batch.valid_count != 2 or valid_batch.invalid_count != 0 or valid_batch.issues:
            raise RuntimeError("inventory_import_valid_preflight_failed")
        confirmed = inventory.confirm_import(valid_batch.id, steven_actor)
        if confirmed.status != "confirmed" or any(row.status != "confirmed" for row in confirmed.rows):
            raise RuntimeError("inventory_import_confirmation_failed")
        imported_items = [inventory.get_item(row.imported_item_id) for row in confirmed.rows]
        imported_count = inventory.create_count(
            InventoryCountCreateRequest(
                count_number=f"DEMO-S3-IMPORT-COUNT-{suffix}",
                count_date=date(2026, 7, 17),
                is_demo=True,
                lines=[
                    InventoryCountLineInput(
                        inventory_item_id=imported_items[0].id,
                        counted_quantity=2,
                        confirmed_order_quantity=10,
                    ),
                    InventoryCountLineInput(
                        inventory_item_id=imported_items[1].id,
                        counted_quantity=15,
                        confirmed_order_quantity=0,
                    ),
                ],
            ),
            steven_actor,
        )
        inventory.submit(imported_count.id, steven_actor)
        inventory.approve(imported_count.id, approver_actor, "已人工复核批量导入盘点")
        import_exports = [inventory.export(imported_count.id, steven_actor) for _ in range(3)]
    import_versions = inventory.list_versions(imported_count.id)
    if [version.version_number for version in import_versions] != [1, 2, 3]:
        raise RuntimeError("inventory_import_workflow_export_versions_failed")
    for version in import_versions:
        InventoryExcelRenderer.verify(inventory.version_file(imported_count.id, version.version_number))
    with engine.connect() as connection:
        persisted = connection.execute(
            text(
                """
                SELECT b.status,count(r.id),count(r.imported_item_id)
                  FROM steven_inventory_import_batches b
                  JOIN steven_inventory_import_rows r ON r.batch_id=b.id
                 WHERE b.id=:batch_id AND b.request_id=:request_id
                 GROUP BY b.status
                """
            ),
            {"batch_id": valid_batch.id, "request_id": request},
        ).one_or_none()
    if persisted is None or tuple(persisted) != ("confirmed", 2, 2):
        raise RuntimeError(f"inventory_import_persistence_failed persisted={persisted}")
    emit(
        "inventory_import",
        "passed",
        invalid_rows=invalid_batch.invalid_count,
        valid_rows=valid_batch.valid_count,
        confirmed_rows=len(confirmed.rows),
        count_status="approved",
        version_numbers=[version.version_number for version in import_versions],
        request_id_linked=True,
    )
    return {
        "batch_id": valid_batch.id,
        "imported_item_ids": [item.id for item in imported_items],
        "count_id": imported_count.id,
        "version_numbers": [version.version_number for version in import_versions],
        "export_result_count": len(import_exports),
    }


def assert_item_validation(inventory, actor: str, suffix: str) -> None:
    invalid_values = (
        {"book_quantity": -1},
        {"safety_stock": -1},
        {"target_stock": -1},
        {"book_quantity": 1.5},
        {"book_quantity": "12"},
    )
    for index, override in enumerate(invalid_values, start=1):
        values = {**DEMO_ITEMS[0], "sku": f"DEMO-S3-INVALID-{suffix}-{index}", **override}
        try:
            InventoryItemCreateRequest(**values)
        except ValidationError:
            continue
        raise RuntimeError(f"invalid_inventory_quantity_not_blocked_{index}")

    try:
        inventory.create_item(
            InventoryItemCreateRequest(
                **{
                    **DEMO_ITEMS[0],
                    "sku": f"DEMO-S3-TARGET-{suffix}",
                    "safety_stock": 10,
                    "target_stock": 9,
                }
            ),
            actor,
        )
    except ApiError as error:
        if error.code != "target_below_safety_stock":
            raise
    else:
        raise RuntimeError("target_below_safety_stock_not_blocked")

    try:
        inventory.create_item(
            InventoryItemCreateRequest(
                **{
                    **DEMO_ITEMS[0],
                    "sku": " ｄｅｍｏ－ｓ３－ｐａｐｅｒ－ａ４ ",
                }
            ),
            actor,
        )
    except ApiError as error:
        if error.code != "duplicate_sku":
            raise
    else:
        raise RuntimeError("normalized_duplicate_sku_not_blocked")
    emit(
        "item_rules",
        "passed",
        non_negative_strict_integer=True,
        target_not_below_safety=True,
        nfkc_casefold_unique=True,
    )


def count_payload(number: str, items, *, lines=None) -> InventoryCountCreateRequest:
    if lines is None:
        lines = [
            InventoryCountLineInput(
                inventory_item_id=items[0].id,
                counted_quantity=8,
                confirmed_order_quantity=22,
                remark="低库存与负差异脱敏案例",
            ),
            InventoryCountLineInput(
                inventory_item_id=items[1].id,
                counted_quantity=25,
                confirmed_order_quantity=0,
                remark="高于账面且建议订货量为零",
            ),
            InventoryCountLineInput(
                inventory_item_id=items[2].id,
                counted_quantity=7,
                confirmed_order_quantity=20,
                manual_reason="人工确认演示：保留两件缓冲量",
            ),
            InventoryCountLineInput(
                inventory_item_id=items[3].id,
                counted_quantity=18,
                confirmed_order_quantity=0,
            ),
            InventoryCountLineInput(
                inventory_item_id=items[4].id,
                counted_quantity=35,
                confirmed_order_quantity=5,
            ),
        ]
    return InventoryCountCreateRequest(
        count_number=number,
        count_date=date(2026, 7, 17),
        lines=lines,
        is_demo=True,
    )


def assert_count_rules(inventory, engine, actor: str, items, suffix: str) -> None:
    try:
        inventory.create_count(
            count_payload(
                f"DEMO-S3-MANUAL-{suffix}",
                items,
                lines=[
                    InventoryCountLineInput(
                        inventory_item_id=items[0].id,
                        counted_quantity=8,
                        confirmed_order_quantity=21,
                    )
                ],
            ),
            actor,
        )
    except ApiError as error:
        if error.code != "manual_reason_required":
            raise
    else:
        raise RuntimeError("manual_reason_not_blocked")

    duplicate_number = f"DEMO-S3-DUP-LINE-{suffix}"
    with engine.connect() as connection:
        before = connection.execute(
            text("SELECT count(*) FROM steven_inventory_counts WHERE count_number=:number"),
            {"number": duplicate_number},
        ).scalar_one()
    try:
        inventory.create_count(
            count_payload(
                duplicate_number,
                items,
                lines=[
                    InventoryCountLineInput(
                        inventory_item_id=items[0].id,
                        counted_quantity=8,
                    ),
                    InventoryCountLineInput(
                        inventory_item_id=items[0].id,
                        counted_quantity=9,
                    ),
                ],
            ),
            actor,
        )
    except ApiError as error:
        if error.code != "duplicate_count_item":
            raise
    else:
        raise RuntimeError("duplicate_count_item_not_blocked")
    with engine.connect() as connection:
        after = connection.execute(
            text("SELECT count(*) FROM steven_inventory_counts WHERE count_number=:number"),
            {"number": duplicate_number},
        ).scalar_one()
    if before != 0 or after != 0:
        raise RuntimeError("duplicate_line_transaction_partially_persisted")

    count = inventory.create_count(count_payload(f"DEMO-S3-CONSTRAINT-{suffix}", items), actor)
    first_line = count.lines[0]
    try:
        with engine.begin() as connection:
            connection.execute(
                text(
                    """
                    INSERT INTO steven_inventory_count_lines
                        (id,inventory_count_id,inventory_item_id,sku_snapshot,item_name_snapshot,location_snapshot,
                         book_quantity_snapshot,safety_stock_snapshot,target_stock_snapshot,counted_quantity,
                         difference_quantity,is_low_stock,suggested_order_quantity,confirmed_order_quantity,
                         manual_reason,remark,updated_by,created_at,updated_at)
                    SELECT
                        :id,inventory_count_id,inventory_item_id,sku_snapshot,item_name_snapshot,location_snapshot,
                        book_quantity_snapshot,safety_stock_snapshot,target_stock_snapshot,counted_quantity,
                        difference_quantity,is_low_stock,suggested_order_quantity,confirmed_order_quantity,
                        manual_reason,remark,updated_by,now(),now()
                      FROM steven_inventory_count_lines
                     WHERE id=:source_id
                    """
                ),
                {"id": str(uuid4()), "source_id": first_line.id},
            )
    except IntegrityError as error:
        constraint = getattr(getattr(error.orig, "diag", None), "constraint_name", None)
        if constraint != "uq_steven_inventory_count_item":
            raise
    else:
        raise RuntimeError("database_composite_unique_not_enforced")
    emit(
        "count_rules",
        "passed",
        manual_reason_required=True,
        duplicate_line_atomic_rollback=True,
        database_composite_unique=True,
    )


def workflow_smoke(settings: Settings, engine) -> dict[str, object]:
    app = create_app(settings)
    if not isinstance(app.state.inventory_application, LazyPostgresInventoryApplication):
        raise RuntimeError("inventory_repository_not_postgresql")
    inventory = app.state.inventory_application
    steven_actor = actor_for_role(engine, "operator")
    approver_actor = actor_for_role(engine, "approver")
    suffix = uuid4().hex[:10]
    items = ensure_demo_items(inventory, steven_actor)
    locations = {item.location for item in items}
    if len(items) < 20 or len(locations) < 3 or any(not value.startswith("DEMO-STORE-") for value in locations):
        raise RuntimeError(
            f"s3_demo_inventory_scale_failed item_count={len(items)} locations={sorted(locations)}"
        )

    assert_item_validation(inventory, steven_actor, suffix)
    assert_count_rules(inventory, engine, steven_actor, items, suffix)
    import_result = assert_import_workflow(
        inventory,
        engine,
        steven_actor,
        approver_actor,
        suffix,
    )

    draft = inventory.create_count(count_payload(f"DEMO-S3-DRAFT-{suffix}", items), steven_actor)
    try:
        inventory.export(draft.id, steven_actor)
    except ApiError as error:
        if error.code != "formal_export_forbidden":
            raise
    else:
        raise RuntimeError("draft_export_not_blocked")

    returned = inventory.create_count(count_payload(f"DEMO-S3-RETURN-{suffix}", items), steven_actor)
    inventory.submit(returned.id, steven_actor)
    try:
        inventory.return_for_revision(returned.id, approver_actor, " ")
    except ApiError as error:
        if error.code != "return_opinion_required":
            raise
    else:
        raise RuntimeError("empty_return_opinion_not_blocked")
    inventory.return_for_revision(returned.id, approver_actor, "请复核脱敏库存差异")
    try:
        inventory.export(returned.id, steven_actor)
    except ApiError as error:
        if error.code != "formal_export_forbidden":
            raise
    else:
        raise RuntimeError("returned_export_not_blocked")

    request = f"s3-smoke-{suffix}"
    with request_id(request):
        approved = inventory.create_count(
            count_payload(f"DEMO-S3-APPROVED-{suffix}", items),
            steven_actor,
        )
        lines = approved.lines
        expected = {
            "DEMO-S3-PAPER-A4": (-4, True, 22, 22),
            "DEMO-S3-PEN-BLUE": (5, False, 0, 0),
            "DEMO-S3-MARKER-WHITEBOARD": (-8, True, 18, 20),
            "DEMO-S3-FOLDER-A4": (0, False, 0, 0),
            "DEMO-S3-STAPLES-246": (5, False, 5, 5),
        }
        actual = {
            line.sku: (
                line.difference_quantity,
                line.is_low_stock,
                line.suggested_order_quantity,
                line.confirmed_order_quantity,
            )
            for line in lines
        }
        if actual != expected:
            raise RuntimeError(f"inventory_calculation_mismatch actual={actual}")
        inventory.submit(approved.id, steven_actor)
        try:
            inventory.approve(approved.id, steven_actor, "不得成功")
        except ApiError as error:
            if error.status_code != 403 or error.code != "self_approval_forbidden":
                raise
        else:
            raise RuntimeError("self_approval_not_blocked")
        inventory.approve(approved.id, approver_actor, "独立审批人已核对脱敏盘点")
        exports = [inventory.export(approved.id, steven_actor) for _ in range(3)]

    versions = inventory.list_versions(approved.id)
    if [item.version_number for item in versions] != [1, 2, 3]:
        raise RuntimeError("export_versions_not_continuous")
    if len({item.storage_key for item in versions}) != 3:
        raise RuntimeError("export_storage_key_reused")
    if any(item.status != "ready" for item in versions):
        raise RuntimeError("export_version_not_ready")
    file_root = Path(settings.file_storage_root).resolve()
    for version in versions:
        path = inventory.version_file(approved.id, version.version_number)
        if file_root not in path.resolve().parents:
            raise RuntimeError("export_path_outside_file_root")
        InventoryExcelRenderer.verify(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if workbook["Inventory Count"]["B5"].value != version.version_number:
                raise RuntimeError("workbook_version_mismatch")
        finally:
            workbook.close()
        if path.stat().st_size != version.size_bytes or not version.sha256 or len(version.sha256) != 64:
            raise RuntimeError("export_metadata_mismatch")

    audits = inventory.list_audit_events(approved.id)
    required_actions = {
        "inventory.count.create",
        "inventory.count.submit",
        "inventory.count.approve",
        "inventory.export_reserved",
        "inventory.export",
    }
    if not required_actions.issubset({item["action"] for item in audits}):
        raise RuntimeError("persistent_inventory_audit_incomplete")
    if not any(item["request_id"] == request for item in audits):
        raise RuntimeError("inventory_audit_request_id_missing")

    restarted = create_app(settings)
    if not isinstance(restarted.state.inventory_application, LazyPostgresInventoryApplication):
        raise RuntimeError("restart_inventory_repository_not_postgresql")
    recovered = restarted.state.inventory_application.get_count(approved.id)
    recovered_versions = restarted.state.inventory_application.list_versions(approved.id)
    recovered_audits = restarted.state.inventory_application.list_audit_events(approved.id)
    if (
        recovered.status != "approved"
        or len(recovered.lines) != 5
        or len(recovered_versions) != 3
        or len(recovered_audits) < len(required_actions)
    ):
        raise RuntimeError("restart_recovery_failed")
    if not restarted.state.quote_application.list_quotes():
        raise RuntimeError("s2_baseline_missing_after_s3")
    if not restarted.state.tender_application.list_tenders():
        raise RuntimeError("s1_baseline_missing_after_s3")

    with engine.connect() as connection:
        session_count = connection.execute(text("SELECT count(*) FROM auth_sessions")).scalar_one()
        platform_audit_count = connection.execute(text("SELECT count(*) FROM platform_audit_events")).scalar_one()
        file_rows = connection.execute(
            text(
                """
                SELECT count(*)
                  FROM files
                 WHERE module='steven'
                   AND document_type='inventory_count'
                   AND purpose='approved_inventory_export'
                   AND request_id=:request_id
                """
            ),
            {"request_id": request},
        ).scalar_one()
        baseline_counts = connection.execute(
            text(
                """
                SELECT
                    (SELECT count(*) FROM steven_quote_jobs),
                    (SELECT count(*) FROM steven_tender_jobs),
                    (SELECT count(*) FROM steven_inventory_items),
                    (SELECT count(*) FROM steven_inventory_counts)
                """
            )
        ).one()
    if session_count < 1 or platform_audit_count < 1 or file_rows != 3:
        raise RuntimeError("session_audit_or_file_metadata_persistence_failed")

    return {
        "inventory_count_id": approved.id,
        "inventory_status": recovered.status,
        "item_count": len(items),
        "location_count": len(locations),
        "import": import_result,
        "line_count": len(recovered.lines),
        "calculation_results": actual,
        "version_numbers": [item.version_number for item in recovered_versions],
        "excel_reopenable": True,
        "request_id": request,
        "audit_count": len(recovered_audits),
        "session_count": session_count,
        "baseline_counts": list(baseline_counts),
        "draft_count_id": draft.id,
        "returned_count_id": returned.id,
        "export_result_count": len(exports),
    }


def main() -> int:
    settings = Settings.from_env()
    settings.validate()
    database_name = make_url(settings.database_url).database if settings.database_url else None
    if settings.app_env != "development" or settings.auth_mode != "session":
        raise RuntimeError("s3_smoke_requires_development_session")
    if settings.demo_seed_enabled:
        raise RuntimeError("s3_smoke_forbids_runtime_demo_seed")
    if database_name != EXPECTED_DATABASE:
        raise RuntimeError("s3_smoke_wrong_database")
    if settings.ocr_enabled or settings.ai_structuring_enabled:
        raise RuntimeError("s3_smoke_requires_offline_ai_ocr")

    engine = create_engine(settings.database_url, pool_pre_ping=True, future=True)
    schema_smoke(engine)
    result = workflow_smoke(settings, engine)
    emit("s3_postgres_closure", "passed", **result)
    emit(
        "verification",
        "passed",
        scope="local_redacted_s3_demo",
        real_postgres=True,
        in_memory_fallback=False,
        external_services_called=False,
        production_ready=False,
        controlled_trial_ready=False,
        d2_authorized=False,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        emit("verification", "failed", error_type=type(error).__name__, message=str(error))
        raise SystemExit(1)
